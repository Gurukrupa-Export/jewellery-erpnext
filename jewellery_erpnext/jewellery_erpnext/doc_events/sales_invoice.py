from datetime import datetime, timedelta

import frappe
from erpnext.setup.utils import get_exchange_rate
from frappe import _
from frappe.query_builder.custom import ConstantColumn
from frappe.utils import flt, get_last_day , getdate,strip

from jewellery_erpnext.jewellery_erpnext.doc_events.bom_utils import _calculate_diamond_amount
from jewellery_erpnext.jewellery_erpnext.doc_events.quotation import update_totals


def _so_gold_rate_changed(si_gold_rate, sales_order):
	if not sales_order:
		return True
	so_gold_rate = frappe.db.get_value("Sales Order", sales_order, "gold_rate")
	if not so_gold_rate:
		return True
	return abs(flt(si_gold_rate) - flt(so_gold_rate)) > 0.001


def before_validate(self, method):
	if self.is_return:
		return
	if self.sales_type != 'Certification':
		if self.gold_rate:
			self.gold_rate_with_gst=round(self.gold_rate * 1.03,3)
		if self.item_same_as_above:
			self.invoice_item = []
			for row in self.items:
				duplicate_row = {}
				for key in row.__dict__:
					duplicate_row[key] = row.get(key)
				duplicate_row["name"] = None

				if duplicate_row:
					self.append("invoice_item", duplicate_row)
		customer_group=frappe.db.get_value('Customer',self.customer,'customer_group')
		prec= frappe.db.get_value('Customer',self.customer,'custom_precision_variable')
		if not (self.company == "KG GK Jewellers Private Limited" or customer_group == "Internal"):
			for row_s in self.items:
					if row_s.bom:
						bom_doc = frappe.get_doc("BOM", row_s.bom)
						gold_gst_rate=frappe.db.get_single_value("Jewellery Settings", "gold_gst_rate")
						for row in bom_doc.metal_detail:
							update_making_charges(row_s, bom_doc, row, self.gold_rate_with_gst) 
							customer_metal_purity = frappe.db.sql(f"""select metal_purity from `tabMetal Criteria` where parent = '{self.customer}' and metal_type = '{row.metal_type}' and metal_touch = '{row.metal_touch}'""",as_dict=True)[0]['metal_purity']
							row.customer_metal_purity=customer_metal_purity
							rate = (float(row.customer_metal_purity) * self.gold_rate_with_gst) / (100 + int(gold_gst_rate))
							row.rate = round(rate,2)
							row.amount=round(row.rate*row.quantity,2 )
							row.wastage_amount = row.amount * row.wastage_rate
						bom_doc.total_metal_amount= sum(row.amount for row in bom_doc.metal_detail)
						for row in bom_doc.finding_detail:
							update_making_charges(row_s, bom_doc, row, self.gold_rate_with_gst) 
							customer_metal_purity = frappe.db.sql(f"""select metal_purity from `tabMetal Criteria` where parent = '{self.customer}' and metal_type = '{row.metal_type}' and metal_touch = '{row.metal_touch}'""",as_dict=True)[0]['metal_purity']
							row.customer_metal_purity=customer_metal_purity
							rate = (float(row.customer_metal_purity) * self.gold_rate_with_gst) / (100 + int(gold_gst_rate))
							row.rate = round(rate,2)
							row.amount=round(row.rate*row.quantity,2 )
							row.wastage_amount = row.amount * row.wastage_rate
						bom_doc.total_finding_amount= sum(row.amount for row in bom_doc.finding_detail)
						bom_doc.diamond_bom_amount = bom_doc.total_diamond_amount
						total_bom_amount = round(
							bom_doc.total_bom_amount
							+ bom_doc.making_charge
							+ flt(bom_doc.certification_amount)
							+ flt(bom_doc.custom_duty_amount)
							+ flt(bom_doc.hallmarking_amount)
							+ flt(bom_doc.freight_amount)
							+ flt(bom_doc.sale_amount),
							prec,
						)
						row_s.rate = total_bom_amount
						row_s.amount = row_s.rate
						row_s.taxable_value=row_s.base_net_amount=row_s.base_net_rate = row_s.net_amount=row_s.net_rate=row_s.base_amount=row_s.base_rate=total_bom_amount
						bom_doc.save(ignore_permissions=True)
						row_s.wastage_amount = bom_doc.total_wastage_amount

		update_income_account(self)
		payment_terms_data = update_si_data(self )
		update_payment_terms(self, payment_terms_data)

def validate(self, method):
	
	if self.is_return:
		set_gst_details(self)
		# payment_terms_data = update_si_data(self )
		# update_payment_terms(self, payment_terms_data)
		self.calculate_taxes_and_totals()
		for row_s in self.items:
			if row_s.bom:
				bom_doc = frappe.get_doc("BOM", row_s.bom)
				row_s.custom_diamond_pcs=bom_doc.total_diamond_pcs
				row_s.custom_gemstone_pcs=bom_doc.total_gemstone_pcs
				row_s.custom_other_weight = bom_doc.total_other_weight
				row_s.custom_metal_weight=bom_doc.total_metal_weight
				row_s.custom_finding_weight=bom_doc.finding_weight
				row_s.custom_diamond_weight=bom_doc.total_diamond_weight_in_gms
				row_s.custom_gemstone_weight=bom_doc.total_gemstone_weight_in_gms
		self.custom_diamond_pcs = sum(flt(r.custom_diamond_pcs) for r in self.items)
		self.custom_gemstone_pcs = sum(flt(r.custom_gemstone_pcs) for r in self.items)
		self.custom_other_weight = sum(flt(r.custom_other_weight) for r in self.items)
		self.custom_metal_weight = sum(flt(r.custom_metal_weight) for r in self.items)
		self.custom_finding_weight = sum(flt(r.custom_finding_weight) for r in self.items)
		self.custom_diamond_weight = sum(flt(r.custom_diamond_weight) for r in self.items)
		self.custom_gemstone_weight = sum(flt(r.custom_gemstone_weight) for r in self.items)
		payment_terms_data = update_si_data(self )
		update_payment_terms(self, payment_terms_data)
		return
	prec = frappe.db.get_value(
		"Customer",
		self.customer,
		"custom_precision_variable"
	)

	update_income_account(self)
	payment_terms_data = update_si_data(self )
	update_payment_terms(self, payment_terms_data)
	customer_group=frappe.db.get_value('Customer',self.customer,'customer_group')
	for row_s in self.items:
		if row_s.bom:
			bom_doc = frappe.get_doc("BOM", row_s.bom)
			row_s.custom_diamond_pcs=bom_doc.total_diamond_pcs
			row_s.custom_gemstone_pcs=bom_doc.total_gemstone_pcs
			row_s.custom_other_weight = bom_doc.total_other_weight
			row_s.custom_metal_weight=bom_doc.total_metal_weight
			row_s.custom_finding_weight=bom_doc.finding_weight
			row_s.custom_diamond_weight=bom_doc.total_diamond_weight_in_gms
			row_s.custom_gemstone_weight=bom_doc.total_gemstone_weight_in_gms
	self.custom_diamond_pcs = sum(flt(r.custom_diamond_pcs) for r in self.items)
	self.custom_gemstone_pcs = sum(flt(r.custom_gemstone_pcs) for r in self.items)
	self.custom_other_weight = sum(flt(r.custom_other_weight) for r in self.items)
	self.custom_metal_weight = sum(flt(r.custom_metal_weight) for r in self.items)
	self.custom_finding_weight = sum(flt(r.custom_finding_weight) for r in self.items)
	self.custom_diamond_weight = sum(flt(r.custom_diamond_weight) for r in self.items)
	self.custom_gemstone_weight = sum(flt(r.custom_gemstone_weight) for r in self.items)
	if not (self.company == "KG GK Jewellers Private Limited" or customer_group == "Internal"):
		self.total = 0
		for row in self.items:
			if row.bom:
				bom_doc = frappe.get_doc("BOM", row.bom)
				for m in bom_doc.metal_detail:
					# if not m.is_customer_item:
						update_making_charges(row, bom_doc, m, self.gold_rate_with_gst)
				for m in bom_doc.finding_detail:
					# if not m.is_customer_item:
						update_making_charges(row, bom_doc, m, self.gold_rate_with_gst)
				bom_doc.diamond_bom_amount = bom_doc.total_diamond_amount
				total_bom_amount = round(
					bom_doc.total_bom_amount
					+ bom_doc.making_charge
					+ flt(bom_doc.certification_amount)
					+ flt(bom_doc.custom_duty_amount)
					+ flt(bom_doc.hallmarking_amount)
					+ flt(bom_doc.freight_amount)
					+ flt(bom_doc.sale_amount),
					prec,
				)
				bom_doc.custom_gk_sell_gold_bom_amount=bom_doc.gold_bom_amount
				bom_doc.custom_gk_sell_total_bom_amount=bom_doc.total_bom_amount
				bom_doc.custom_gk_sell_making_charge= bom_doc.making_charge
				bom_doc.custom_gk_sell_other_bom_amount= bom_doc.other_bom_amount
				bom_doc.custom_gk_sell_finding_bom_amount= bom_doc.finding_bom_amount
				bom_doc.custom_gk_sell_gemstone_bom_amount= bom_doc.gemstone_bom_amount
				bom_doc.custom_gk_sell_diamond_bom_amount= bom_doc.diamond_bom_amount
				row.rate = total_bom_amount
				row.amount = row.rate
				row.taxable_value=row.base_net_amount=row.base_net_rate = row.net_amount=row.net_rate=row.base_amount=row.base_rate=total_bom_amount
			self.total+=row.amount
	
	set_gst_details(self)
	self.calculate_taxes_and_totals()

def get_allowed_item_types(customer, sales_type):
	"""Return the set of E Invoice Item names whose Customer Payment Terms entry
	has a matching sales_type row. Computed once per invoice save."""
	allowed_item_types = set()
	customer_payment_term_name = frappe.db.get_value(
		"Customer Payment Terms", {"customer": customer}
	)
	if not customer_payment_term_name:
		return allowed_item_types

	customer_payment_term_doc = frappe.get_doc("Customer Payment Terms", customer_payment_term_name)
	for row in customer_payment_term_doc.customer_payment_details:
		item_type = row.item_type
		e_invoice_item = frappe.get_doc("E Invoice Item", item_type)
		matched_sales_type_row = None
		for st_row in e_invoice_item.sales_type:
			if st_row.sales_type == sales_type:
				matched_sales_type_row = st_row
				break
		if matched_sales_type_row:
			allowed_item_types.add(item_type)

	return allowed_item_types
	
def on_submit(self,method):
	if self.company == 'Sadguru Diamond':
		return
	if self.is_return:
		total_making_charge = 0
		for row in self.items:
			if row.item_code=='Subcontracting Charges':
				return
			# frappe.msgprint(f"Row: {row.idx}, Item: {row.item_code}, BOM: {row.bom}")
			bom_doc = frappe.get_doc("BOM", row.bom)
			total_making_charge += bom_doc.making_charge
		new_si = frappe.copy_doc(self)
		new_si.sales_invoice = self.name
		new_si.set("items", [])
		new_si.set("invoice_item", [])
		new_si.set("taxes", [])
			
		new_si.append("items", {
					"item_code": "Subcontracting Charges",
					"item_name": "Subcontracting Charges",
					"description": "Subcontracting Charges",
					"qty": -1,
					"uom": "Nos",
					"conversion_factor": 1,
					"rate": total_making_charge,
					"amount": total_making_charge,
					"sales_order":row.sales_order,
					# "gst_hsn_code": gst_hsn_code,
					# "item_tax_template": sc_template,
					"custom_is_subcontracting_charge_row": 1,
				})
		custom_item, hsn_code, uom = frappe.db.get_value(
			"E Invoice Item", {"is_for_labour": 1,"metal_purity":bom_doc.metal_touch}, ["name", "hsn_code", "uom"]
		)
		new_si.append("invoice_item", {
					"amount":total_making_charge,
					"base_amount":total_making_charge,
					"base_rate":total_making_charge,
					"rate":total_making_charge,
					"qty": -1,
					"item_name":custom_item,
					"conversion_factor":1,
					"item_code":custom_item,
					"hsn_code": hsn_code,
					"income_account": row.income_account,
					"uom":uom,})
		new_si.total = flt(sum(flt(d.amount) for d in new_si.items))

		set_gst_details(new_si)
		new_si.calculate_taxes_and_totals()
		new_si.insert(ignore_permissions=True)
		frappe.db.set_value("Serial No", row.get("serial_no"), {"status": "Active","warehouse":"Product Allocation FG - KGJPL"})
		return
	if self.sales_type=='Hybrid' :
		# _template = SALES_TYPE_ITEM_TAX_TEMPLATE["Outwork"].get(self.company)
		
		total_making_charge = 0
		for row in self.items:
			if row.item_code=='Subcontracting Charges':
				return
			# frappe.msgprint(f"Row: {row.idx}, Item: {row.item_code}, BOM: {row.bom}")
			bom_doc = frappe.get_doc("BOM", row.bom)
			total_making_charge += bom_doc.making_charge
		new_si = frappe.copy_doc(self)
		new_si.sales_invoice = self.name
		new_si.set("items", [])
		new_si.set("invoice_item", [])
		new_si.set("taxes", [])
			
		new_si.append("items", {
					"item_code": "Subcontracting Charges",
					"item_name": "Subcontracting Charges",
					"description": "Subcontracting Charges",
					"qty": 1,
					"uom": "Nos",
					"conversion_factor": 1,
					"rate": total_making_charge,
					"amount": total_making_charge,
					"sales_order":row.sales_order,
					# "gst_hsn_code": gst_hsn_code,
					# "item_tax_template": sc_template,
					"custom_is_subcontracting_charge_row": 1,
				})
		custom_item, hsn_code, uom = frappe.db.get_value(
			"E Invoice Item", {"is_for_labour": 1,"metal_purity":bom_doc.metal_touch}, ["name", "hsn_code", "uom"]
		)
		new_si.append("invoice_item", {
					"amount":total_making_charge,
					"base_amount":total_making_charge,
					"base_rate":total_making_charge,
					"rate":total_making_charge,
					"qty": 1,
					"item_name":custom_item,
					"conversion_factor":1,
					"item_code":custom_item,
					"hsn_code": hsn_code,
					"income_account": row.income_account,
					"uom":uom,})
		new_si.total = flt(sum(flt(d.amount) for d in new_si.items))

		set_gst_details(new_si)
		new_si.calculate_taxes_and_totals()
		new_si.insert(ignore_permissions=True)


		
	if self.sales_type not in  ['Certification','']:
		separate_hallmarking_invoice = frappe.db.get_value(
			"Customer", self.customer, "custom_separate_hallmarking_invoice"
		)
		certification_items = []
		certification_invoice = []
		for row in self.items:
			if not row.bom:
				continue
			bom_doc = frappe.get_doc("BOM", row.bom)
			if bom_doc.certification_amount and separate_hallmarking_invoice:
					custom_item, hsn_code, uom = frappe.db.get_value(
					"E Invoice Item", {"is_for_certification": 1}, ["name", "hsn_code", "uom"]
				)
					certification_items.append({
							"qty": 1,
							"hsn_code": hsn_code,
							"uom": uom,
							"bom":row.bom,
							"amount":bom_doc.certification_amount,
							"rate":bom_doc.certification_amount,
							"income_account": row.income_account,
							"cost_center": row.cost_center,
							"delivery_note":row.delivery_note,
							"item_code":row.item_code,
							"sales_order":row.sales_order
						})
					certification_invoice.append({
						"amount":bom_doc.certification_amount,
						"base_amount":bom_doc.certification_amount,
						"base_rate":bom_doc.certification_amount,
						"rate":bom_doc.certification_amount,
						"qty": 1,
						"item_name":custom_item,
						"conversion_factor":1,
						"item_code":custom_item,
						"hsn_code": hsn_code,
						"income_account": row.income_account,
						"uom":uom,
					})
					if certification_items:
						certification_si = frappe.new_doc("Sales Invoice")
						certification_si.customer = self.customer
						certification_si.tax_category=self.tax_category
						certification_si.sales_type="Certification"
						certification_si.company = self.company 
						certification_si.posting_date = self.posting_date
						for item in certification_items:
							certification_si.append("items", item)
						for invoice in certification_invoice:
							certification_si.append("invoice_item",invoice)
						certification_si.insert(ignore_permissions=True,ignore_mandatory=True)
						certification_si.save()

def set_gst_details(self):
    if self.sales_type not in ("Outright", "Outwork","Hybrid"):
        return

    customer_state = frappe.db.get_value("Address", self.customer_address, "gst_state_number")
    company_state  = frappe.db.get_value("Address", self.company_address,  "gst_state_number")

    if not customer_state or not company_state:
        return

    self.tax_category = "In-State" if customer_state == company_state else "Out-State"

    item_template_map = {
        "Outright": {
            "Gurukrupa Export Private Limited": "GST 3% - GEPL",
            "KG GK Jewellers Private Limited":  "GST 3% - KGJPL",
        },
		"Branch Sales": {
        "Gurukrupa Export Private Limited": "GST 3% - GEPL"
        },
        "Outwork": {
            "Gurukrupa Export Private Limited": "GST 5% - GEPL",
            "KG GK Jewellers Private Limited":  "GST 5% - KGJPL",
        },
		
        "Hybrid": {
            "Gurukrupa Export Private Limited": "GST 5% - GEPL",
            "KG GK Jewellers Private Limited":  "GST 5% - KGJPL",
         },
    }
    if self.sales_type == "Hybrid":
        has_real_items = any(row.item_code != "Subcontracting Charges" for row in self.items)
        template_key = "Outright" if has_real_items else "Outwork"
    else:
        template_key = self.sales_type
    item_tax_template = item_template_map.get(self.sales_type, {}).get(self.company)
    if not item_tax_template:
        return

    taxes_and_charges = frappe.db.get_value(
        "Sales Taxes and Charges Template",
        {
            "company":      self.company,
            "tax_category": self.tax_category,
            "disabled":     0,
        },
        "name"
    )

    if not taxes_and_charges:
        frappe.log_error(
            f"No Sales Taxes and Charges Template found for "
            f"Company: {self.company}, Tax Category: {self.tax_category}",
            "set_gst_details"
        )
        return

    self.taxes_and_charges = taxes_and_charges

    template_rates = frappe.get_all(
        "Item Tax Template Detail",
        filters={"parent": item_tax_template},
        fields=["tax_type", "tax_rate"]
    )

    cgst_rate = sgst_rate = igst_rate = 0.0
    for r in template_rates:
        tax_type = r.tax_type or ""
        if "Output" not in tax_type or "RCM" in tax_type:
            continue
        if "CGST" in tax_type:
            cgst_rate = flt(r.tax_rate)
        elif "SGST" in tax_type:
            sgst_rate = flt(r.tax_rate)
        elif "IGST" in tax_type:
            igst_rate = flt(r.tax_rate)

    account_rate_map = {}
    for r in template_rates:
        tax_type = r.tax_type or ""
        if "Output" not in tax_type or "RCM" in tax_type:
            continue
        account_rate_map[r.tax_type] = flt(r.tax_rate)

    self.taxes = []
    tax_rows = frappe.get_all(
        "Sales Taxes and Charges",
        filters={"parent": self.taxes_and_charges},
        fields=["charge_type", "account_head", "description", "rate", "cost_center"],
        order_by="idx asc"
    )
    for t in tax_rows:
        correct_rate = account_rate_map.get(t.account_head, t.rate)
        self.append("taxes", {
            "charge_type":  t.charge_type,
            "account_head": t.account_head,
            "description":  t.description,
            "rate":         correct_rate,
            "cost_center":  t.cost_center,
        })
    for item in self.items:
        if not item.item_code:
            continue

        item.item_tax_template = item_tax_template
        item.gst_treatment     = "Taxable"
        item.cgst_rate         = 0.0
        item.sgst_rate         = 0.0
        item.igst_rate         = 0.0
        item.cgst_amount       = 0.0
        item.sgst_amount       = 0.0
        item.igst_amount       = 0.0

        taxable_value=self.total
        if self.tax_category == "In-State":
            item.cgst_rate   = cgst_rate
            item.sgst_rate   = sgst_rate
            item.cgst_amount = flt(taxable_value * cgst_rate / 100, 2)
            item.sgst_amount = flt(taxable_value * sgst_rate / 100, 2)
        else:
            item.igst_rate   = igst_rate
            item.igst_amount = flt(taxable_value * igst_rate / 100, 2)

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from io import BytesIO

@frappe.whitelist()
def xl_preview_sales_invoice(docname):
    doc = frappe.get_doc("Sales Invoice", docname)
    rows_diamond = []

    # Excel columns (Diamond Rate pehle, Diamond Amount baadme shift kiya hai)
    columns = [
        "Index","Item Code","Serial No","Item Name","Diamond Quality","PCS","Diamond Weight","Average",
        "Total Cts","Grams","Total Diamond Rate","Diamond Amount","Gross Weight","Gemstone Weight",
        "Other Weight","Gold Rate","Net Weight","Gold Amount","Customer Purity","Chain Weight",
        "Chain Amount","Chain Purity","Per Gram MC","Chain MC","Chain Wastage %","Chain Wastage Amount",
        "Jewellery Per Gram MC","Jewellery MC","Gold Wastage %","Jewellery Wastage","Gemstone Pcs",
        "Gemstone Cts","Gemstone Amount","Cert Charge","Hallmark Charge","Total Amt"
    ]

    # --- Populate rows_diamond ---
    for item in doc.items:
        #  New logic for BOM selection
        bom_name = item.bom
        # if item.quotation_bom:
        #     bom_name = item.quotation_bom
        # elif hasattr(item, "bom") and item.bom:   # Check if BOM field exists in Sales Order Item
        #     bom_name = item.bom

        if not bom_name:
            continue  # Skip if no BOM found

        bom_doc = frappe.get_doc("BOM", bom_name)

        total_qty = sum([float(d.quantity or 0) for d in bom_doc.diamond_detail])
        grams = total_qty * 0.2

        gross_weight = float(bom_doc.gross_weight or 0)
        gross_weight = round(gross_weight, 2)

        gemstone_weight = float(bom_doc.total_gemstone_weight_in_gms or 0)
        other_weight = float(bom_doc.other_weight or 0)
        net_weight = float(bom_doc.metal_and_finding_weight or 0)

        gemstone_pcs_rows = [float(g.pcs or 0) for g in bom_doc.gemstone_detail] if bom_doc.gemstone_detail else []
        gemstone_cts_rows = [float(g.quantity or 0) for g in bom_doc.gemstone_detail] if bom_doc.gemstone_detail else []
        gemstone_amount_rows = [float(g.gemstone_rate_for_specified_quantity or 0) for g in bom_doc.gemstone_detail] if bom_doc.gemstone_detail else []

        chain_weight_val, chain_mc_val, chain_wastage_val = 0.0, 0.0, 0.0
        chain_weight, chain_amount, chain_mc, chain_wastage, chain_purity = 0, 0, 0, 0, 0
        per_gram_mc, chain_wastage_amount = 0, 0
        net_weight_from_findings = 0.0

        if bom_doc.finding_detail:
            for f in bom_doc.finding_detail:
                qty = float(f.quantity or 0)
                if f.finding_category and f.finding_category.lower() == "chains":
                    chain_weight_val += qty
                    chain_purity = float(f.customer_metal_purity or 0)
                    per_gram_mc = float(f.making_rate or 0)
                    chain_mc_val = float(f.making_amount or 0)
                    chain_wastage_val = float(f.wastage_rate or 0)
                else:
                    net_weight_from_findings += qty

        # --- FIXED CHAIN AMOUNT CALCULATION ---
        if chain_weight_val > 0:
            chain_weight = chain_weight_val
            quotation_gold_rate = float(doc.gold_rate or 0)
            chain_amount = (quotation_gold_rate * chain_purity / 100) * chain_weight
            chain_mc = chain_mc_val
            chain_wastage = chain_wastage_val
            chain_wastage_amount = (chain_amount * chain_wastage_val) if chain_wastage_val else 0

        net_weight_display = net_weight + net_weight_from_findings

        #  Net Weight se chain weight minus karna ---
        if chain_weight > 0:
            net_weight_display = net_weight_display - chain_weight

        if bom_doc.metal_detail:
            customer_metal_purity = float(bom_doc.metal_detail[0].customer_metal_purity or 0)
            gold_wastage = float(bom_doc.metal_detail[0].wastage_rate or 0)
            jewellery_per_gram_mc = float(bom_doc.metal_detail[0].making_rate or 0)
        else:
            customer_metal_purity, gold_wastage, jewellery_per_gram_mc = 0.0, 0, 0

        quotation_gold_rate = float(doc.gold_rate or 0)
        calculated_gold_rate = (quotation_gold_rate * customer_metal_purity) / 100
        calculated_gold_rate = float(f"{calculated_gold_rate:.2f}")   # Always 2 decimals

        cert_charge = float(bom_doc.certification_amount or 0)
        hallmark_charge = float(bom_doc.hallmarking_amount or 0)

        for i, diamond in enumerate(bom_doc.diamond_detail):
            pcs = float(diamond.pcs or 0)
            qty = float(diamond.quantity or 0)   #
            qty = float(f"{qty:.2f}")            #
            avg = (qty / pcs) if pcs else 0
            rate = float(diamond.total_diamond_rate or 0)
            diamond_amount = rate * qty

            gold_amount_val = calculated_gold_rate * net_weight_display if i == 0 else 0
            jewellery_wastage_val = gold_amount_val * (gold_wastage / 100) if i == 0 else 0

            gemstone_pcs_val = gemstone_pcs_rows[i] if i < len(gemstone_pcs_rows) else 0
            gemstone_cts_val = gemstone_cts_rows[i] if i < len(gemstone_cts_rows) else 0
            gemstone_amount_val = gemstone_amount_rows[i] if i < len(gemstone_amount_rows) else 0

            jewellery_mc_val = net_weight_display * jewellery_per_gram_mc if i == 0 else 0

            total_amt = (
                hallmark_charge + cert_charge + jewellery_mc_val +
                gemstone_amount_val + gold_amount_val +
                jewellery_wastage_val + diamond_amount
            )

            rows_diamond.append([
                item.idx if i == 0 else "",
                item.item_code if i == 0 else "",
                item.serial_no if i == 0 else "",
                item.item_name if i == 0 else "",
                # item.diamond_quality,
                pcs,
                f"{qty:.2f}",
                f"{avg:.3f}",   # 
                round(total_qty, 2) if (i == 0 and total_qty != 0) else "",
                round(grams, 2) if (i == 0 and grams != 0) else "",
                round(rate, 2),   # 
                round(diamond_amount, 2),  # 
                round(gross_weight, 2) if (i == 0 and gross_weight != 0) else "",
                round(gemstone_weight, 2) if (i == 0 and gemstone_weight != 0) else "",
                round(other_weight, 2) if (i == 0 and other_weight != 0) else "",
                f"{calculated_gold_rate:.2f}" if i == 0 else "",
                round(net_weight_display, 2) if i == 0 else "",
                f"{gold_amount_val:.2f}" if i == 0 else "",
                customer_metal_purity if i == 0 else "",
                round(chain_weight, 2) if i == 0 else "",
                round(chain_amount, 2) if i == 0 else "",
                chain_purity if i == 0 else "",
                round(per_gram_mc, 2) if i == 0 else "",
                round(chain_mc, 2) if i == 0 else "",
                round(chain_wastage, 2) if i == 0 else "",
                round(chain_wastage_amount, 2) if i == 0 else "",
                round(jewellery_per_gram_mc, 2) if i == 0 else "",
                round(jewellery_mc_val, 2) if i == 0 else "",
                round(gold_wastage, 2) if i == 0 else "",
                round(jewellery_wastage_val, 2) if i == 0 else "",
                gemstone_pcs_val if gemstone_pcs_val != 0 else "",
                round(gemstone_cts_val, 2) if gemstone_cts_val != 0 else "",
                round(gemstone_amount_val, 2) if gemstone_amount_val != 0 else "",
                round(cert_charge, 2) if i == 0 else "",
                round(hallmark_charge, 2) if i == 0 else "",
                round(total_amt, 2)
            ])

    # --- SUM ROW ---
    sum_row = [""] * len(columns)
    sum_row[5]  = round(sum(float(r[5] or 0) for r in rows_diamond), 2)
    sum_row[6]  = round(sum(float(r[6] or 0) for r in rows_diamond), 2)
    sum_row[8]  = round(sum(float(r[8] or 0) for r in rows_diamond), 2)
    sum_row[10] = round(sum(float(r[10] or 0) for r in rows_diamond), 2)  # Total Diamond Rate
    sum_row[11] = round(sum(float(r[11] or 0) for r in rows_diamond), 2)  # Diamond Amount
    sum_row[12] = round(sum(float(r[12] or 0) for r in rows_diamond), 2)
    sum_row[13] = round(sum(float(r[13] or 0) for r in rows_diamond), 2)
    sum_row[14] = round(sum(float(r[14] or 0) for r in rows_diamond), 2)
    sum_row[16] = round(sum(float(r[16] or 0) for r in rows_diamond), 2)
    sum_row[17] = round(sum(float(r[17] or 0) for r in rows_diamond), 2)
    sum_row[19] = round(sum(float(r[19] or 0) for r in rows_diamond), 2)
    sum_row[20] = round(sum(float(r[20] or 0) for r in rows_diamond), 2)
    sum_row[23] = round(sum(float(r[23] or 0) for r in rows_diamond), 2)
    sum_row[25] = round(sum(float(r[25] or 0) for r in rows_diamond), 2)
    sum_row[27] = round(sum(float(r[27] or 0) for r in rows_diamond), 2)  #  Jewellery MC total
    sum_row[29] = round(sum(float(r[29] or 0) for r in rows_diamond), 2)
    sum_row[30] = round(sum(float(r[30] or 0) for r in rows_diamond), 2)
    sum_row[31] = round(sum(float(r[31] or 0) for r in rows_diamond), 2)
    sum_row[32] = round(sum(float(r[32] or 0) for r in rows_diamond), 2)
    sum_row[33] = round(sum(float(r[33] or 0) for r in rows_diamond), 2)
    sum_row[34] = round(sum(float(r[34] or 0) for r in rows_diamond), 2)
    # sum_row[35] = round(sum(float(r[35] or 0) for r in rows_diamond), 2)

    rows_diamond.append(sum_row)

    # --- Create Workbook ---
    wb = Workbook()
    ws = wb.active
    ws.title = "Diamond Detail"

    # --- Add Company Name ---
    company_name = "M/S. GURUKRUPA EXPORT PVT LIMITED"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    cell = ws.cell(row=1, column=1, value=company_name)
    cell.font = Font(bold=True, size=15)
    cell.alignment = Alignment(horizontal="center", vertical="center")

    # --- Add Headers ---
    for col_num, column_title in enumerate(columns, 1):
        c = ws.cell(row=2, column=col_num, value=column_title)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")

    # --- Add Data Rows ---
    for row_num, row_data in enumerate(rows_diamond, 3):
        for col_num, cell_value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=cell_value)

    # --- Auto column width ---
    for i, column in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(i)].width = 15

    # --- Save to BytesIO and Download ---
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    frappe.local.response.filecontent = output.read()
    frappe.local.response.filename = f"Diamond_Detail_SO_{docname}.xlsx"
    frappe.local.response.type = "download"

def update_si_data(self):
	self.is_customer_metal = False
	self.is_customer_diamond = False
	invoice_data = {}
	payment_terms_data = {}
	hallmarking_invoice_items = []
	is_branch_customer = frappe.db.get_value(
		"Sales Type Multiselect", {"parent": self.customer, "sales_type": "Branch"}
	)
	separate_hallmarking_invoice = frappe.db.get_value(
		"Customer", self.customer, "custom_separate_hallmarking_invoice"
	)
	allowed_item_types = get_allowed_item_types(self.customer, self.sales_type)
	for row in self.items:
		if row.bom and not self.item_same_as_above:
			
			exchange_rate = 1
			bom_doc = frappe.get_doc("BOM", row.bom)
			if bom_doc.currency != self.currency:
				exchange_rate = get_exchange_rate(
					bom_doc.currency, self.currency, transaction_date=self.posting_date
				)
			gold_rate_changed = True
			update_bom_details(self, row, bom_doc, is_branch_customer, invoice_data,gold_rate_changed)
			if bom_doc.hallmarking_amount :
				custom_item, hsn_code, uom = frappe.db.get_value(
				"E Invoice Item", {"is_for_hallmarking": 1}, ["name", "hsn_code", "uom"]
			)
				if invoice_data.get(custom_item):
					invoice_data[custom_item]["qty"] += 1
					invoice_data[custom_item]["amount"] += bom_doc.hallmarking_amount
					invoice_data[custom_item]["rate"] = invoice_data[custom_item]["amount"]/invoice_data[custom_item]["qty"]
				else:
					invoice_data[custom_item] = {
						"qty": 1,
						"hsn_code": hsn_code,
						"uom": uom,
						"amount": bom_doc.hallmarking_amount,
						"rate":bom_doc.hallmarking_amount,
						"income_account": row.income_account,
						"cost_center": row.cost_center,
					}
					
			if bom_doc.certification_amount and not separate_hallmarking_invoice:
				custom_item, hsn_code, uom = frappe.db.get_value(
				"E Invoice Item", {"is_for_certification": 1}, ["name", "hsn_code", "uom"]
			)
				if invoice_data.get(custom_item):
					invoice_data[custom_item]["qty"] += 1
					invoice_data[custom_item]["amount"] += bom_doc.certification_amount
					invoice_data[custom_item]["rate"] = invoice_data[custom_item]["amount"]/invoice_data[custom_item]["qty"]
				else:
					invoice_data[custom_item] = {
						"qty": 1,
						"hsn_code": hsn_code,
						"uom": uom,
						"rate":bom_doc.certification_amount,
						"amount": bom_doc.certification_amount,
						"income_account": row.income_account,
						"cost_center": row.cost_center,
					}
			
			update_einvoice_items(self, invoice_data, payment_terms_data ,allowed_item_types)
			if row.get("custom_freight_amount"):
				custom_item, hsn_code, uom = frappe.db.get_value(
					"E Invoice Item", {"is_for_freight": 1}, ["name", "hsn_code", "uom"]
				)
				if invoice_data.get(custom_item):
					invoice_data[custom_item]["qty"] += 1
					invoice_data[custom_item]["amount"] += row.custom_freight_amount
				else:
					invoice_data[custom_item] = {
						"qty": 1,
						"hsn_code": hsn_code,
						"uom": uom,
						"amount": row.custom_freight_amount,
						"income_account": row.income_account,
						"cost_center": row.cost_center,
					}
			if self.sales_type != 'Certification':
				bom_doc = frappe.get_doc("BOM", row.bom)
				# Sum directly from BOM detail rows — rates already updated in before_validate
				# if gold rate changed vs SO, otherwise original BOM amounts are used as-is
				if self.sales_type == "Hybrid":
					# Hybrid: customer-supplied material was already split out
					# into a separate "Subcontracting Charges" row on the Sales
					# Order (see add_hybrid_outwork_row in sales_order.py). Only
					# sum the company-owned material here, or that amount gets
					# counted twice — once on this row, once on the
					# Subcontracting Charges row.
					#
					# Making charge (metal/finding) and handling charge
					# (diamond) are always job-work value regardless of
					# is_customer_item — they never belong on this row, only
					# on the Subcontracting Charges row (see the matching
					# split in _update_bom_totals() in sales_order.py).
					metal_rows     = [r for r in bom_doc.metal_detail if not r.is_customer_item]
					finding_rows   = [r for r in bom_doc.finding_detail if not r.is_customer_item]
					diamond_total  = sum(flt(r.diamond_rate_for_specified_quantity) for r in bom_doc.diamond_detail if not r.is_customer_item)
					gemstone_total = sum(flt(r.gemstone_rate_for_specified_quantity) for r in bom_doc.gemstone_detail if not r.is_customer_item)
				else:
					metal_rows, finding_rows = bom_doc.metal_detail, bom_doc.finding_detail
					diamond_total  = bom_doc.total_diamond_amount
					gemstone_total = bom_doc.gemstone_bom_amount
				row.metal_amount = flt(sum(r.amount for r in metal_rows), 3) * exchange_rate
				row.making_amount = (
					0
					if self.sales_type == "Hybrid"
					else flt(sum(r.making_amount for r in finding_rows)) + flt(sum(r.making_amount for r in metal_rows))
				)
				row.finding_amount = flt(sum(r.amount for r in finding_rows), 3) * exchange_rate
				row.diamond_amount = diamond_total * exchange_rate
				row.gemstone_amount = gemstone_total * exchange_rate
				row.custom_certification_amount = bom_doc.certification_amount * exchange_rate
				if bom_doc.certification_amount and separate_hallmarking_invoice:
					row.custom_certification_amount = 0
				row.custom_freight_amount = bom_doc.freight_amount * exchange_rate
				row.custom_hallmarking_amount = bom_doc.hallmarking_amount * exchange_rate
				row.custom_custom_duty_amount = bom_doc.custom_duty_amount * exchange_rate
				row.rate = flt(
					row.metal_amount
					+ row.making_amount
					+ row.diamond_amount
					+ row.finding_amount
					+ row.gemstone_amount
					+ row.custom_certification_amount
					+ row.custom_freight_amount
					+ row.custom_hallmarking_amount
					+ row.custom_custom_duty_amount,
					3,
				)
				row.amount = row.qty * row.rate
	return payment_terms_data

def update_einvoice_items(self, invoice_data, payment_terms_data,allowed_item_types):
	if not self.get("invoice_item"):
		self.invoice_item = []
	else:
		self.set("invoice_item", [])
	precision = frappe.db.get_value('Customer',self.customer,"custom_precision_variable")
	for row in invoice_data:
		if row not in allowed_item_types:
			continue
		if invoice_data[row]["amount"] >= 0:
			amount = invoice_data[row]["amount"]
			qty = invoice_data[row]["qty"]
			if self.is_return:
				amount = -abs(flt(amount))
				qty = -abs(flt(qty))
			if payment_terms_data.get(row):
				payment_terms_data[row] += round(invoice_data[row]["amount"], precision)
			else:
				payment_terms_data[row] = round(invoice_data[row]["amount"], precision)
			self.append(
				"invoice_item",
				{
					"item_code": row,
					"item_name": row,
					"uom": invoice_data[row]["uom"] or "Nos",
					"gst_hsn_code": invoice_data[row]["hsn_code"],
					"conversion_factor": 1,
					"qty": invoice_data[row]["qty"],
					"rate": invoice_data[row].get("rate", 0),
					"base_rate": invoice_data[row].get("rate", 0),
					"amount": flt(invoice_data[row]["amount"], 3),
					"base_amount": invoice_data[row]["amount"],
					"income_account": invoice_data[row]["income_account"],
					"cost_center": invoice_data[row]["cost_center"],
				},
			)

def update_bom_details(self, row, bom_doc, is_branch_customer, invoice_data, gold_rate_changed=True):
	gold_item = None
	gold_making_item = None
	bom_doc.customer = self.customer
	precision = frappe.db.get_value("Customer", self.customer, "custom_precision_variable")
	so_doc = frappe.get_doc("Sales Order", row.sales_order)
	so_item_map = {}

	if not self.is_return and row.sales_order:
		so_doc = frappe.get_doc("Sales Order", row.sales_order)
		for item in so_doc.custom_invoice_item:
			so_item_map[item.item_code] = item

	def add_to_invoice(item_code, so_item, fallback_amount=0, fallback_qty=0,fallback_rate=0, hsn=None, uom=None):

		if so_item:
			amount = so_item.amount
			qty = so_item.qty
			rate = so_item.rate
		else:
			amount = fallback_amount
			qty = fallback_qty
			rate = fallback_rate

		if item_code in invoice_data:
			if so_item:
				invoice_data[item_code]["amount"] = amount
				invoice_data[item_code]["qty"] = qty
				invoice_data[item_code]["rate"] = rate
			else:
				# fallback case → can accumulate
				invoice_data[item_code]["amount"] += amount
				invoice_data[item_code]["qty"] += qty
				invoice_data[item_code]["rate"] = rate
		else:
			invoice_data[item_code] = {
				"amount": amount,
				"qty": qty,
				"rate": rate,
				"hsn_code": hsn,
				"uom": uom,
				"income_account": row.income_account,
				"cost_center": row.cost_center,
			}
	for i in bom_doc.metal_detail:
		update_making_charges(row, bom_doc, i, self.gold_rate_with_gst) 
		einvoice_item, hsn_code, uom = frappe.db.get_value(
			"E Invoice Item",
			{
				"is_for_metal": 1,
				"metal_type": i.metal_type,
				"metal_purity": i.metal_touch,
			},
			["name", "hsn_code", "uom"],
		) or (None, None, None)

		# Hybrid: making charge is always job-work value (Outwork rate),
		# regardless of is_customer_item — same rule as validate_item_dharm()
		# in sales_order.py.
		filter_value = (
			"is_for_labour"
			if (i.is_customer_item or self.sales_type == "Hybrid")
			else "is_for_making"
		)

		making_item, making_hsn, making_uom = frappe.db.get_value(
			"E Invoice Item",
			{
				filter_value: 1,
				"metal_type": i.metal_type,
				"metal_purity": i.metal_touch,
			},
			["name", "hsn_code", "uom"],
		) or (None, None, None)

		so_metal = so_item_map.get(einvoice_item)
		so_making = so_item_map.get(making_item)

		if einvoice_item and not i.is_customer_item:
			add_to_invoice(
				einvoice_item,
				so_metal,
				fallback_amount=round(i.amount,precision),
				fallback_qty=i.quantity,

				fallback_rate=i.rate,
				hsn=hsn_code,
				uom=uom,
			)

		if making_item and not is_branch_customer and self.sales_type != "Hybrid":
			is_metal_per_pc = (
				flt(i.making_rate) > 0
				and abs(flt(i.making_amount) - flt(i.making_rate)) < 0.01
			)
			making_amount_with_wastage = flt(i.making_amount) + flt(i.wastage_amount)
			# metal_making_qty = 1 if is_metal_per_pc else i.quantity
			add_to_invoice(
				making_item,
				so_making,
				fallback_amount=round(making_amount_with_wastage,precision),
				fallback_qty=i.quantity,
				fallback_rate=i.making_rate,
				hsn=making_hsn,
				uom=making_uom,
			)

		if i.is_customer_item:
			self.is_customer_metal = True


	for i in bom_doc.finding_detail:
		update_making_charges(row, bom_doc, i, self.gold_rate_with_gst)
		einvoice_item = hsn_code = uom = None

		# ---------------- Finding amount: category-specific match ----------------
		result = frappe.db.get_value(
			"E Invoice Item",
			{
				"is_for_finding": 1,
				"metal_type": i.metal_type,
				"metal_purity": i.metal_touch,
				"finding_category": i.finding_category,
			},
			["name", "hsn_code", "uom"],
		)
		if result:
			einvoice_item, hsn_code, uom = result
		else:
			# FIXED: fallback must only match a genuinely generic record
			# (finding_category is unset) — not just any record sharing
			# metal_type/metal_purity, which could accidentally be a
			# different category's record.
			result = frappe.db.get_value(
				"E Invoice Item",
				{
					"is_for_metal": 1,
					"metal_type": i.metal_type,
					"metal_purity": i.metal_touch,
					"finding_category": ["is", "not set"],
				},
				["name", "hsn_code", "uom"],
			)
			if result:
				einvoice_item, hsn_code, uom = result

		# Hybrid: making charge is always job-work value (Outwork rate),
		# regardless of is_customer_item — same rule as validate_item_dharm()
		# in sales_order.py.
		filter_value = (
			"is_for_labour"
			if (i.is_customer_item or self.sales_type == "Hybrid")
			else "is_for_finding_making"
		)

		making_item = making_hsn = making_uom = None

		# ---------------- Making charge: category-specific match ----------------
		result = frappe.db.get_value(
			"E Invoice Item",
			{
				filter_value: 1,
				"metal_type": i.metal_type,
				"metal_purity": i.metal_touch,
				"finding_category": i.finding_category,
			},
			["name", "hsn_code", "uom"],
		)
		if result:
			making_item, making_hsn, making_uom = result
		else:
			# Fallback mirrors the SO's behavior — drop into the generic
			# metal bucket for whichever filter_value we resolved above
			# (is_for_labour for Hybrid/customer-supplied, is_for_making
			# otherwise), not hardcoded to is_for_making, so Hybrid making
			# charges don't fall back into the Outright-taxed bucket.
			fallback_filter_value = (
				"is_for_labour"
				if (i.is_customer_item or self.sales_type == "Hybrid")
				else "is_for_making"
			)
			result = frappe.db.get_value(
				"E Invoice Item",
				{
					fallback_filter_value: 1,
					"metal_type": i.metal_type,
					"metal_purity": i.metal_touch,
				},
				["name", "hsn_code", "uom"],
			)
			if result:
				making_item, making_hsn, making_uom = result

		so_finding = so_item_map.get(einvoice_item)
		so_making = so_item_map.get(making_item)

		if einvoice_item and not i.is_customer_item:
			add_to_invoice(
				einvoice_item,
				so_finding,
				fallback_amount=round(i.amount,precision),
				fallback_qty=i.quantity,
				fallback_rate=i.rate,
				hsn=hsn_code,
				uom=uom,
			)

		if making_item and self.sales_type != "Hybrid":
			making_amount_with_wastage = flt(i.making_amount) + flt(i.wastage_amount)
			add_to_invoice(
				making_item,
				so_making,
				fallback_amount=round(making_amount_with_wastage,precision),
				fallback_qty=i.quantity,
				fallback_rate=i.making_rate,
				hsn=making_hsn,
				uom=making_uom,
			)
		if i.is_customer_item:
			self.is_customer_metal = True
	
	for i in bom_doc.diamond_detail:
		einvoice_item = hsn_code = uom = None
		result = frappe.db.get_value(
			"E Invoice Item",
			{
				"is_for_diamond": 1,
				"diamond_type": i.diamond_type
			},
			["name", "hsn_code", "uom"],
		)

		if result:
			einvoice_item, hsn_code, uom = result
		else:
			continue  

		so_item = so_item_map.get(einvoice_item)

		if so_item:
			amount = so_item.amount
			qty = so_item.qty
			rate = so_item.rate

		else:
			amount = _calculate_diamond_amount(bom_doc, i, {}, {})
			if is_branch_customer:
				amount = i.se_rate * i.quantity
			amount=i.diamond_rate_for_specified_quantity
			qty = i.quantity
			rate=amount/qty
			# rate = 0

		if not i.is_customer_item:
			if einvoice_item in invoice_data:
				invoice_data[einvoice_item]["amount"] = round(amount , precision)
				invoice_data[einvoice_item]["qty"] = qty
			else:
				invoice_data[einvoice_item] = {
					"amount": amount,
					"qty": qty,
					"rate": rate,
					"hsn_code": hsn_code,
					"uom": uom,
					"income_account": row.income_account,
					"cost_center": row.cost_center,
				}

		# if self.sales_type == "Hybrid":
		# 	# Handling charge is always job-work value, whether the stone
		# 	# is company-owned or customer-supplied — route it to the same
		# 	# labour/Subcontracting bucket used for metal/finding making
		# 	# above (mirrors the diamond handling split added to
		# 	# validate_item_dharm() in sales_order.py). Not filtered by uom —
		# 	# "Subcontracting Income" is Gram-based even though diamonds
		# 	# are Carat, matching the existing is_for_labour convention
		# 	# elsewhere in this function.
		# 	labour_item, labour_hsn, labour_uom = frappe.db.get_value(
		# 		"E Invoice Item",
		# 		{"is_for_labour": 1},
		# 		["name", "hsn_code", "uom"],
		# 	) or (None, None, None)
		# 	if labour_item:
		# 		so_labour = so_item_map.get(labour_item)
		# 		handling_amount = round(flt(i.quantity) * flt(i.handling_rate), precision)
		# 		add_to_invoice(
		# 			labour_item,
		# 			so_labour,
		# 			fallback_amount=handling_amount,
		# 			fallback_qty=i.quantity,
		# 			fallback_rate=i.handling_rate,
		# 			hsn=labour_hsn,
		# 			uom=labour_uom,
		# 		)

		if i.is_customer_item:
			self.is_customer_diamond = True
	for i in bom_doc.gemstone_detail:

		einvoice_item, hsn_code, uom = frappe.db.get_value(
			"E Invoice Item",
			{"is_for_gemstone": 1},
			["name", "hsn_code", "uom"],
		) or (None, None, None)

		if not einvoice_item:
			continue

		so_item = so_item_map.get(einvoice_item)

		if so_item:
			amount = so_item.amount
			qty = so_item.qty
			rate = so_item.rate
		else:
			
			amount = i.gemstone_rate_for_specified_quantity
			qty = i.quantity
			rate = i.total_gemstone_rate

			if is_branch_customer:
				amount = i.se_rate * i.quantity
				rate = i.se_rate

		
		if einvoice_item and not i.is_customer_item:
			if so_item_map.get(einvoice_item):
				invoice_data[f"{einvoice_item}"] = {
					"amount": so_item_map[einvoice_item].amount,
					"qty": so_item_map[einvoice_item].qty,
					"rate": so_item_map[einvoice_item].rate,
					"hsn_code": hsn_code,
					"uom": uom,
					"income_account": row.income_account,
					"cost_center": row.cost_center,
				}
			else:
				invoice_data[f"{einvoice_item}"] = {
					"amount": amount,
					"qty": i.quantity,
					"rate": i.rate,
					"hsn_code": hsn_code,
					"uom": uom,
					"income_account": row.income_account,
					"cost_center": row.cost_center,
				}

		if i.is_customer_item:
			self.is_customer_gemstone = True
	
	bom_doc.gold_rate_with_gst = self.gold_rate_with_gst
	customer_group=frappe.db.get_value('Customer',self.customer,'customer_group')
	if not (self.company == "KG GK Jewellers Private Limited" or customer_group == "Internal"):
		bom_doc.validate()
		bom_doc.save()
		update_totals("BOM", bom_doc.name)

def update_making_charges(row, bom_doc, bom_row, gold_rate):

	bom_doc.set_additional_rate = False

	item_details = frappe.db.get_value(
		"Item",
		row.item_code,
		["item_subcategory", "setting_type"],
		as_dict=True
	)

	sub_category = (item_details.get("item_subcategory") or "").strip()
	setting_type = item_details.get("setting_type")

	item_category = frappe.db.get_value("Item", row.item_code, "item_category")
	customer_group = frappe.db.get_value("Customer", bom_doc.customer, "customer_group")

	override_internal = (
		bom_doc.company == "KG GK Jewellers Private Limited"
		or customer_group == "Internal"
	)
	if not override_internal :
		MCP = frappe.qb.DocType("Making Charge Price")
		MCPIS = frappe.qb.DocType("Making Charge Price Item Subcategory")
		MCPFS = frappe.qb.DocType("Making Charge Price Finding Subcategory")

		child_table = MCPIS if bom_row.parentfield == "metal_detail" else MCPFS

		subcat_subcategory = (bom_row.get("finding_type") or sub_category or "").strip()

		# ---------------- MAIN QUERY ----------------
		query = (
			frappe.qb.from_(MCP)
			.left_join(child_table)
			.on(child_table.parent == MCP.name)
			.select(
				MCP.name.as_("mcp_name"),
				child_table.rate_per_gm,
				child_table.rate_per_pc,
				child_table.rate_per_gm_threshold,
				child_table.wastage,
				child_table.subcontracting_rate,
				child_table.subcontracting_wastage,
				child_table.wastage_per_pcs,
				child_table.supplier_fg_purchase_rate,
			)
			.where(
				(MCP.customer == bom_doc.customer)
				& (MCP.setting_type == setting_type)
				& (MCP.metal_type == bom_row.metal_type)
				& (MCP.metal_touch == bom_row.metal_touch)
				& (MCP.from_gold_rate <= gold_rate)
				& (MCP.to_gold_rate >= gold_rate)
			)
		)

		# ---------------- SUBCATEGORY MATCH (RELAXED) ----------------
		query = query.where(
			(child_table.subcategory == subcat_subcategory)
			| (child_table.subcategory.isnull())
			| (child_table.subcategory == "")
		)

		if bom_row.parentfield != "metal_detail":
			query = query.where(child_table.metal_touch == bom_row.metal_touch)
			# frappe.msgprint(f"huyt{bom_row.making_amount}")

		query = query.limit(1)
		# frappe.msgprint(f"yt{query}")
		making_charge_details = query.run(as_dict=True)

		

		# ---------------- FALLBACK FOR FINDING ----------------
		if not making_charge_details and bom_row.parentfield != "metal_detail":
			query = (
				frappe.qb.from_(MCP)
				.left_join(MCPFS)
				.on(MCPFS.parent == MCP.name)
				.select(
					MCPFS.rate_per_gm,
					MCPFS.rate_per_pc,
					MCPFS.rate_per_gm_threshold,
					MCPFS.wastage,
					MCPFS.wastage_per_pcs,
					MCPFS.subcontracting_rate,
					MCPFS.subcontracting_wastage,
					MCPFS.subcategory,
					MCPFS.supplier_fg_purchase_rate,
					ConstantColumn(1).as_("non_finding_rate"),
				)
				.where(
					(MCP.customer == bom_doc.customer)
					& (MCP.setting_type == setting_type)
					& (MCP.metal_type == bom_row.metal_type)
					& (MCP.metal_touch == bom_row.metal_touch)
					& (MCP.from_gold_rate <= gold_rate)
					& (MCP.to_gold_rate >= gold_rate)
				)
				.where(
					(MCPFS.subcategory == subcat_subcategory)
					| (MCPFS.subcategory.isnull())
					| (MCPFS.subcategory == "")
				)
				.limit(1)
			)

		
			making_charge_details = query.run(as_dict=True)
			if len(making_charge_details) > 0:
				making_charges = making_charge_details[0]
				rate_per_gm = flt(making_charges.get("rate_per_gm") or 0)
				rate_per_pc = flt(making_charges.get("rate_per_pc") or 0)
				threshold = flt(making_charges.get("rate_per_gm_threshold") or 2)
				subcontracting_rate=flt(making_charges.get("subcontracting_rate") or 0)
				subcontracting_wastage=flt(making_charges.get("subcontracting_wastage") or 0)

				# 🔥 IMPORTANT FIX: ALWAYS COMPARE WITH BOM LEVEL WEIGHT
				weight = flt(bom_doc.metal_and_finding_weight or 0)
				# ---------------- THRESHOLD LOGIC (FIXED FOR FINDING) ----------------
			
		# ---------------- APPLY RESULT ----------------
		if len(making_charge_details) > 0:
			making_charges = making_charge_details[0]
			rate_per_gm = flt(making_charges.get("rate_per_gm") or 0)
			rate_per_pc = flt(making_charges.get("rate_per_pc") or 0)
			threshold = flt(making_charges.get("rate_per_gm_threshold") or 2)
			subcontracting_rate=flt(making_charges.get("subcontracting_rate") or 0)
			subcontracting_wastage=flt(making_charges.get("subcontracting_wastage") or 0)
			# 🔥 IMPORTANT FIX: ALWAYS COMPARE WITH BOM LEVEL WEIGHT
			weight = flt(bom_doc.metal_and_finding_weight or 0)
			# frappe.throw(f"{weight}")

			# ---------------- THRESHOLD LOGIC (FIXED FOR FINDING) ----------------
			if bom_row.is_customer_item:
				bom_row.rate=0
				bom_row.making_rate =subcontracting_rate
				metal_making_charges = subcontracting_rate * bom_row.quantity
				wastage_rate = subcontracting_wastage
			else:
				if weight <= threshold :
					metal_making_charges = rate_per_pc
					# metal_making_charges = 5
					bom_row.making_rate = rate_per_pc
					bom_row.making_amount = bom_row.making_rate
					wastage_rate = flt(making_charges.get("wastage_per_pcs") or 0)
				else:
					metal_making_charges = rate_per_gm * bom_row.quantity
					bom_row.making_rate = rate_per_gm
					bom_row.making_amount = metal_making_charges
					wastage_rate = flt(making_charges.get("wastage") or 0)
			bom_row.making_amount = metal_making_charges
			bom_row.wastage_rate = wastage_rate
			bom_row.wastage_amount = bom_row.wastage_rate * bom_row.amount / 100

			additional_net_weight = 0

			# ---------------- DIAMOND LOGIC ----------------
			if not bom_doc.set_additional_rate and bom_row.parentfield == "metal_detail":
				if (
					frappe.db.get_value("Customer", bom_doc.customer, "compute_making_charges_on")
					== "Diamond Inclusive"
					and flt(bom_row.metal_purity) == flt(bom_doc.metal_purity)
				):
					if not bom_doc.total_diamond_weight_per_gram:
						bom_doc.total_diamond_weight_per_gram = flt(
							flt(bom_doc.total_diamond_weight) / 5, 3
						)

					bom_row.additional_net_weight = bom_doc.total_diamond_weight_per_gram
					additional_net_weight = bom_row.additional_net_weight
					bom_doc.set_additional_rate = True

			bom_row.fg_purchase_rate = flt(making_charges.get("supplier_fg_purchase_rate") or 0)
			bom_row.fg_purchase_amount = bom_row.fg_purchase_rate * (
				bom_row.quantity + additional_net_weight
			)


	bom_doc.making_charge = flt(
		sum(r.making_amount for r in bom_doc.metal_detail)
		+ sum(r.making_amount for r in bom_doc.finding_detail)
	)


def update_payment_terms(self, payment_terms_data=None):
	if self.payment_terms_template:
		return
	
	if not self.grand_total:
		return

	payment_term_dict = {}
	due_date_list = []
	item_to_append = []

	remaining_terms = []
	for row in payment_terms_data:
		payment_terms = frappe.db.get_value(
			"Customer Payment Terms Details",
			{"parent": self.customer, "item_type": row},
			["payment_term", "due_date_based_on", "due_days"],
			as_dict=1,
		)
		if payment_terms:
			if not payment_term_dict.get(payment_terms.payment_term):
				payment_term_dict.update(
					{
						payment_terms.payment_term: {
							"item_type": [row],
							"due_days": payment_terms.due_days,
							"due_date_based_on": payment_terms.due_date_based_on,
						}
					}
				)
			else:
				payment_term_dict[payment_terms.payment_term]["item_type"].append(row)

		else:
			remaining_terms.append(row)

	if remaining_terms:
		frappe.throw(
			_("Following Items not mentioned in Customer Payment Terms. <br><b>{0}</b>").format(
				"<br>".join(remaining_terms)
			)
		)
	# if custom_term:
	# 	for row in custom_term.customer_payment_details:
	# 		if not payment_term_dict.get(row.payment_term):
	# 			payment_term_dict.update(
	# 				{
	# 					row.payment_term: {
	# 						"item_type": [row.item_type],
	# 						"due_days": row.due_days,
	# 						"due_date_based_on": row.due_date_based_on,
	# 					}
	# 				}
	# 			)
	# 		else:
	# 			payment_term_dict[row.payment_term]["item_type"].append(row.item_type)

	total_metal_amount = 0
	total_making_amount = 0
	total_finding_amount = 0
	total_diamond_amount = 0
	total_gemstone_amount = 0

	if payment_term_dict:
		for row in self.items:
			if row.bom:
				total_metal_amount += (
					row.metal_amount
					+ row.custom_custom_duty_amount
					+ row.custom_hallmarking_amount
					+ row.custom_freight_amount
					+ row.custom_certification_amount
				)
				total_making_amount += row.making_amount
				total_finding_amount += row.finding_amount
				total_diamond_amount += row.diamond_amount
				total_gemstone_amount += row.gemstone_amount
	self.payment_schedule = []
	if payment_term_dict:
		due_date = None
		self.payment_terms_template = None
		for row in payment_term_dict:
			payment_amount = 0

			description = []
			for item_type in payment_term_dict[row]["item_type"]:
				charge_type = frappe.db.get_value("E Invoice Item", item_type, "charge_type")
				if charge_type in ["Making Charges", "Labour Charges"] and total_making_amount > 0:
					if charge_type == "Making Charges" and not self.is_customer_metal:
						
						# payment_amount += total_making_amount
						payment_amount += payment_terms_data.get(item_type)
						# payment_amount += self.total_taxes_and_charges
						description.append(item_type)
					elif charge_type != "Making Charges" and self.is_customer_metal:
						payment_amount += total_making_amount
						# payment_amount += self.total_taxes_and_charges
						description.append(item_type)
				elif charge_type == "Studded Metal" and total_metal_amount > 0:
					payment_amount += payment_terms_data.get(item_type)
					description.append(item_type)
				elif charge_type in ["Studded Diamond", "Handling Charges"] and total_diamond_amount > 0:
					if charge_type == "Studded Diamond" and not self.is_customer_diamond:
						payment_amount += payment_terms_data.get(item_type)
						# payment_amount += total_diamond_amount
						description.append(item_type)
					elif self.is_customer_diamond and charge_type != "Studded Diamond":
						payment_amount += total_diamond_amount
						description.append(item_type)
				elif charge_type == "Studded Gemstone" and total_gemstone_amount > 0:
					payment_amount += total_gemstone_amount
					description.append(item_type)

			if payment_term_dict[row]["due_date_based_on"] == "Day(s) after invoice date":
				due_date = getdate(self.posting_date) + timedelta(
						days=int(payment_terms["due_days"])
							)

			elif payment_term_dict[row]["due_date_based_on"] == "Day(s) after the end of the invoice month":
				posting_date = get_last_day(self.posting_date)
				due_date = datetime.strptime(posting_date, "%Y-%m-%d") + timedelta(
					days=int(payment_term_dict[row]["due_days"])
				)

			due_date_list.append(due_date)
			if payment_amount > 0:
				# if self.disable_rounded_total == 0:
				payment_amount = flt(payment_amount, 3)

				item_to_append.append(
					{
						"due_date": due_date,
						"description": ", ".join(item_type for item_type in description),
						"payment_term": row,
						"payment_amount": payment_amount,
						"custom_invoice_portion": flt((payment_amount / self.grand_total) * 100)
						if self.grand_total > 0
						else 0,
					}
				)
		self.payment_schedule = []
		self.due_date = max(due_date_list)
		self.extend("payment_schedule", item_to_append)

def update_income_account(self):
	if self.is_opening == "No":
		income_account = frappe.db.get_value(
			"Account", {"company": self.company, "custom_sales_type": self.sales_type}, "name"
		)
		if income_account:
			for row in self.items:
				row.income_account = income_account

@frappe.whitelist()
@frappe.validate_and_sanitize_search_inputs
def get_completed_product_return_orders(doctype, txt, searchfield, start, page_len, filters):

    return frappe.db.sql("""
        SELECT pro.name
        FROM `tabProduct Return Order Form` pro
        WHERE pro.docstatus = 1
          AND pro.name LIKE %(txt)s
          AND NOT EXISTS (
                SELECT 1
                FROM `tabProduct Return Order` prf
                WHERE prf.product_return_order_form = pro.name
                  AND prf.docstatus != 1
          )
          AND EXISTS (
                SELECT 1
                FROM `tabProduct Return Order` prf
                WHERE prf.product_return_order_form = pro.name
          )
        ORDER BY pro.modified DESC
        LIMIT %(start)s, %(page_len)s
    """, {
        "txt": f"%{txt}%",
        "start": start,
        "page_len": page_len,
    })

