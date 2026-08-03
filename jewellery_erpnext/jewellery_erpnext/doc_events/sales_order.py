import frappe
import requests
from frappe import _
from frappe.model.mapper import get_mapped_doc
from frappe.utils import flt

from jewellery_erpnext.jewellery_erpnext.customization.sales_order.doc_events.branch_utils import (
	create_branch_so,
)
from jewellery_erpnext.jewellery_erpnext.doc_events.bom_utils import (
	calculate_gst_rate,
	set_bom_item_details,
	set_bom_rate,
)

# Companies allowed to use Sales Type "Hybrid" (3% on company-owned material +
# 5% on customer-supplied material within the same BOM). Extend this tuple to
# roll Hybrid out to more companies later.
HYBRID_ENABLED_COMPANIES = ("KG GK Jewellers Private Limited",)

# Placeholder item that carries the aggregated customer-supplied (5%) amount
# for Hybrid Sales Orders — see add_hybrid_outwork_row().
HYBRID_OUTWORK_ITEM = "Subcontracting Charges"

# Item Tax Templates for Hybrid rows, by company — same "Outright"
# (owned, 3%) / "Outwork" (customer-supplied, 5%) template naming
# already used in tax() and set_gst_details() above. Also used directly by
# set_sales_type_tax_template() for plain (non-Hybrid) Outright /
# Outwork orders.
SALES_TYPE_ITEM_TAX_TEMPLATE = {
	"Outright": {
		"Gurukrupa Export Private Limited": "GST 3% - GEPL",
		"KG GK Jewellers Private Limited": "GST 3% - KGJPL",
	},
	"Outwork": {
		"Gurukrupa Export Private Limited": "GST 5% - GEPL",
		"KG GK Jewellers Private Limited": "GST 5% - KGJPL",
	},
}


def before_validate(self, method):
	# Drop any Hybrid Outwork-charge row left over from a previous
	# save before create_new_bom1 runs, so it never gets treated as a real
	# serial-no/BOM row. add_hybrid_outwork_row() rebuilds it fresh.
	self.items = [row for row in self.items if not row.get("custom_is_subcontracting_charge_row")]

	validate_sales_type(self)
	validate_quotation_item(self)
	set_repair_serial_bom(self)
	create_new_bom1(self)
	add_hybrid_outwork_row(self)
	set_missing_tax_category_and_template(self)
	set_sales_type_tax_template(self)
	validate_serial_number(self)
	# set_gst_details(self)  # superseded by add_hybrid_outwork_row() for Hybrid; see history for why
	validate_item_dharm(self)
	if not self.get("__islocal") and self.docstatus == 0:
		set_bom_item_details(self)

def set_missing_tax_category_and_template(self):
	"""
	Re-derive tax_category and taxes_and_charges when either is blank (e.g.
	cleared by a user on an existing order), so GST still comes through
	automatically instead of staying empty forever. Only fills in what's
	missing - never overrides a value already on the doc.

	Deliberately narrower than the old (disabled) set_gst_details(): it does
	NOT set item-level cgst_rate/sgst_rate/igst_rate using the
	customer-specific "Sales Type Multiselect" rate - that's left entirely
	to set_sales_type_tax_template()/sync_header_tax_rate(), which already
	drives the real 3%/5% rate from the company's Item Tax Template.

	The taxes child table itself IS rebuilt here (copied straight from the
	template's own Sales Taxes and Charges rows) rather than relying on
	core's AccountsController.set_taxes_and_charges() - that only fires when
	Accounts Settings.add_taxes_from_taxes_and_charges_template is enabled,
	which isn't guaranteed to be on on every site (confirmed off on at least
	one), so doing it explicitly here keeps this independent of that setting.
	"""
	if self.sales_type not in ("Outright", "Outwork", "Branch Sales", "Hybrid"):
		return
	if self.tax_category and self.taxes_and_charges and self.get("taxes"):
		return

	customer_state = frappe.db.get_value("Address", self.customer_address, "gst_state_number")
	company_state = frappe.db.get_value("Address", self.company_address, "gst_state_number")
	if not customer_state or not company_state:
		return

	if not self.tax_category:
		self.tax_category = "In-State" if customer_state == company_state else "Out-State"

	if not self.taxes_and_charges:
		taxes_and_charges = frappe.db.get_value(
			"Sales Taxes and Charges Template",
			{"company": self.company, "tax_category": self.tax_category, "disabled": 0},
			"name",
		)
		if not taxes_and_charges:
			frappe.log_error(
				f"No Sales Taxes and Charges Template found for "
				f"Company: {self.company}, Tax Category: {self.tax_category}",
				"set_missing_tax_category_and_template",
			)
			return
		self.taxes_and_charges = taxes_and_charges

	if not self.get("taxes"):
		tax_rows = frappe.get_all(
			"Sales Taxes and Charges",
			filters={"parent": self.taxes_and_charges},
			fields=["charge_type", "account_head", "description", "rate", "cost_center"],
			order_by="idx asc",
		)
		for t in tax_rows:
			self.append("taxes", {
				"charge_type": t.charge_type,
				"account_head": t.account_head,
				"description": t.description,
				"rate": t.rate,
				"cost_center": t.cost_center,
			})




def set_repair_serial_bom(self):
	"""Bridge the repair BOM Quotation -> Sales Order for header-order_type repairs.

	A repair is marked at the header level via ``order_type == "Repair"``. The manufacturing BOM for
	such rows comes from the source Quotation item, but the Quotation field ``custom_serial_id_bom``
	has a different name than the Sales Order Item field ``serial_id_bom``, so ``get_mapped_doc`` does
	NOT auto-copy it (and ``quotation_item`` is a Data field, so ``fetch_from`` can't bridge it).
	Here we copy it explicitly for each mapped repair row that doesn't already have a ``serial_id_bom``.
	The downstream Manufacturing Plan repair fetch resolves the row BOM from this field.
	"""
	if self.get("order_type") != "Repair":
		return

	for item in self.items:
		if item.get("serial_id_bom") or not item.get("quotation_item"):
			continue
		bom = frappe.db.get_value(
			"Quotation Item",
			item.quotation_item,
			["custom_serial_id_bom", "custom_repair_bom"],
			as_dict=True,
		)
		if not bom:
			continue
		item.serial_id_bom = bom.custom_serial_id_bom or bom.custom_repair_bom


def on_submit(self, method):
	validate_snc(self)

def before_submit(self, method):
	if not self.get("custom_invoice_item"):
		frappe.throw(_("Invoice Item table is mandatory for submission."))

def on_cancel(self, method):
	for row in self.items:
		# Cancel tracking BOM if present
		if row.custom_tracking_bom:
			tb = frappe.get_doc("Tracking Bom", row.custom_tracking_bom)
			if tb.docstatus == 1:
				tb.cancel()
		# Also handle legacy BOM references
		if row.bom:
			bom = frappe.get_doc("BOM", row.bom)
			bom.is_active = 0
			row.bom = ""
	# cancel_bom(self)
	validate_snc(self)

def tax(self):
	for row in self.items:
		item_tax_template = ''
		account_list = []
		customer_state = frappe.db.get_value("Address", {"name": self.customer_address}, "gst_state_number")
		company_state = frappe.db.get_value("Address", {"name": self.company_address}, "gst_state_number")
		self.tax_category = 'In-State' if customer_state == company_state else 'Out-State'
		# Map Sales Type + Company to appropriate Item Tax Template
		template_map = {
			'Outright': {
				'Gurukrupa Export Private Limited': 'GST 3% - GEPL',
				'KG GK Jewellers Private Limited': 'GST 3% - KGJPL',
			},
			'Outwork': {
				'Gurukrupa Export Private Limited': 'GST 5% - GEPL',
				'KG GK Jewellers Private Limited': 'GST 5% - KGJPL',
			},
		}
		item_tax_template = template_map.get(self.sales_type, {}).get(self.company, '')
		if frappe.db.get_value("Item", row.item_code, "item_subcategory"):

					if item_tax_template:
						row.item_tax_template = item_tax_template
                    
                    
                    # Per-line indicative GST split for UI; actual accounts come from template
					if self.tax_category == 'Out-State':
						row.igst = 5.0 if self.sales_type == 'Outwork' else 3.0
						row.igst_amount = round((row.net_rate or 0) * (row.igst / 100), 2)
						row.cgst_amount = 0
						row.sgst_amount = 0
					
					else:
						rate = 5.0 if self.sales_type == 'Outwork' else 3.0
						row.cgst = rate / 2
						row.sgst = rate / 2
						row.cgst_amount = (row.net_rate or 0) * (row.cgst / 100)
						row.sgst_amount = (row.net_rate or 0) * (row.sgst / 100)
						row.igst_amount = 0

			
		self.taxes = []

			
		if item_tax_template:
			
			if item_tax_template not in ['Exempted - GEPL', 'Exempted - KGJPL', 'Exempted - SHC', 'Exempted - SD']:
				row.item_tax_template = item_tax_template
				row.gst_treatment = 'Taxable'

				if self.tax_category == 'In-State':
					if not self.is_reverse_charge:
						tax = frappe.db.sql(
							f"""select tax_type,tax_rate
								from `tabItem Tax Template Detail`
								where parent = '{item_tax_template}'
									and tax_type not like '%IGST%'
									and tax_type like 'Output%'
									and tax_type not like '%RCM%'""",
							as_dict=1,
						)
					else:
						tax = frappe.db.sql(
							f"""select tax_type,tax_rate
								from `tabItem Tax Template Detail`
								where parent = '{item_tax_template}'
									and (tax_type like '%RCM%' or (tax_type like 'Output%' and tax_type not like 'Input%'))
									and tax_type not like '%IGST%'""",
							as_dict=1,
						)
						
				else:
					if not self.is_reverse_charge:
						tax = frappe.db.sql(
							f"""select tax_type,tax_rate
								from `tabItem Tax Template Detail`
								where parent = '{item_tax_template}'
									and tax_type like '%IGST%'
									and tax_type like 'Output%'
									and tax_type not like '%RCM%'""",
							as_dict=1,
						)
					else:
						tax = frappe.db.sql(
							f"""select tax_type,tax_rate
								from `tabItem Tax Template Detail`
								where parent = '{item_tax_template}'
									and (tax_type like '%RCM%' or tax_type like 'Output%')
									and tax_type like '%IGST%'""",
							as_dict=1,
						)
					# frappe.throw(f"{tax}")

				account_list = []
				for j in tax:
					if j.get("tax_type") in account_list:
						continue
					account_list.append(j.get("tax_type"))

					if 'IGST RCM' in j.get("tax_type"):
						gst_tax_type = 'igst_rcm'
					elif 'SGST RCM' in j.get("tax_type"):
						gst_tax_type = 'sgst_rcm'
					elif 'CGST RCM' in j.get("tax_type"):
						gst_tax_type = 'cgst_rcm'
					elif 'IGST' in j.get("tax_type"):
						gst_tax_type = 'igst'
					elif 'SGST' in j.get("tax_type"):
						gst_tax_type = 'sgst'
					elif 'CGST' in j.get("tax_type"):
						gst_tax_type = 'cgst'
					else:
						gst_tax_type = None

					add_deduct_tax = "Deduct" if 'RCM' in j.get("tax_type") else "Add"
					
					self.append("taxes", {
						"category": "Total",
						"add_deduct_tax": add_deduct_tax,
						"charge_type": "On Net Total",
						"account_head": j.get("tax_type"),
						"description": j.get("tax_type").replace(" - GE", ""),
						"rate": j.get("tax_rate"),
						"tax_amount":(self.total or 0) * (j.get("tax_rate", 0) / 100),
						"total":self.total + (self.total or 0) * (j.get("tax_rate", 0) / 100) ,
						"gst_tax_type": gst_tax_type
					})
				self.grand_total = self.total + (self.total or 0) * (j.get("tax_rate", 0) / 100)
				self.rounded_total =self.grand_total

def set_gst_details(self):
    if self.sales_type not in ("Outright", "Outwork", "Branch Sales", "Hybrid"):
        return

    customer_state = frappe.db.get_value("Address", self.customer_address, "gst_state_number")
    company_state  = frappe.db.get_value("Address", self.company_address,  "gst_state_number")

    if not customer_state or not company_state:
        return

    self.tax_category = "In-State" if customer_state == company_state else "Out-State"

    item_template_map = {
        "Outright": {
            "Gurukrupa Export Private Limited": "GST 3% - GEPL",
            "KG GK Jewellers Private Limited":  "GST 3% - KGJPL",
        },
        "Branch Sales": {
        "Gurukrupa Export Private Limited": "GST 3% - GEPL"
        },
        "Outwork": {
            "Gurukrupa Export Private Limited": "GST 5% - GEPL",
            "KG GK Jewellers Private Limited":  "GST 5% - KGJPL",
        },
    }
    # Hybrid has no rate/template of its own — for account-head selection
    # only, it borrows the Outright template; the real 3%/5% rates
    # come from the customer's Outright / Outwork rows (see
    # _set_hybrid_gst_details below).
    item_template_map["Hybrid"] = item_template_map["Outright"]

    item_tax_template = item_template_map.get(self.sales_type, {}).get(self.company)
    # frappe.throw(f"{item_tax_template}")
    if not item_tax_template:
        return

    taxes_and_charges = frappe.db.get_value(
        "Sales Taxes and Charges Template",
        {
            "company":      self.company,
            "tax_category": self.tax_category,
            "disabled":     0,
        },
        "name"
    )

    if not taxes_and_charges:
        frappe.log_error(
            f"No Sales Taxes and Charges Template found for "
            f"Company: {self.company}, Tax Category: {self.tax_category}",
            "set_gst_details"
        )
        return

    self.taxes_and_charges = taxes_and_charges

    if self.sales_type == "Hybrid":
        _set_hybrid_gst_details(self, item_tax_template)
        return

    template_rates = frappe.get_all(
        "Item Tax Template Detail",
        filters={"parent": item_tax_template},
        fields=["tax_type", "tax_rate"]
    )

    tax_rate = frappe.db.get_value(
        "Sales Type Multiselect",
        {
            "parent": self.customer,
            "sales_type": self.sales_type
        },
        "tax_rate"
    )

    tax_rate_f = flt(tax_rate)
    cgst_rate = sgst_rate = igst_rate = 0.0
    for r in template_rates:
        tax_type = r.tax_type or ""
        if "Output" not in tax_type or "RCM" in tax_type:
            continue
        if "CGST" in tax_type:
            cgst_rate = flt(tax_rate_f)
        elif "SGST" in tax_type:
            sgst_rate = flt(tax_rate_f)
        elif "IGST" in tax_type:
            igst_rate = flt(tax_rate_f)

    account_rate_map = {}
    for r in template_rates:
        tax_type = r.tax_type or ""
        if "Output" not in tax_type or "RCM" in tax_type:
            continue
        account_rate_map[r.tax_type] = flt(tax_rate_f)

    self.taxes = []
    tax_rows = frappe.get_all(
        "Sales Taxes and Charges",
        filters={"parent": self.taxes_and_charges},
        fields=["charge_type", "account_head", "description", "rate", "cost_center"],
        order_by="idx asc"
    )
    for t in tax_rows:
        correct_rate = account_rate_map.get(t.account_head, t.rate)
        self.append("taxes", {
            "charge_type":  t.charge_type,
            "account_head": t.account_head,
            "description":  t.description,
            "rate":         correct_rate,
            "cost_center":  t.cost_center,
            "tax_amount":correct_rate * self.total /100
        })
    for item in self.items:
        if not item.item_code:
            continue

        item.item_tax_template = item_tax_template
        item.gst_treatment     = "Taxable"
        item.cgst_rate         = 0.0
        item.sgst_rate         = 0.0
        item.igst_rate         = 0.0
        item.cgst_amount       = 0.0
        item.sgst_amount       = 0.0
        item.igst_amount       = 0.0

        taxable_value = flt(item.taxable_value)

        if self.tax_category == "In-State":
            item.cgst_rate   = tax_rate/2
            item.sgst_rate   = tax_rate/2
            item.cgst_amount = flt(taxable_value * item.cgst_rate / 100, 2)
            item.sgst_amount = flt(taxable_value * item.sgst_rate / 100, 2)
        else:
            item.igst_rate   = tax_rate
            item.igst_amount = flt(taxable_value * item.igst_rate / 100, 2)

def _set_hybrid_gst_details(self, item_tax_template):
    """
    Hybrid: each Sales Order Item row carries both company-owned material
    (taxed at the customer's Outright rate, e.g. 3%) and
    customer-supplied material (taxed at the customer's Outwork
    rate, e.g. 5%) — split earlier into row.custom_company_owned_amount /
    row.custom_customer_supplied_amount by _update_bom_totals().

    A single row can't carry two distinct legal GST rates, so the row's
    cgst_rate/sgst_rate/igst_rate become a blended (weighted-average) rate
    while cgst_amount/sgst_amount/igst_amount are the exact correct rupee
    totals. The header "taxes" table is not limited that way, so it gets
    the clean real-rate breakdown instead (e.g. CGST 1.5% + SGST 1.5% on
    the owned portion, CGST 2.5% + SGST 2.5% on the supplied portion).
    """
    if self.company not in HYBRID_ENABLED_COMPANIES:
        frappe.throw(_("Hybrid Sales Type is not yet enabled for company {0}").format(self.company))

    fg_rate = flt(frappe.db.get_value(
        "Sales Type Multiselect",
        {"parent": self.customer, "sales_type": "Outright"},
        "tax_rate",
    ))
    sc_rate = flt(frappe.db.get_value(
        "Sales Type Multiselect",
        {"parent": self.customer, "sales_type": "Outwork"},
        "tax_rate",
    ))

    total_owned    = sum(flt(item.custom_company_owned_amount)     for item in self.items)
    total_supplied = sum(flt(item.custom_customer_supplied_amount) for item in self.items)

    self.taxes = []
    tax_rows = frappe.get_all(
        "Sales Taxes and Charges",
        filters={"parent": self.taxes_and_charges},
        fields=["charge_type", "account_head", "description", "rate", "cost_center"],
        order_by="idx asc",
    )

    # One template row per account head (CGST/SGST/IGST) — reused for both
    # the Outright and Outwork rows we append below.
    account_by_type = {}
    for t in tax_rows:
        head = t.account_head or ""
        if "IGST" in head:
            account_by_type.setdefault("IGST", t)
        elif "CGST" in head:
            account_by_type.setdefault("CGST", t)
        elif "SGST" in head:
            account_by_type.setdefault("SGST", t)

    def _append_tax_row(account_key, rate, amount, label):
        template = account_by_type.get(account_key)
        if not template:
            return
        self.append("taxes", {
            "charge_type":  template.charge_type,
            "account_head": template.account_head,
            "description":  f"{template.description} ({label})",
            "rate":         rate,
            "cost_center":  template.cost_center,
            "tax_amount":   amount,
        })

    if self.tax_category == "In-State":
        _append_tax_row("CGST", fg_rate / 2, round(total_owned * fg_rate / 2 / 100, 2), "Outright")
        _append_tax_row("SGST", fg_rate / 2, round(total_owned * fg_rate / 2 / 100, 2), "Outright")
        _append_tax_row("CGST", sc_rate / 2, round(total_supplied * sc_rate / 2 / 100, 2), "Outwork")
        _append_tax_row("SGST", sc_rate / 2, round(total_supplied * sc_rate / 2 / 100, 2), "Outwork")
    else:
        _append_tax_row("IGST", fg_rate, round(total_owned * fg_rate / 100, 2), "Outright")
        _append_tax_row("IGST", sc_rate, round(total_supplied * sc_rate / 100, 2), "Outwork")

    for item in self.items:
        if not item.item_code:
            continue

        item.item_tax_template = item_tax_template
        item.gst_treatment     = "Taxable"

        owned         = flt(item.custom_company_owned_amount)
        supplied      = flt(item.custom_customer_supplied_amount)
        taxable_value = owned + supplied

        if self.tax_category == "In-State":
            cgst_amount = round(owned * fg_rate / 2 / 100 + supplied * sc_rate / 2 / 100, 2)
            item.cgst_amount = cgst_amount
            item.sgst_amount = cgst_amount
            item.igst_amount = 0.0
            item.cgst_rate   = round(cgst_amount / taxable_value * 100, 4) if taxable_value else 0.0
            item.sgst_rate   = item.cgst_rate
            item.igst_rate   = 0.0
        else:
            igst_amount = round(owned * fg_rate / 100 + supplied * sc_rate / 100, 2)
            item.igst_amount = igst_amount
            item.cgst_amount = 0.0
            item.sgst_amount = 0.0
            item.cgst_rate   = 0.0
            item.sgst_rate   = 0.0
            item.igst_rate   = round(igst_amount / taxable_value * 100, 4) if taxable_value else 0.0


def add_hybrid_outwork_row(self):
	"""
	Hybrid: each real row's amount is currently owned + supplied combined
	(see the custom_company_owned_amount / custom_customer_supplied_amount
	split computed in _update_bom_totals). Move the supplied portion out of
	every row and into a single aggregated HYBRID_Outwork_ITEM row,
	so each row ends up taxed at one clean rate instead of a blended one —
	real rows at the Outright rate, this row at the Outwork
	rate (the item itself carries 5% Item Tax Templates).
	"""
	if self.sales_type != "Hybrid":
		return

	ctx   = _get_bom_context(self)
	_prec = int(ctx.precision or 2)

	fg_template = SALES_TYPE_ITEM_TAX_TEMPLATE["Outright"].get(self.company)
	sc_template = SALES_TYPE_ITEM_TAX_TEMPLATE["Outwork"].get(self.company)

	total_supplied = 0.0
	for row in self.items:
		# Every remaining row is owned-only after the split below, so it
		# always gets the Outright (3%) template.
		if fg_template:
			row.item_tax_template = fg_template

		supplied = flt(row.custom_customer_supplied_amount)
		if not supplied:
			continue
		row.amount = round(flt(row.amount) - supplied, _prec)
		row.rate   = row.amount
		row.custom_customer_supplied_amount = 0
		total_supplied += supplied

	if total_supplied:
		gst_hsn_code = frappe.db.get_value("Item", HYBRID_OUTWORK_ITEM, "gst_hsn_code")
		self.append("items", {
			"item_code": HYBRID_OUTWORK_ITEM,
			"item_name": HYBRID_OUTWORK_ITEM,
			"description": HYBRID_OUTWORK_ITEM,
			"qty": 1,
			"uom": "Nos",
			"conversion_factor": 1,
			"rate": round(total_supplied, _prec),
			"amount": round(total_supplied, _prec),
			"gst_hsn_code": gst_hsn_code,
			"item_tax_template": sc_template,
			"custom_is_subcontracting_charge_row": 1,
		})

	self.total = round(sum(flt(item.amount) for item in self.items), _prec)


def clear_hybrid_header_tax_rate(self, method=None):
	"""
	Hybrid mixes two rates (3% owned + 5% supplied) in one order, so unlike
	Outright / Outwork there's no single correct number for
	sync_header_tax_rate() to put on the header row. Every row already
	carries its own item_tax_template (set in add_hybrid_outwork_row),
	which is what actually drives the computed tax_amount — so the header
	row's own rate is just as cosmetically unused here as it is there.
	Rather than leave the Sales Taxes and Charges Template's generic
	nominal rate (e.g. 9%) showing — which asserts a specific number
	that's actually wrong for both halves of the order — clear it to 0
	instead of picking either one. The correct per-rate detail still shows
	in the GST breakup table.

	Hooked on before_save (not before_validate): the `taxes` table isn't
	populated from the template yet during before_validate on a brand-new
	order — it only exists once core's tax calculation has run as part of
	validate(). Running this before_save (after that, and after India
	Compliance's own before_save hook, since this app is last in
	apps.txt) guarantees there's something to clear on every save,
	first save included.
	"""
	if self.sales_type != "Hybrid":
		return

	for row in self.taxes:
		if row.gst_tax_type:
			row.rate = 0


def set_sales_type_tax_template(self):
	"""
	Outright / Outwork (non-Hybrid): set each row's
	item_tax_template from sales_type + company, same as
	add_hybrid_outwork_row() does for Hybrid rows. Without this,
	item_tax_template stays blank and India Compliance silently falls back
	to the header Sales Taxes and Charges Template's own flat nominal rate
	(e.g. 9%+9%) instead of the correct 3%/5% — confirmed against several
	live Sales Orders where this was wrong.
	"""
	if self.sales_type not in ("Outright", "Outwork"):
		return

	template = SALES_TYPE_ITEM_TAX_TEMPLATE[self.sales_type].get(self.company)
	if not template:
		return

	for row in self.items:
		if row.item_code:
			row.item_tax_template = template

	sync_header_tax_rate(self, template)


def sync_header_tax_rate(self, template):
	"""
	Outright / Outwork: the header Sales Taxes and Charges
	Template row carries a generic nominal rate (e.g. 9%) that has nothing
	to do with jewellery GST — the real 3%/5% comes entirely from each
	item's item_tax_template (set above), which already drives the actual
	computed tax_amount regardless of what the header row's own rate says.
	That leaves the header showing "9%" next to a correctly-computed
	3%/5% amount, which reads as wrong even though the money is right.

	Cosmetic-only: overwrite the header row's displayed rate to match the
	same Item Tax Template already driving the real calculation, so the
	number on screen can never drift from the number being charged. Only
	safe here because these sales types have exactly one rate for the
	whole order — Hybrid mixes two rates and needs the header row's rate
	to stay as-is (see _set_hybrid_gst_details / the GST breakup table).
	"""
	if not self.taxes:
		return

	rate_by_account = {
		d.tax_type: d.tax_rate
		for d in frappe.get_all(
			"Item Tax Template Detail",
			filters={"parent": template},
			fields=["tax_type", "tax_rate"],
		)
	}

	for row in self.taxes:
		if row.account_head in rate_by_account:
			row.rate = rate_by_account[row.account_head]


CHUNK_SIZE = 10
ENQUEUE_THRESHOLD = 10


_gold_rate_cache = {}
_company_ctx_cache = {}  # (company, key)            → cctx
_making_charge_cache = {}  # (customer, metal_type, setting_type, gold_rate, touch) → (mc_name, sub_info, threshold)
_metal_purity_cache = {}  # (customer, metal_type, metal_touch) → metal_purity
_ccp_cache = {}  # customer → ccp_doc | None
_gemstone_pl_cache = {}  # customer → price_list_type


def _clear_caches():
	_company_ctx_cache.clear()
	_making_charge_cache.clear()
	_metal_purity_cache.clear()
	_ccp_cache.clear()
	_gemstone_pl_cache.clear()

def _get_bom_context(self):
	gold_gst_rate = frappe.db.get_single_value("Jewellery Settings", "gold_gst_rate")
	(
		customer_group,
		precision,
		metal_precision,
		stone_precision,
		precision_for_net_weight,
		precision_for_gross_weight,
	) = frappe.db.get_value(
		"Customer",
		self.customer,
		[
			"customer_group",
			"custom_precision_variable",
			"custom_precision_for_metal",
			"custom_precision_for_stone",
			"custom_precision_for_net_weight",
			"custom_precision_for_gross_weight",
		],
	)
	if self.get("custom_precision"):
		metal_precision = 3
	if self.get("custom_precision_for_stone"):
		stone_precision = 3
	return frappe._dict(
		gold_gst_rate=gold_gst_rate,
		customer_group=customer_group,
		precision=precision,
		metal_precision=metal_precision or 3,
		stone_precision=stone_precision or 3,
		precision_for_net_weight=precision_for_net_weight or 3,
		precision_for_gross_weight=precision_for_gross_weight or 3,
	)


def _get_company_context(self, row, ctx):
	"""
	Cached version — most rows on the same SO share identical company context.
	KG GK  : keyed by serial_no (each serial has its own manufacturing chain)
	Gurukrupa internal : keyed by custom_parent_sales_order (same for whole SO)
	everything else    : nothing is queried, return empty dict immediately
	"""
	if self.company == "KG GK Jewellers Private Limited":
		cache_key = ("KGGK", row.serial_no)
	elif (
		self.company == "Gurukrupa Export Private Limited"
		and ctx.customer_group == "Internal"
	):
		cache_key = ("GEPL", self.custom_parent_sales_order)
	else:
		# No DB work needed — skip cache overhead
		return frappe._dict(
			reference_customer=None,
			billing_currency=None,
			exchange_rate=None,
			reference_customer_c2c=None,
		)

	if cache_key not in _company_ctx_cache:
		cctx = frappe._dict(
			reference_customer=None,
			billing_currency=None,
			exchange_rate=None,
			reference_customer_c2c=None,
		)

		if self.company == "KG GK Jewellers Private Limited":
			creation_no = frappe.get_value("Serial No", row.serial_no, "reference_name")
			serial_no_creator = frappe.get_value(
				"Stock Entry", creation_no, "custom_serial_number_creator"
			)
			snc = frappe.get_value(
				"Serial Number Creator", serial_no_creator, "parent_manufacturing_order"
			)
			ref_customer = frappe.get_value(
				"Parent Manufacturing Order", snc, "ref_customer"
			)
			if not ref_customer:
				sales_order = frappe.get_value(
					"Parent Manufacturing Order", snc, "sales_order"
				)
				ref_customer = frappe.db.get_value(
					"Sales Order", sales_order, "ref_customer"
				)
			exchange_rate = frappe.db.sql(
				"""SELECT exchange_rate FROM `tabCurrency Exchange`
                   WHERE for_selling = 1 ORDER BY modified DESC LIMIT 1""",
				pluck="exchange_rate",
			)
			cctx.reference_customer = ref_customer
			cctx.exchange_rate = exchange_rate[0] if exchange_rate else None
			cctx.billing_currency = frappe.get_value(
				"Customer", ref_customer, "default_currency"
			)

		elif (
			self.company == "Gurukrupa Export Private Limited"
			and ctx.customer_group == "Internal"
		):
			cctx.reference_customer_c2c = frappe.get_value(
				"Sales Order", self.custom_parent_sales_order, "customer"
			)

		_company_ctx_cache[cache_key] = cctx

	cctx = _company_ctx_cache[cache_key]
	if self.company == "KG GK Jewellers Private Limited":
		row.ref_customer = cctx.reference_customer

	return cctx


def _get_making_charge(self, doc, touch, ctx, cctx):
	"""
	Cached — keyed by subcategory too so different BOM types
	don't share each other's making charge rows.
	"""
	# if self.company == "KG GK Jewellers Private Limited":
	#     effective_customer = cctx.reference_customer
	#     # frappe.throw(f"{effective_customer}")
	# elif self.company == "Gurukrupa Export Private Limited" and ctx.customer_group == "Internal":
	#     effective_customer = cctx.reference_customer_c2c
	# else:
	effective_customer = self.customer

	# ← FIX: include item_subcategory in cache key
	cache_key = (
		effective_customer,
		doc.metal_type,
		doc.setting_type,
		self.gold_rate_with_gst,
		touch,
		doc.item_subcategory,  # ← added
	)
	from_site = frappe.db.get_single_value("Data Migration in KGGK", "from_site_1")
	api_key = frappe.db.get_single_value("Data Migration in KGGK", "api_key")
	api_secret = frappe.db.get_single_value("Data Migration in KGGK", "api_secret")

	use_api = bool(from_site)

	if cache_key not in _making_charge_cache:
		if use_api:
			url = f"{from_site}/api/method/gke_customization.gke_order_forms.doc_events.item.get_making_charge"
			headers = {"Authorization": f"token {api_key}:{api_secret}"}
			data = {
				"customer": effective_customer,
				"metal_type": doc.metal_type,
				"setting_type": doc.setting_type,
				"gold_rate": self.gold_rate_with_gst,
				"metal_touch": touch,
				"subcategory": doc.item_subcategory,
			}
			try:
				response = requests.post(
					url,
					headers=headers,
					json=data,
					timeout=30,
				)

				response.raise_for_status()
				response = response.json()

			except requests.exceptions.RequestException as e:
				frappe.throw(f"Failed to fetch Making Charge: {e}")
			if not response or not response.get("message"):
				frappe.throw(
					f"Unable to fetch Making Charge for Customer: {effective_customer}"
				)
			sub_info = response["message"]
			mc_name = sub_info.get("parent")
			threshold = sub_info.get("rate_per_gm_threshold") or 2
			diamond_pcs = doc.total_diamond_pcs or 0
		else:
			filters = {
				"customer": effective_customer,
				"metal_type": doc.metal_type,
				"setting_type": doc.setting_type,
				"from_gold_rate": ["<=", self.gold_rate_with_gst],
				"to_gold_rate": [">=", self.gold_rate_with_gst],
				"metal_touch": touch,
			}
			mc = frappe.get_all(
				"Making Charge Price", filters=filters, fields=["name"], limit=1
			)
			if not mc:
				frappe.throw(
					f"Create a valid Making Charge Price for Customer: {effective_customer}, "
					f"Metal Type: {touch}, Setting Type: {doc.setting_type}"
				)

			mc_name = mc[0]["name"]

			# Fetch ALL subcategory rows for this mc_name upfront
			all_sub_rows = frappe.db.get_all(
				"Making Charge Price Item Subcategory",
				filters={"parent": mc_name},
				fields=[
					"subcategory",
					"rate_per_gm",
					"rate_per_pc",
					"supplier_fg_purchase_rate",
					"wastage",
					"wastage_per_pcs",
					"subcontracting_rate",
					"Subcontracting_wastage",
					"rate_per_gm_threshold",
					"to_diamond",
					"from_diamond",
				],
			)

			# Filter to this BOM's subcategory
			sub_rows = [
				r for r in all_sub_rows if r.subcategory == doc.item_subcategory
			]

			if not sub_rows:
				frappe.throw(
					f"No Making Charge Price subcategory row found for subcategory: "
					f'"{doc.item_subcategory}" in Making Charge Price: {mc_name}'
				)

			sub_info = sub_rows[0]
			threshold = sub_info.get("rate_per_gm_threshold") or 2

			diamond_pcs = doc.total_diamond_pcs or 0
		if doc.metal_and_finding_weight < threshold:
			if sub_info.get("from_diamond") and int(
				sub_info.get("from_diamond")
			) <= int(diamond_pcs) <= int(sub_info.get("to_diamond")):
				pass
		_making_charge_cache[cache_key] = (mc_name, sub_info, threshold)
	# frappe.throw(str(_making_charge_cache[cache_key]))
	return _making_charge_cache[cache_key]


def _get_metal_purity(customer, metal_type, metal_touch):
	key = (customer, metal_type, metal_touch)

	if key not in _metal_purity_cache:
		_metal_purity_cache[key] = frappe.db.get_value(
			"Metal Criteria",
			{"parent": customer, "metal_type": metal_type, "metal_touch": metal_touch},
			"metal_purity",
		)
	return _metal_purity_cache[key]


def _get_calculated_gold_rate(
	customer, metal_type, metal_touch, gold_rate_with_gst, gold_gst_rate
):
	key = (customer, metal_type, metal_touch, gold_rate_with_gst, gold_gst_rate)
	if key not in _gold_rate_cache:
		purity = _get_metal_purity(customer, metal_type, metal_touch)
		_gold_rate_cache[key] = (float(purity) * gold_rate_with_gst) / (
			100 + int(gold_gst_rate)
		)
	return _gold_rate_cache[key]


def get_stock_entry_additional_cost(self, doc):
	creation_docname = getattr(doc, "custom_creation_docname", None)
	if not creation_docname:
		return 0

	stock_entry = frappe.db.get_value(
		"Stock Entry",
		{"custom_serial_number_creator": creation_docname},
		["name", "total_additional_costs"],
		as_dict=True,
	)

	return stock_entry.total_additional_costs if stock_entry else 0


def _process_metal_detail1(self, doc, ctx, cctx):
    if not hasattr(doc, "metal_detail"):
        return
    metal_prec = int(ctx.metal_precision or 3)

    operational_cost = get_stock_entry_additional_cost(self,doc)
    chain_weight = sum(
        r.quantity for r in doc.finding_detail
        if r.finding_category == 'Chains'
    )
    total_weight = doc.metal_and_finding_weight + chain_weight
    for s in doc.metal_detail:
        
        customer_metal_purity = _metal_purity_cache.get(
                (self.customer, s.metal_type, s.metal_touch)
            )
        s.quantity = round(s.quantity, metal_prec)

        if self.company == "Gurukrupa Export Private Limited" and ctx.customer_group == "Internal":
            if s.is_customer_item:
                s.rate           = 0
                s.making_rate=operational_cost/total_weight
                s.wastage_rate   = 0
                s.wastage_amount = 0
            else:
                s.rate= round(s.se_rate,2)
                s.wastage_rate          = 0
                s.making_rate=operational_cost/total_weight
                s.wastage_amount        = 0
                s.customer_metal_purity = customer_metal_purity
            s.amount        = round(s.rate * s.quantity, 2)
            s.making_amount = round(s.making_rate * s.quantity, 2)

        elif self.company == "KG GK Jewellers Private Limited" and ctx.customer_group == "Internal":
            _, sub_info, threshold = _get_making_charge(self, doc, s.metal_touch, ctx, cctx)
            # frappe.throw(str(sub_info))
            calculated_gold_rate  = _get_calculated_gold_rate(
                self.customer, s.metal_type, s.metal_touch,
                self.gold_rate_with_gst, ctx.gold_gst_rate,
            )
            if s.is_customer_item:
                s.rate          = 0
                s.amount = 0
                # s.making_rate=operational_cost/total_weight
                s.making_rate = sub_info.get("rate_per_gm", 0)
                s.making_amount = round(s.making_rate * s.quantity, 2)
                s.wastage_rate   = 0
                s.wastage_amount = 0
                s.fg_purchase_rate = 0
                s.fg_purchase_amount = 0
            else:
                if cctx.billing_currency == "USD":
                    s.se_rate     = s.se_rate * cctx.exchange_rate
                    # s.making_rate=(operational_cost/total_weight)*cctx.exchange_rate
                    s.making_rate = sub_info.get("rate_per_gm", 0)*cctx.exchange_rate
                else:
                    # s.making_rate=operational_cost/total_weight
                    s.making_rate = sub_info.get("rate_per_gm", 0)

                # s.rate           = s.se_rate
                s.rate = round(calculated_gold_rate, 2)
                s.wastage_rate   = 0
                s.wastage_amount = 0
                s.making_amount  = round(s.making_rate * s.quantity, 2)
            s.amount = round(s.rate * s.quantity, 2)

        else:
            _, sub_info, threshold = _get_making_charge(self, doc, s.metal_touch, ctx, cctx)
            # frappe.throw(str(sub_info))
            calculated_gold_rate  = _get_calculated_gold_rate(
                self.customer, s.metal_type, s.metal_touch,
                self.gold_rate_with_gst, ctx.gold_gst_rate,
            )
            
            if doc.metal_and_finding_weight < threshold:
                making_rate        = sub_info.get("rate_per_pc", 0)
                wastage_rate_value = sub_info.get("wastage_per_pcs", 0) / 100.0
            else:
                making_rate        = sub_info.get("rate_per_gm", 0)
                wastage_rate_value = sub_info.get("wastage", 0) / 100.0


            if s.is_customer_item:
                s.rate          = 0
                s.amount        = 0
                s.making_rate   = sub_info.get("subcontracting_rate", 0)
                s.making_amount = s.making_rate * s.quantity
            else:
                wastage = (
                    sub_info.get("subcontracting_wastage", 0) / 100.0
                    if getattr(doc, "is_customer_item", False)
                    else wastage_rate_value
                )
                s.customer_metal_purity = customer_metal_purity
                s.rate                  = round(calculated_gold_rate, 2)
                s.amount                = round(s.rate * s.quantity, 2)
                s.making_rate           = making_rate
                s.making_amount         = (
                    s.making_rate if doc.metal_and_finding_weight < threshold
                    else s.making_rate * s.quantity
                )
                s.wastage_rate   = wastage
                s.wastage_amount = (
                    s.wastage_rate * s.amount
                    if self.customer != "TNCU0101"
                    else s.wastage_rate * s.quantity * self.gold_rate
                )
                # frappe.throw(str(s.making_rate))

    doc.total_metal_amount   = sum(flt(r.amount)        for r in doc.get("metal_detail",   []))
    doc.total_wastage_amount = sum(flt(r.wastage_amount) for r in doc.get("metal_detail",   []))
    # frappe.msgprint(f"it7t6{doc.total_wastage_amount},{sum(flt(r.wastage_amount) for r in doc.get("finding_detail",   []))}")
    doc.total_making_amount  = sum(flt(r.making_amount)  for r in doc.get("metal_detail",   []))


_finding_sub_cache = {}


def _get_all_finding_sub_rows(mc_name):
	if mc_name not in _finding_sub_cache:
		_finding_sub_cache[mc_name] = frappe.db.get_all(
			"Making Charge Price Finding Subcategory",
			filters={"parent": mc_name},
			fields=[
				"subcategory",
				"rate_per_gm",
				"rate_per_pc",
				"wastage",
				"wastage_per_pcs",
				"supplier_fg_purchase_rate",
				"subcontracting_rate",
				"subcontracting_wastage",
			],
		)
	return _finding_sub_cache[mc_name]


def _get_finding_sub_info(mc_name, finding_type, doc):
	"""
	Returns find_data for a given finding_type, using the
	fetch-all cache. Falls back to item_subcategory rows
	(also from cache) if no direct match.
	"""
	from_site = frappe.db.get_single_value("Data Migration in KGGK", "from_site_1")
	api_key = frappe.db.get_single_value("Data Migration in KGGK", "api_key")
	api_secret = frappe.db.get_single_value("Data Migration in KGGK", "api_secret")
	use_api = bool(from_site)
	if use_api:
		url = f"{from_site}/api/method/gke_customization.gke_order_forms.doc_events.item.get_finding_charge"
		headers = {"Authorization": f"token {api_key}:{api_secret}"}
		data = {"parent": mc_name, "subcategory": finding_type}
		response = requests.post(url=url, headers=headers, data=data)
		result = response.json()
		find_data = result.get("message") if response else None
		if find_data:
			return find_data

		else:
			url = f"{from_site}/api/method/gke_customization.gke_order_forms.doc_events.item.get_making_charge_price"
			headers = {"Authorization": f"token {api_key}:{api_secret}"}
			payload = {"parent": mc_name, "subcategory": doc.item_subcategory}
			response = requests.post(url, headers=headers, data=payload)
			result = response.json()
			find_data = result.get("message") if response else None
			# frappe.throw(str(find_data))
		f_threshold = find_data.get("rate_per_gm_threshold") or 2
		diamond_pcs = doc.total_diamond_pcs or 0

		return find_data
	else:
		all_finding_rows = _get_all_finding_sub_rows(mc_name)

		# Try direct match by finding_type (subcategory)
		match = next(
			(r for r in all_finding_rows if r.subcategory == finding_type), None
		)
		if match:
			return match

		# Fall back: use item_subcategory rows from Making Charge Price Item Subcategory
		# (already fetched & cached inside _get_making_charge via all_sub_rows)
		# Re-fetch from DB only if not already in the making charge cache path.
		all_item_sub = frappe.db.get_all(
			"Making Charge Price Item Subcategory",
			filters={"parent": mc_name},
			fields=[
				"subcategory",
				"rate_per_gm",
				"rate_per_pc",
				"supplier_fg_purchase_rate",
				"wastage",
				"wastage_per_pcs",
				"subcontracting_rate",
				"subcontracting_wastage",
				"name",
				"to_diamond",
				"from_diamond",
				"rate_per_gm_threshold",
			],
		)
		sub_rows = [r for r in all_item_sub if r.subcategory == doc.item_subcategory]
		find_data = sub_rows[0] if sub_rows else frappe._dict()
		f_threshold = find_data.get("rate_per_gm_threshold") or 2
		diamond_pcs = doc.total_diamond_pcs or 0

		if doc.metal_and_finding_weight < f_threshold:
			for sf_row in sub_rows:
				if sf_row.from_diamond and (
					int(sf_row.from_diamond)
					<= int(diamond_pcs)
					<= int(sf_row.to_diamond)
				):
					find_data = sf_row
					break

		return find_data


def _process_finding_detail1(self, doc, ctx, cctx):
	if not hasattr(doc, "finding_detail") or not doc.finding_detail:
		doc.total_finding_amount = 0
		return

	finding_cache = {}  # local per-BOM-doc: finding_type → find_data
	f_metal_prec = int(ctx.metal_precision or 3)
	operational_cost = get_stock_entry_additional_cost(self, doc)
	chain_weight = sum(
		r.quantity for r in doc.finding_detail if r.finding_category == "Chains"
	)
	total_weight = doc.metal_and_finding_weight + chain_weight
	for f in doc.finding_detail:
		# frappe.msgprint("jhgh")
		customer_metal_purity = _metal_purity_cache.get(
			(self.customer, f.metal_type, f.metal_touch)
		)

		f.customer_metal_purity = customer_metal_purity
		mc_name, sub_info, threshold = _get_making_charge(self, doc, f.metal_touch, ctx, cctx)

		finding_type = f.finding_type
		if finding_type not in finding_cache:
			finding_cache[finding_type] = _get_finding_sub_info(mc_name, finding_type, doc)

		find_data = finding_cache[finding_type]

		f.customer_metal_purity = customer_metal_purity
		f.quantity              = round(f.quantity, f_metal_prec)

		# if f.is_customer_item:
		# 	f.rate = 0
		# 	f.amount = 0
		# 	f.making_rate    = find_data.get("subcontracting_rate")
		# 	f.wastage_rate = 0
		# 	f.wastage_amount = 0
		# 	f.making_amount = round(f.making_rate * f.quantity, 2)

		if (
			self.company == "Gurukrupa Export Private Limited"
			and ctx.customer_group == "Internal"
		):
			if f.is_customer_item:
				f.rate = 0
				f.amount = 0
				f.making_rate    = find_data.get("rate_per_gm")
				if self.sales_type =='Hybrid' and f.finding_category=='Chains':
					f.making_rate     = 0
				f.wastage_rate = 0
				f.wastage_amount = 0
				f.making_amount = round(f.making_rate * f.quantity, 2)
			#  f.rate           = round(calculated_gold_rate, 2)
			else:
				f.rate = round(f.se_rate, 2)
				f.amount = round(f.rate * f.quantity, 2)
				# f.making_rate    = 550 if f.finding_category != 'Chains' else 200
				f.making_rate = operational_cost / total_weight
				f.wastage_rate = 0
				f.wastage_amount = 0
				f.making_amount = round(f.making_rate * f.quantity, 2)

		elif (
			self.company == "KG GK Jewellers Private Limited"
			and ctx.customer_group == "Internal"
		):
			mc_name, sub_info, threshold = _get_making_charge(self, doc, f.metal_touch, ctx, cctx)

			finding_type = f.finding_type
			if finding_type not in finding_cache:
				finding_cache[finding_type] = _get_finding_sub_info(mc_name, finding_type, doc)
                
			find_data = finding_cache[finding_type]
			if f.is_customer_item:
				f.rate          = 0
				f.amount = 0
				# f.making_rate=operational_cost/total_weight
				f.making_rate     = find_data.get("rate_per_gm")
				if self.sales_type =='Hybrid' and f.finding_category=='Chains':
					f.making_rate     = 0
				f.making_amount = round(f.making_rate * f.quantity, 2)
				f.wastage_rate   = 0
				f.wastage_amount = 0
				f.fg_purchase_rate = 0
				f.fg_purchase_amount = 0
			else:
				if cctx.billing_currency == "USD":
					f.se_rate = f.se_rate * cctx.exchange_rate
				else:
					# f.making_rate = operational_cost / total_weight
					f.making_rate     = find_data.get("rate_per_gm", 0)
					
				calculated_gold_rate  = _get_calculated_gold_rate(self.customer, f.metal_type, f.metal_touch,self.gold_rate_with_gst, ctx.gold_gst_rate,)
				f.rate=round(calculated_gold_rate, 2)
				# f.rate = round(f.se_rate, 2)
				f.amount = round(f.rate * f.quantity, 2)
				f.wastage_rate = 0
				f.wastage_amount = 0
				f.making_amount = round(f.making_rate * f.quantity, 2)
				f.fg_purchase_rate = 0
				f.fg_purchase_amount = 0

		else:
			mc_name, sub_info, threshold = _get_making_charge(
				self, doc, f.metal_touch, ctx, cctx
			)

			finding_type = f.finding_type
			if finding_type not in finding_cache:
				finding_cache[finding_type] = _get_finding_sub_info(
					mc_name, finding_type, doc
				)

			find_data = finding_cache[finding_type]

			calculated_gold_rate = _get_calculated_gold_rate(
				self.customer,
				f.metal_type,
				f.metal_touch,
				self.gold_rate_with_gst,
				ctx.gold_gst_rate,
			)
			if f.is_customer_item:
				f.rate = 0
				f.amount = 0
				f.making_rate    = find_data.get("rate_per_gm")
				if self.sales_type =='Hybrid' and f.finding_category=='Chains':
					f.making_rate     = 0
				f.wastage_rate = 0
				f.wastage_amount = 0
				f.making_amount = round(f.making_rate * f.quantity, 2)
			else:
				f.rate = round(calculated_gold_rate, 2)
				f.amount = round(f.rate * f.quantity, 2)
				finding_weight = getattr(doc, "metal_and_finding_weight", None)
				if finding_weight is not None and finding_weight < threshold:
					making_rate = find_data.get("rate_per_pc", 0)
					wastage_rate = find_data.get("wastage_per_pcs", 0) / 100.0
					f.making_amount = making_rate
				else:
					making_rate = find_data.get("rate_per_gm", 0)
					wastage_rate = find_data.get("wastage", 0) / 100.0
					f.making_amount = making_rate * f.quantity
				f.making_rate = making_rate
				# frappe.msgprint(f"kjh{f.making_amount}")
				f.wastage_rate = wastage_rate
				f.wastage_amount = (
					f.wastage_rate * f.amount
					if self.customer != "TNCU0101"
					else f.wastage_rate * f.quantity * self.gold_rate
				)

	doc.total_finding_amount = sum(flt(r.amount) for r in doc.get("finding_detail", []))

	doc.total_finding_making_amount = sum(
		flt(r.making_amount) for r in doc.get("finding_detail", [])
	)
	doc.total_finding_wastage_amount = sum(
		flt(r.wastage_amount) for r in doc.get("finding_detail", [])
	)


# Condition if  Data Migration in KGGK doc from site and to site is there
def _process_gemstone_detail(self, doc, ctx, cctx):
	if not hasattr(doc, "gemstone_detail"):
		return

	if self.customer not in _gemstone_pl_cache:
		_gemstone_pl_cache[self.customer] = frappe.db.get_value(
			"Customer", self.customer, "custom_gemstone_price_list_type"
		)
	gemstone_price_list_customer = _gemstone_pl_cache[self.customer]

	from_site = frappe.db.get_single_value("Data Migration in KGGK", "from_site_1")
	api_key = frappe.db.get_single_value("Data Migration in KGGK", "api_key")
	api_secret = frappe.db.get_single_value("Data Migration in KGGK", "api_secret")
	use_api = bool(from_site)

	HEADERS = {"Authorization": f"token {api_key}:{api_secret}"}

	stone_prec = int(ctx.stone_precision or 3)
	for gem in doc.gemstone_detail:
		gem.quantity = round(gem.quantity, stone_prec)

		if (
			self.company == "Gurukrupa Export Private Limited"
			and ctx.customer_group == "Internal"
		):
			gem.total_gemstone_rate = round(gem.se_rate , 2)

			gem.gemstone_rate_for_specified_quantity = round(
				float(gem.total_gemstone_rate) / 100 * float(gem.quantity) , 2
			)

		elif (
			self.company == "KG GK Jewellers Private Limited"
			and ctx.customer_group == "Internal"
		):
			if gem.is_customer_item:
				gem.total_gemstone_rate = 0
				gem.fg_purchase_rate = 0
				gem.gemstone_rate_for_specified_quantity = 0
				gem.fg_purchase_amount = 0
				gem.se_rate = 0
			
			else:
				gem.total_gemstone_rate = round(
					gem.se_rate * cctx.exchange_rate
					if cctx.billing_currency == "USD"
					else gem.se_rate, 2
				)
				gem.gemstone_rate_for_specified_quantity = round(
					float(gem.total_gemstone_rate) * float(gem.quantity)
					if gem.per_pc_or_per_carat == "Per Carat"
					else float(gem.total_gemstone_rate) * float(gem.pcs) , 2
				)
				# gem.total_gemstone_rate = 0
				gem.fg_purchase_rate = 0
				gem.fg_purchase_amount = 0

		# ---------------- Fixed ----------------
		elif gemstone_price_list_customer == "Fixed" and ctx.customer_group != "Retail":
			if use_api:
				response = requests.post(
					url=f"{from_site}/api/method/gke_customization.gke_order_forms.doc_events.item.get_fixed_retail_rate",
					headers=HEADERS,
					data={
						"customer": self.customer,
						"price_list_type": gemstone_price_list_customer,
						"per_pc_or_per_carat": gem.get("per_pc_or_per_carat"),
						"cut_or_cab": gem.get("cut_or_cab"),
						"gemstone_type": gem.get("gemstone_type"),
						"stone_shape": gem.get("stone_shape"),
						"gemstone_grade": gem.get("gemstone_grade"),
					},
				)
				gpc = response.json().get("message", {})
			else:
				gpc = frappe.get_all(
					"Gemstone Price List",
					filters={
						"customer": self.customer,
						"price_list_type": gemstone_price_list_customer,
						"per_pc_or_per_carat": gem.get("per_pc_or_per_carat"),
						"cut_or_cab": gem.get("cut_or_cab"),
						"gemstone_type": gem.get("gemstone_type"),
						"stone_shape": gem.get("stone_shape"),
						"gemstone_grade": gem.get("gemstone_grade"),
					},
					fields=[
						"name",
						"price_list_type",
						"rate",
						"handling_rate",
						"outwork_handling_charges_rate",
					],
				)
			if not gpc:
				gem.gemstone_rate_for_specified_quantity = 0
				gem.total_gemstone_rate = 0
				frappe.msgprint(
					f'No Gemstone Price List found: {gem.get("per_pc_or_per_carat")}, '
					f'{gem.get("cut_or_cab")}, {gem.get("gemstone_type")}, {gem.get("stone_shape")}'
				)

		# ---------------- Retail ----------------
		elif ctx.customer_group == "Retail":
			gpc = frappe.get_all(
				"Gemstone Price List",
				filters={
					"is_retail_customer": 1,
					"price_list_type": gemstone_price_list_customer,
					"per_pc_or_per_carat": gem.get("per_pc_or_per_carat"),
					"cut_or_cab": gem.get("cut_or_cab"),
					"gemstone_type": gem.get("gemstone_type"),
					"stone_shape": gem.get("stone_shape"),
				},
				fields=[
					"name",
					"price_list_type",
					"rate",
					"handling_rate",
					"outwork_handling_charges_rate",
				],
			)
			if not gpc:
				frappe.throw("No Retail Gemstone Price List found")
			rate = (
				gpc[0]["outwork_handling_charges_rate"]
				if gem.is_customer_item
				else gpc[0]["rate"] or 0
			)
			gem.total_gemstone_rate = round(rate, 2)
			gem.gemstone_rate_for_specified_quantity = round(
				float(rate) * float(gem.quantity)
				if gem.per_pc_or_per_carat == "Per Carat"
				else float(rate) * float(gem.pcs),
				2,
			)

		# ---------------- Diamond Range ----------------
		elif (
			gemstone_price_list_customer == "Diamond Range"
			and ctx.customer_group != "Retail"
		):
			if use_api:
				response = requests.post(
					url=f"{from_site}/api/method/gke_customization.gke_order_forms.doc_events.item.get_diamond_range_rate",
					headers=HEADERS,
					data={
						"customer": self.customer,
						"cut_or_cab": gem.get("cut_or_cab"),
						"gemstone_grade": gem.get("gemstone_grade"),
					},
				)
				result = response.json().get("message", {})
				if not result:
					frappe.msgprint("No Multiplier Price List found")
				else:
					gpc_doc = result.get("doc", {})
					rate = 0
					quality_map = {
						"Precious": (
							"outwork_precious_percentage",
							"precious_percentage",
						),
						"Semi-Precious": (
							"outwork_semi_precious_percentage",
							"semi_precious_percentage",
						),
						"Synthetic": (
							"outwork_synthetic_percentage",
							"synthetic_percentage",
						),
					}
					for mul in gpc_doc.get("gemstone_multiplier", []):
						if (
							mul.get("gemstone_type") == gem.gemstone_type
							and flt(doc.diamond_weight) >= flt(mul.get("from_weight"))
							and flt(doc.diamond_weight) <= flt(mul.get("to_weight"))
						):
							outwork_field, standard_field = quality_map.get(
								gem.gemstone_quality, (None, None)
							)
							rate = (
								mul.get(outwork_field, 0)
								if gem.is_customer_item
								else mul.get(standard_field, 0)
							)
							gem.total_gemstone_rate = round(rate, 2)
							gem.gemstone_rate_for_specified_quantity = round(
								float(rate) * float(gem.gemstone_pr)
								if mul.get("is_rate")
								else float(rate) / 100 * float(gem.gemstone_pr),
								2,
							)
							break
			else:
				gpc = frappe.get_all(
					"Gemstone Price List",
					filters={
						"customer": self.customer,
						"price_list_type": gemstone_price_list_customer,
						"cut_or_cab": gem.get("cut_or_cab"),
						"gemstone_grade": gem.get("gemstone_grade"),
					},
					fields=["name", "price_list_type"],
				)
				if not gpc:
					frappe.msgprint("No Multiplier Price List found")
				else:
					gpc_doc = frappe.get_doc("Gemstone Price List", gpc[0].name)
					rate = 0
					quality_map = {
						"Precious": (
							"outwork_precious_percentage",
							"precious_percentage",
						),
						"Semi-Precious": (
							"outwork_semi_precious_percentage",
							"semi_precious_percentage",
						),
						"Synthetic": (
							"outwork_synthetic_percentage",
							"synthetic_percentage",
						),
					}
					for mul in gpc_doc.get("gemstone_multiplier", []):
						if mul.gemstone_type == gem.gemstone_type and (
							flt(doc.diamond_weight) >= flt(mul.from_weight)
							and flt(doc.diamond_weight) <= flt(mul.to_weight)
						):
							outwork_field, standard_field = quality_map.get(
								gem.gemstone_quality, (None, None)
							)
							rate = (
								getattr(mul, outwork_field, 0)
								if gem.is_customer_item
								else getattr(mul, standard_field, 0)
							)
						gem.total_gemstone_rate = round(rate, 2)
						gem.gemstone_rate_for_specified_quantity = round(
							float(rate) * float(gem.gemstone_pr)
							if mul.is_rate
							else float(rate) / 100 * float(gem.gemstone_pr),
							2,
						)

			gem.price_list_type = "Diamond Range"

	doc.total_gemstone_amount = sum(
		flt(r.gemstone_rate_for_specified_quantity)
		for r in doc.get("gemstone_detail", [])
	)


# Condition if  Data Migration in KGGK doc from site and to site is there
# def _process_diamond_detail(self, doc, ctx, row, cctx):
#     if not hasattr(doc, "diamond_detail"):
#         return

#     from_site = frappe.db.get_single_value("Data Migration in KGGK", "from_site_1")
#     api_key = frappe.db.get_single_value("Data Migration in KGGK", "api_key")
#     api_secret = frappe.db.get_single_value("Data Migration in KGGK", "api_secret")
#     use_api   = bool(from_site)

#     customer_key = (
#         cctx.reference_customer
#         if self.company == "KG GK Jewellers Private Limited" and ctx.customer_group == "Internal"
#         else self.customer
#     )
#     if self.custom_diamond_quality:
#         row.diamond_quality = self.custom_diamond_quality
#     stone_prec = int(ctx.stone_precision or 3)

#     _DIAMOND_RATE_URL     = f"{from_site}/api/method/gke_customization.gke_order_forms.doc_events.item.get_diamond_rate"
#     _DIAMOND_RATE_HEADERS = {"Authorization": f"token {api_key}:{api_secret}"}

#     for d in doc.diamond_detail:
#         d.quality        = row.diamond_quality
#         d.quantity       = round(d.quantity, stone_prec)
#         d.weight_per_pcs = d.quantity / d.pcs
#         if 0.001 < d.weight_per_pcs > 0.005:
#             d.weight_per_pcs = round(d.weight_per_pcs, 3)

#         pl_result = frappe.db.sql(
#             """SELECT diamond_price_list FROM `tabDiamond Price List Table`
#                 WHERE parent = %s AND diamond_shape = %s""",
#             (customer_key, d.stone_shape), as_dict=True,
#         )
#         if not pl_result:
#             d.total_diamond_rate                  = 0
#             d.diamond_rate_for_specified_quantity = 0
#             continue

#         price_list_type = pl_result[0]["diamond_price_list"]

#         # -------- fetch latest rate --------
#         if use_api:
#             try:
#                 response = requests.post(
#                     url=_DIAMOND_RATE_URL,
#                     headers=_DIAMOND_RATE_HEADERS,
#                     data={
#                         "customer":          customer_key,
#                         "diamond_type":      d.diamond_type,
#                         "stone_shape":       d.stone_shape,
#                         "diamond_quality":   d.quality,
#                         "price_list_type":   price_list_type,
#                         "sieve_size_range":  d.sieve_size_range   if price_list_type == "Sieve Size Range"  else None,
#                         "weight_per_pcs":    d.weight_per_pcs     if price_list_type == "Weight (in cts)"   else None,
#                         "diamond_size_in_mm": d.diamond_sieve_size if price_list_type == "Size (in mm)"     else None,
#                     },
#                 )
#                 latest = response.json().get("message", {})
#             except Exception as e:
#                 frappe.log_error(frappe.get_traceback(), f"Diamond Rate API Error: {e}")
#                 latest = {}

#         else:
#             common_filters = {
#                 "price_list":      "Standard Selling",
#                 "price_list_type": price_list_type,
#                 "customer":        customer_key,
#                 "diamond_type":    d.diamond_type,
#                 "stone_shape":     d.stone_shape,
#                 "diamond_quality": d.quality,
#             }
#             fields = [
#                 "rate", "outright_handling_charges_rate",
#                 "outright_handling_charges_in_percentage",
#                 "outwork_handling_charges_rate",
#                 "outwork_handling_charges_in_percentage",
#                 "supplier_fg_purchase_rate",
#             ]

#             if price_list_type == "Sieve Size Range":
#                 latest = frappe.db.get_value(
#                     "Diamond Price List",
#                     {**common_filters, "sieve_size_range": d.sieve_size_range},
#                     fields, as_dict=True,
#                 )
#             elif price_list_type == "Weight (in cts)":
#                 conds = " AND ".join(f"{k} = %s" for k in common_filters)
#                 rows  = frappe.db.sql(
#                     f"""SELECT {", ".join(fields)} FROM `tabDiamond Price List`
#                         WHERE {conds} AND %s BETWEEN from_weight AND to_weight LIMIT 1""",
#                     list(common_filters.values()) + [d.weight_per_pcs], as_dict=True,
#                 )
#                 latest = rows[0] if rows else None
#             elif price_list_type == "Size (in mm)":
#                 latest = frappe.db.get_value(
#                     "Diamond Price List",
#                     {**common_filters, "diamond_size_in_mm": d.diamond_sieve_size},
#                     fields, as_dict=True,
#                 )
#             else:
#                 latest = None

#         # -------- no rate found --------
#         if not(self.company == "KG GK Jewellers Private Limited" or ctx.customer_group == "Internal"):
#             if not latest :
#                 frappe.msgprint(f'No Diamond Pricelist for {d.quality}, {d.weight_per_pcs}')
#                 d.total_diamond_rate                  = 0
#                 d.diamond_rate_for_specified_quantity = 0
#                 continue
#         if self.company == "KG GK Jewellers Private Limited" and ctx.customer_group == "Internal":
#             if cctx.billing_currency == "USD":
#                 d.se_rate = d.se_rate * cctx.exchange_rate
#             d.total_diamond_rate = d.se_rate * 1.15
#             if d.quantity > 0.005:
#                 d.quantity = round(d.quantity, stone_prec)
#             d.diamond_rate_for_specified_quantity = round(
#                 d.quantity * (d.handling_rate + d.se_rate), 2
#             )

#         elif self.company == "Gurukrupa Export Private Limited" and ctx.customer_group == "Internal":
#             # d.fg_purchase_rate   = latest.get("supplier_fg_purchase_rate", 0)
#             d.total_diamond_rate = d.se_rate
#             # frappe.throw(f"{ d.total_diamond_rate}")
#             d.quantity           = round(d.quantity, stone_prec)
#             d.weight_per_pcs     = d.quantity / d.pcs
#             d.quantity_3         = round(d.quantity, 2)
#             d.diamond_rate_for_specified_quantity = round(d.quantity * d.total_diamond_rate, 2)

#         # -------- extract rate fields --------
#         base_rate = latest.get("rate", 0)
#         out_rate  = latest.get("outright_handling_charges_rate", 0)
#         out_pct   = latest.get("outright_handling_charges_in_percentage", 0)
#         work_rate = latest.get("outwork_handling_charges_rate", 0)
#         work_pct  = latest.get("outwork_handling_charges_in_percentage", 0)
#         is_cust   = getattr(d, "is_customer_item", False)

#         d.handling_rate = (
#             work_rate or (base_rate * (work_pct / 100))
#             if is_cust
#             else out_rate or (base_rate + base_rate * (out_pct / 100)) if out_pct else 0
#         )

#         # -------- apply rate by company / customer_group --------
#         if self.company == "KG GK Jewellers Private Limited" and ctx.customer_group == "Internal":
#             if cctx.billing_currency == "USD":
#                 d.se_rate = d.se_rate * cctx.exchange_rate
#             d.total_diamond_rate = d.se_rate
#             if d.quantity > 0.005:
#                 d.quantity = round(d.quantity, stone_prec)
#             d.diamond_rate_for_specified_quantity = round(
#                 d.quantity * (d.handling_rate + d.se_rate), 2
#             )

#         elif self.company == "Gurukrupa Export Private Limited" and ctx.customer_group == "Internal":
#             d.fg_purchase_rate   = latest.get("supplier_fg_purchase_rate", 0)
#             d.total_diamond_rate = d.fg_purchase_rate
#             d.quantity           = round(d.quantity, stone_prec)
#             d.weight_per_pcs     = d.quantity / d.pcs
#             d.quantity_3         = round(d.quantity, 2)
#             d.diamond_rate_for_specified_quantity = round(d.quantity * d.total_diamond_rate, 2)

#         else:
#             d.total_diamond_rate = round(base_rate, 2)
#             d.quantity_3         = round(d.quantity, 2)
#             d.diamond_rate_for_specified_quantity = round(
#                 d.quantity * (d.handling_rate + d.total_diamond_rate), 2
#             )
#             d.weight_per_pcs = d.quantity / d.pcs
#             if 0.001 < d.weight_per_pcs > 0.005:
#                 d.weight_per_pcs = round(d.quantity / d.pcs, 3)

#         doc.total_diamond_amount = sum(
#             flt(r.diamond_rate_for_specified_quantity) for r in doc.get("diamond_detail", [])
#         )
#         doc.diamond_bom_amount = doc.total_diamond_amount
def _process_diamond_detail(self, doc, ctx, row, cctx):
    if not hasattr(doc, "diamond_detail"):
        doc.total_diamond_amount=0
        return

    from_site = frappe.db.get_single_value("Data Migration in KGGK", "from_site_1")
    api_key = frappe.db.get_single_value("Data Migration in KGGK", "api_key")
    api_secret = frappe.db.get_single_value("Data Migration in KGGK", "api_secret")
    use_api   = bool(from_site)

    # customer_key = (
    #     cctx.customer_key
    #     if self.company == "KG GK Jewellers Private Limited" and ctx.customer_group == "Internal"
    #     else self.customer
    # )
    customer_key = self.customer
    if self.custom_diamond_quality:
        row.diamond_quality = self.custom_diamond_quality
    stone_prec = int(ctx.stone_precision or 3)
    diamond_additional_cost = frappe.db.get_single_value("Jewellery Settings","inter_company_diamond_additional_cost")
    _DIAMOND_RATE_URL     = f"{from_site}/api/method/gke_customization.gke_order_forms.doc_events.item.get_diamond_rate"
    _DIAMOND_RATE_HEADERS = {"Authorization": f"token {api_key}:{api_secret}"}

    for d in doc.diamond_detail:
        d.quality        = row.diamond_quality
        d.quantity       = round(d.quantity, stone_prec)
        d.weight_per_pcs = d.quantity / d.pcs if d.pcs else d.quantity
        if 0.001 < d.weight_per_pcs > 0.005:
            d.weight_per_pcs = round(d.weight_per_pcs, 3)

        pl_result = frappe.db.sql(
            """SELECT diamond_price_list FROM `tabDiamond Price List Table`
                WHERE parent = %s AND diamond_shape = %s""",
            (customer_key, d.stone_shape), as_dict=True,
        )
        if not pl_result:
            d.total_diamond_rate                  = 0
            d.diamond_rate_for_specified_quantity = 0
            continue

        price_list_type = pl_result[0]["diamond_price_list"]

        # -------- fetch latest rate --------
        if use_api:
            try:
                response = requests.post(
                    url=_DIAMOND_RATE_URL,
                    headers=_DIAMOND_RATE_HEADERS,
                    data={
                        "customer":          customer_key,
                        "diamond_type":      d.diamond_type,
                        "stone_shape":       d.stone_shape,
                        "diamond_quality":   d.quality,
                        "price_list_type":   price_list_type,
                        "sieve_size_range":  d.sieve_size_range   if price_list_type == "Sieve Size Range"  else None,
                        "weight_per_pcs":    d.weight_per_pcs     if price_list_type == "Weight (in cts)"   else None,
                        "diamond_size_in_mm": d.diamond_sieve_size if price_list_type == "Size (in mm)"     else None,
                    },
                )
                latest = response.json().get("message", {})
            except Exception as e:
                frappe.log_error(frappe.get_traceback(), f"Diamond Rate API Error: {e}")
                latest = {}

        else:
            common_filters = {
                "price_list":      "Standard Selling",
                "price_list_type": price_list_type,
                "customer":        customer_key,
                "diamond_type":    d.diamond_type,
                "stone_shape":     d.stone_shape,
            }
            if self.company != "KG GK Jewellers Private Limited":
                common_filters["diamond_quality"] = d.quality
            fields = [
                "rate", "outright_handling_charges_rate",
                "outright_handling_charges_in_percentage",
                "outwork_handling_charges_rate",
                "outwork_handling_charges_in_percentage",
                "supplier_fg_purchase_rate",
            ]

            if price_list_type == "Sieve Size Range":
                latest = frappe.db.get_value(
                    "Diamond Price List",
                    {**common_filters, "sieve_size_range": d.sieve_size_range},
                    fields, as_dict=True,
                )
            elif price_list_type == "Weight (in cts)":
                conds = " AND ".join(f"{k} = %s" for k in common_filters)
                rows  = frappe.db.sql(
                    f"""SELECT {", ".join(fields)} FROM `tabDiamond Price List`
                        WHERE {conds} AND %s BETWEEN from_weight AND to_weight LIMIT 1""",
                    list(common_filters.values()) + [d.weight_per_pcs], as_dict=True,
                )
                latest = rows[0] if rows else None
            elif price_list_type == "Size (in mm)":
                latest = frappe.db.get_value(
                    "Diamond Price List",
                    {**common_filters, "diamond_size_in_mm": d.diamond_sieve_size},
                    fields, as_dict=True,
                )
            else:
                latest = None

        # -------- no rate found --------
        if not(self.company == "KG GK Jewellers Private Limited" or ctx.customer_group == "Internal"):
            if not latest :
                frappe.msgprint(f'No Diamond Pricelist for {d.quality}, {d.weight_per_pcs}')
                d.total_diamond_rate                  = 0
                d.diamond_rate_for_specified_quantity = 0
                doc.total_diamond_amount = sum(
                flt(r.diamond_rate_for_specified_quantity) for r in doc.get("diamond_detail", [])
            )
                continue
        if self.company == "KG GK Jewellers Private Limited" and ctx.customer_group == "Internal":
            if d.is_customer_item:
                d.fg_purchase_rate   = 0
                d.fg_purchase_amount = 0
                # d.handling_rate   = latest.get("supplier_fg_purchase_rate", 0)
                # d.diamond_rate_for_specified_quantity = round(
                #     d.quantity * (d.handling_rate), 2
                # )
                d.handling_rate   = 0
                d.diamond_rate_for_specified_quantity = 0
                d.total_diamond_rate = 0
            else:
                if cctx.billing_currency == "USD":
                    d.se_rate = d.se_rate * cctx.exchange_rate
                # d.total_diamond_rate = d.se_rate * 1.15
                d.total_diamond_rate = d.se_rate * (1 + (int(diamond_additional_cost)/100))
                if d.quantity > 0.005:
                    d.quantity = round(d.quantity, stone_prec)
                d.diamond_rate_for_specified_quantity = round(
                    d.quantity * (d.handling_rate + d.total_diamond_rate), 2
                )
        
        elif self.company == "Gurukrupa Export Private Limited" and ctx.customer_group == "Internal":
            # d.fg_purchase_rate   = latest.get("supplier_fg_purchase_rate", 0)
            d.total_diamond_rate = d.se_rate
            # frappe.throw(f"{ d.total_diamond_rate}")
            d.quantity           = round(d.quantity, stone_prec)
            d.weight_per_pcs     = d.quantity / d.pcs
            d.quantity_3         = round(d.quantity, 2)
            d.diamond_rate_for_specified_quantity = round(d.quantity * d.total_diamond_rate, 2)
        # -------- extract rate fields --------
        

        # -------- apply rate by company / customer_group --------
        

        else:
            base_rate = latest.get("rate", 0)
            out_rate  = latest.get("outright_handling_charges_rate", 0)
            out_pct   = latest.get("outright_handling_charges_in_percentage", 0)
            work_rate = latest.get("outwork_handling_charges_rate", 0)
            work_pct  = latest.get("outwork_handling_charges_in_percentage", 0)
            is_cust   = getattr(d, "is_customer_item", False)

            d.handling_rate = (
                work_rate or (base_rate * (work_pct / 100))
                if is_cust
                else out_rate or (base_rate + base_rate * (out_pct / 100)) if out_pct else 0
            )
            d.total_diamond_rate = round(base_rate, 2)
            d.quantity_3         = round(d.quantity, 2)
            d.diamond_rate_for_specified_quantity = round(
                d.quantity * (d.handling_rate + d.total_diamond_rate), 2
            )
            d.weight_per_pcs = d.quantity / d.pcs
            if 0.001 < d.weight_per_pcs > 0.005:
                d.weight_per_pcs = round(d.quantity / d.pcs, 3)

        doc.total_diamond_amount = sum(
            flt(r.diamond_rate_for_specified_quantity) for r in doc.get("diamond_detail", [])
        )
        doc.diamond_bom_amount = doc.total_diamond_amount

def _reconcile_metal_weights(doc, ctx, target=None, so_self=None, cctx=None):
	"""
	After gross/net weights are computed, adjust the highest-quantity metal row
	so that sum(metal + non-chain finding quantities) == target.
	Finding rows (chain or non-chain) are never touched.
	Only the difference caused by gross-weight rounding is absorbed here.

	target must be the pure metal+finding weight (diamond excluded) so that
	diamond_gms are never absorbed into metal row quantities across saves.
	"""
	if target is None:
		target = flt(doc.metal_and_finding_weight)
	if not doc.metal_detail or not target:
		return

	precision = int(ctx.metal_precision or 3)

	non_chain_finding_sum = sum(
		flt(r.quantity)
		for r in (doc.finding_detail or [])
		if r.finding_category != "Chains"
	)
	current = sum(flt(r.quantity) for r in doc.metal_detail) + non_chain_finding_sum

	diff = round(current - target, precision)
	if not diff:
		return

	# Adjust only the highest-quantity metal row
	highest_row = max(doc.metal_detail, key=lambda r: flt(r.quantity))
	new_qty = round(flt(highest_row.quantity) - diff, precision)
	if new_qty <= 0:
		# Difference is too large to absorb safely — skip reconcile rather than
		# produce a negative quantity that would break e-invoice aggregation.
		return
	highest_row.quantity = new_qty
	highest_row.quantity_3 = round(flt(highest_row.quantity), 2)
	highest_row.amount = round(flt(highest_row.rate) * flt(highest_row.quantity), 2)

	if (
		doc.company == "KG GK Jewellers Private Limited"
		or ctx.customer_group == "Internal"
	):
		highest_row.making_amount = round(
			flt(highest_row.making_rate) * flt(highest_row.quantity), 2
		)

	if so_self and cctx:
		_, _, _threshold = _get_making_charge(
			so_self, doc, highest_row.metal_touch, ctx, cctx
		)
		if doc.metal_and_finding_weight < _threshold:
			highest_row.making_amount = round(flt(highest_row.making_rate), 2)
		else:
			highest_row.making_amount = round(
				flt(highest_row.making_rate) * flt(highest_row.quantity), 2
			)
	else:
		highest_row.making_amount = round(
			flt(highest_row.making_rate) * flt(highest_row.quantity), 2
		)
	# frappe.msgprint(f"{highest_row.making_amount}")
	if flt(highest_row.wastage_rate):
		highest_row.wastage_amount = round(
			flt(highest_row.wastage_rate) * flt(highest_row.amount), 2
		)

	# Recompute metal-level totals so downstream steps use corrected values
	doc.total_metal_amount = sum(flt(r.amount) for r in doc.metal_detail)
	doc.total_wastage_amount = sum(flt(r.wastage_amount) for r in doc.metal_detail)
	doc.total_making_amount = sum(flt(r.making_amount) for r in doc.metal_detail)
	doc.total_metal_weight = sum(flt(r.quantity) for r in doc.metal_detail)


def _update_bom_totals(self, doc, row, ctx, item_code, serial_no, cctx=None):
	making_charges_on = frappe.db.get_value(
		"Customer", self.customer, "compute_making_charges_on"
	)

	# Safe precision locals — prevents TypeError if customer fields are unset
	_prec = int(ctx.precision or 2)
	_net_wt_prec = int(ctx.precision_for_net_weight or 3)
	_gross_wt_prec = int(ctx.precision_for_gross_weight or 3)

	# ── 1. Diamond weight ────────────────────────────────────────
	doc.diamond_weight = sum(r.quantity for r in doc.diamond_detail)
	doc.total_metal_weight = sum(r.quantity for r in doc.metal_detail)
	doc.total_diamond_weight_in_gms = round(doc.diamond_weight / 5, 3)

	# ── 2. Chain weight ──────────────────────────────────────────
	chain_weight = sum(
		r.quantity for r in doc.finding_detail if r.finding_category == "Chains"
	)

	# ── 3. Gemstone weights ──────────────────────────────────────
	doc.total_gemstone_weight = sum(r.quantity for r in doc.gemstone_detail)
	doc.custom_total_gemstone_weight2_digits = sum(
		r.quantity_3 for r in doc.gemstone_detail
	)
	doc.gemstone_weight = doc.custom_total_gemstone_weight2_digits
	doc.total_gemstone_weight_in_gms = round(doc.total_gemstone_weight / 5, 3)

	# ── 4. Finding weights ───────────────────────────────────────
	doc.finding_weight = sum(r.quantity for r in doc.finding_detail)
	doc.custom_finding_weight2_digits = sum(r.quantity_3 for r in doc.finding_detail)
	doc.finding_weight_ = doc.custom_finding_weight2_digits
	doc.total_finding_weight_per_gram = doc.finding_weight

	# ── 5. Pcs and other ─────────────────────────────────────────
	doc.total_diamond_pcs = sum(flt(r.pcs) for r in doc.diamond_detail)
	doc.total_gemstone_pcs = sum(flt(r.pcs) for r in doc.gemstone_detail)
	doc.total_other_weight = sum(r.quantity for r in doc.other_detail)
	doc.other_weight = doc.total_other_weight

	# ── 6. Raw component weights (from current doc rows) ─────────
	raw_metal_weight = sum(r.quantity for r in doc.metal_detail)
	raw_finding_weight = sum(
		r.quantity for r in doc.finding_detail if r.finding_category != "Chains"
	)
	doc.metal_weight = raw_metal_weight  # for ratio calc

	# ── 7. Net weight DIRECTLY from components ──────────────────────
	# Compute net before touching metal rows so the reconcile target
	# is always derived from the SNC-sourced (pre-adjustment) quantities.
	diamond_gms = flt(doc.total_diamond_weight_in_gms)
	gem_gms = flt(doc.total_gemstone_weight_in_gms)

	if making_charges_on == "Diamond Inclusive":
		# Diamond weight is part of the net metal pool
		doc.metal_and_finding_weight = round(
			raw_metal_weight + raw_finding_weight + diamond_gms, _net_wt_prec
		)
	else:
		# Diamond Exclusive — only metal + non-chain finding
		doc.metal_and_finding_weight = round(
			raw_metal_weight + raw_finding_weight, _net_wt_prec
		)

	# ── 8. Reconcile metal rows to match net weight ──────────────
	if not (
		self.company == "KG GK Jewellers Private Limited"
		or ctx.customer_group == "Internal"
	):
		_reconcile_metal_weights(
			doc, ctx, doc.metal_and_finding_weight, so_self=self, cctx=cctx
		)

	# ── 9. Gross weight ONCE from reconciled rows ────────────────
	# DI: diamond_gms is absorbed into metal rows by reconcile — do
	#     NOT add it again or gross will double-count it every save.
	# DE: diamond is separate, add normally.
	if making_charges_on == "Diamond Inclusive":
		doc.gross_weight = round(
			sum(flt(r.quantity) for r in doc.metal_detail)
			+ raw_finding_weight
			+ chain_weight
			+ gem_gms
			+ flt(doc.total_other_weight),
			_gross_wt_prec,
		)
	else:
		doc.gross_weight = round(
			sum(flt(r.quantity) for r in doc.metal_detail)
			+ raw_finding_weight
			+ chain_weight
			+ diamond_gms
			+ gem_gms
			+ flt(doc.total_other_weight),
			_gross_wt_prec,
		)

	# ── 9. Ratios ────────────────────────────────────────────────
	doc.gold_to_diamond_ratio = (
		flt(doc.metal_and_finding_weight) / flt(doc.diamond_weight)
		if doc.diamond_weight
		else 0
	)
	doc.diamond_ratio = (
		flt(doc.diamond_weight) / flt(doc.total_diamond_pcs)
		if doc.total_diamond_pcs
		else 0
	)
	doc.metal_to_diamond_ratio_excl_of_finding = (
		flt(doc.metal_weight) / flt(doc.diamond_weight) if doc.diamond_weight else 0
	)

	# ── 10. Pure weight ──────────────────────────────────────────
	doc.custom_total_pure_weight = sum(
		r.quantity * (flt(r.metal_purity) / 100) for r in doc.metal_detail
	)
	doc.custom_total_pure_finding_weight = sum(
		r.quantity * (flt(r.metal_purity) / 100) for r in doc.finding_detail
	)
	doc.custom_net_pure_weight = (
		doc.custom_total_pure_weight + doc.custom_total_pure_finding_weight
	)

	# ── 11. Certification / hallmarking ──────────────────────────
	if self.customer not in _ccp_cache:
		ccp = frappe.db.get_all(
			"Customer Certification Price",
			filters={"customer": self.customer},
			limit=1,
		)
		_ccp_cache[self.customer] = (
			frappe.get_doc("Customer Certification Price", ccp[0].name) if ccp else None
		)

	ccp_doc = _ccp_cache[self.customer]
	if ccp_doc:
		doc.certification_amount = round(
			ccp_doc.per_pc_rate
			if doc.diamond_weight <= ccp_doc.wt_threshold
			else ccp_doc.per_carat_rate * doc.diamond_weight,
			_prec,
		)
		doc.hallmarking_amount = ccp_doc.hallmarking_amount

	# if "Earrings" in (doc.item_subcategory or ""):
	# 	doc.hallmarking_amount = (doc.hallmarking_amount or 0) * 2

	# ── 12. BOM amounts ──────────────────────────────────────────
	doc.diamond_bom_amount = round(doc.total_diamond_amount, _prec)
	doc.gold_bom_amount = round(doc.total_metal_amount, _prec)
	doc.gemstone_bom_amount = round(doc.total_gemstone_amount, _prec)
	doc.finding_bom_amount = round(doc.total_finding_amount, _prec)

	doc.total_bom_amount = round(
		doc.diamond_bom_amount
		+ doc.gold_bom_amount
		+ doc.gemstone_bom_amount
		+ doc.finding_bom_amount
		+ doc.total_wastage_amount
		+ sum(flt(r.wastage_amount) for r in doc.get("finding_detail", [])),
		_prec,
	)

	doc.making_charge = round(
		sum(r.making_amount for r in doc.metal_detail)
		+ sum(r.making_amount for r in doc.finding_detail),
		_prec,
	)
	if (self.company == "KG GK Jewellers Private Limited" or ctx.customer_group == "Internal"):
		doc.custom_kg_selling_gold_bom_amount=doc.gold_bom_amount
		doc.custom_kg_selling_total_bom_amount=doc.total_bom_amount
		doc.custom_kg_selling_making_charge= doc.making_charge
		doc.custom_kg_selling_other_bom_amount= doc.other_bom_amount
		doc.custom_kg_selling_finding_bom_amount= doc.finding_bom_amount
		doc.custom_kg_selling_gemstone_bom_amount= doc.gemstone_bom_amount
		doc.custom_kg_selling_diamond_bom_amount= doc.diamond_bom_amount
	else:
		doc.custom_gk_sell_gold_bom_amount=doc.gold_bom_amount
		doc.custom_gk_sell_total_bom_amount=doc.total_bom_amount
		doc.custom_gk_sell_making_charge= doc.making_charge
		doc.custom_gk_sell_other_bom_amount= doc.other_bom_amount
		doc.custom_gk_sell_finding_bom_amount= doc.finding_bom_amount
		doc.custom_gk_sell_gemstone_bom_amount= doc.gemstone_bom_amount
		doc.custom_gk_sell_diamond_bom_amount= doc.diamond_bom_amount
	# ── 13. Total amount ─────────────────────────────────────────
	total_amount = round(
		doc.total_bom_amount
		+ doc.making_charge
		+ flt(doc.certification_amount)
		+ flt(doc.custom_duty_amount)
		+ flt(doc.hallmarking_amount)
		+ flt(doc.freight_amount)
		+ flt(doc.sale_amount),
		_prec,
	)

	if self.sales_type == "Repairing":
		total_amount = doc.total_bom_amount

	# ── 14. Assign row fields ────────────────────────────────────
	row.item_code = item_code
	row.serial_no = serial_no
	row.qty = 1
	row.rate = round(total_amount, _prec)
	row.amount = round(total_amount, _prec)
	row.gold_bom_rate = round(doc.gold_bom_amount, _prec)
	row.diamond_bom_rate = round(doc.diamond_bom_amount, _prec)
	row.gemstone_bom_rate = round(doc.gemstone_bom_amount, _prec)
	row.other_bom_rate = round(doc.other_bom_amount, _prec)
	row.making_charge = round(doc.making_charge, _prec)
	row.custom_diamond_pcs=doc.total_diamond_pcs
	row.custom_gemstone_pcs=doc.total_gemstone_pcs
	row.custom_other_weight = doc.total_other_weight
	row.custom_metal_weight=doc.total_metal_weight
	row.custom_finding_weight=doc.finding_weight
	row.custom_diamond_weight=doc.total_diamond_weight_in_gms
	row.custom_gemstone_weight=doc.total_gemstone_weight_in_gms
	row.custom_gross_weight=doc.gross_weight
	self.custom_diamond_pcs = sum(flt(r.custom_diamond_pcs) for r in self.items)
	self.custom_gemstone_pcs = sum(flt(r.custom_gemstone_pcs) for r in self.items)
	self.custom_other_weight = sum(flt(r.custom_other_weight) for r in self.items)
	self.custom_metal_weight = sum(flt(r.custom_metal_weight) for r in self.items)
	self.custom_finding_weight = sum(flt(r.custom_finding_weight) for r in self.items)
	self.custom_diamond_weight = sum(flt(r.custom_diamond_weight) for r in self.items)
	self.custom_gemstone_weight = sum(flt(r.custom_gemstone_weight) for r in self.items)
	self.custom_gross_weight= sum(flt(r.custom_gross_weight) for r in self.items)

	def _split_weight(detail, factor=1.0):
		return (
			sum(r.quantity for r in detail if not r.is_customer_item) * factor,
			sum(r.quantity for r in detail if r.is_customer_item) * factor,
		)

	m_co, m_ci = _split_weight(doc.metal_detail)
	f_co, f_ci = _split_weight(doc.finding_detail)
	d_co, d_ci = _split_weight(doc.diamond_detail, 0.2)
	g_co, g_ci = _split_weight(doc.gemstone_detail, 0.2)

	# row.custom_company_rm_weight = m_co + f_co + d_co + g_co
	# row.custom_customer_weight = m_ci + f_ci + d_ci + g_ci
	row.custom_company_rm_weight = m_co + f_co + d_co + g_co
	row.custom_customer_weight   = m_ci + f_ci + d_ci + g_ci

	# ── Hybrid: split this row's taxable amount into company-owned
	# (is_customer_item = 0, taxed at the Outright rate) vs
	# customer-supplied (is_customer_item = 1, taxed at the Outwork
	# rate) material, so set_gst_details() can blend the two rates without
	# needing to re-walk the BOM. Certification/hallmarking/freight/custom
	# duty/sale amount are always treated as company-owned.
	if self.sales_type == "Hybrid":
		def _material_vs_labour(detail):
			"""Metal/Finding: material amount is owned-gated (already 0
			when is_customer_item=1); making+wastage is always labour —
			it's job-work value whether it's our metal or the customer's."""
			owned = supplied = 0.0
			for r in detail:
				owned    += flt(r.amount)
				supplied += flt(r.making_amount) + flt(r.wastage_amount)
			return owned, supplied

		def _material_vs_handling(detail):
			"""Diamond: base stone rate is owned-gated (already 0 when
			is_customer_item=1); handling_rate is always labour, set for
			both owned and supplied rows."""
			owned = supplied = 0.0
			for r in detail:
				owned    += flt(r.quantity) * flt(r.total_diamond_rate)
				supplied += flt(r.quantity) * flt(r.handling_rate)
			return owned, supplied

		def _owned_supplied_gemstone(detail):
			# Gemstone has no separate handling_rate field — total_gemstone_rate
			# is either the material rate or the handling rate depending on
            # is_customer_item, never both — so it stays ownership-gated.
			owned = supplied = 0.0
			for r in detail:
				value = flt(r.gemstone_rate_for_specified_quantity)
				if r.is_customer_item:
					supplied += value
				else:
					owned += value
			return owned, supplied

		m_owned, m_supplied = _material_vs_labour(doc.metal_detail)
		f_owned, f_supplied = _material_vs_labour(doc.finding_detail)
		d_owned, d_supplied = _material_vs_handling(doc.diamond_detail)
		g_owned, g_supplied = _owned_supplied_gemstone(doc.gemstone_detail)
		misc_owned = (
            flt(doc.certification_amount) + flt(doc.hallmarking_amount)
            + flt(doc.custom_duty_amount) + flt(doc.freight_amount) + flt(doc.sale_amount)
        )

		row.custom_company_owned_amount = round(
            m_owned + f_owned + d_owned + g_owned + misc_owned, _prec
        )
		row.custom_customer_supplied_amount = round(
            m_supplied + f_supplied + d_supplied + g_supplied, _prec
        )
	else:
		row.custom_company_owned_amount     = 0
		row.custom_customer_supplied_amount = 0

	if self.custom_diamond_quality:
		row.diamond_quality = self.custom_diamond_quality

	# ── 15. Accumulate SO total ──────────────────────────────────
	self.total = round(self.total + row.amount, _prec)

	if self.custom_diamond_quality:
		row.diamond_quality = self.custom_diamond_quality

	# ── 15. Accumulate SO total ──────────────────────────────────
	self.total = round(self.total + row.amount, _prec)


def create_serial_no_bom(self, row, ctx=None):
	serial_no_bom = frappe.db.get_value("Serial No", row.serial_no, "custom_bom_no")
	if not serial_no_bom:
		return
	# frappe.msgprint(f"{serial_no_bom}")
	bom_doc = frappe.get_doc("BOM", serial_no_bom)
	product_certification = frappe.db.get_value(
		"Customer", self.customer, "custom_ignore_po_creation_for_certification"
	)
	doc = frappe.copy_doc(bom_doc)
	doc.customer = self.customer
	if product_certification:
		doc.hallmarking_amount = 0
		doc.certification_amount = 0
	doc.gold_rate_with_gst = self.gold_rate_with_gst
	if hasattr(doc, "diamond_detail"):
		for diamond in doc.diamond_detail or []:
			diamond.quality = row.diamond_quality

	# Apply customer precision rounding to all quantities so the
	# new Sales Order BOM starts with correctly rounded design weights.
	# frappe.throw(f"{ctx}")
	if ctx:
		_m_prec = int(ctx.metal_precision or 3)
		_s_prec = int(ctx.stone_precision or 3)
		for m in doc.metal_detail:
			m.quantity = round(flt(m.quantity), _m_prec)
		for f in doc.finding_detail or []:
			f.quantity = round(flt(f.quantity), _m_prec)
		for d in doc.diamond_detail or []:
			d.quantity = round(flt(d.quantity), _s_prec)
		for g in doc.gemstone_detail or []:
			g.quantity = round(flt(g.quantity), _s_prec)
		for o in doc.other_detail or []:
			o.quantity = round(flt(o.quantity), _m_prec)

	doc.save(ignore_permissions=True)
	row.bom = doc.name
	row.bom_no = doc.name
	# frappe.msgprint(f"{row.bom}")


def _process_single_row(self, row, ctx):
	serial_no = row.serial_no
	item_code = row.item_code
	cctx = _get_company_context(self, row, ctx)  # ← now cached

	# ── Step 1: create BOM if it doesn't exist yet ──────────────
	if not row.custom_tracking_bom:
		create_serial_no_bom(self, row, ctx)
		if not row.bom:
			return
		# ── Step 2: always process the BOM (new or existing) ─────────
		if frappe.db.get_value("BOM", row.bom, "docstatus") == 1:
			frappe.db.set_value("BOM", row.bom, "docstatus", "0")
		doc = frappe.get_doc("BOM", row.bom)
		# frappe.throw(f"{doc.as_dict()}")
		# ── Reset quantities from original Serial No BOM ─────────────
		# On every save, restore design quantities from the source BOM
		# (Serial No → custom_bom_no) with customer precision rounding.
		# This prevents reconcile drift from accumulating across saves —
		# each save always starts from the same clean design weights.
		_snc_bom_name = (
			frappe.db.get_value("Serial No", serial_no, "custom_bom_no")
			if serial_no
			else None
		)
		if _snc_bom_name:
			_snc = frappe.get_doc("BOM", _snc_bom_name)
			_m_prec = int(ctx.metal_precision or 3)
			_s_prec = int(ctx.stone_precision or 3)
			for m_d, m_s in zip(doc.metal_detail, _snc.metal_detail):
				m_d.quantity = round(flt(m_s.quantity), _m_prec)
			for f_d, f_s in zip(doc.finding_detail, _snc.finding_detail):
				f_d.quantity = round(flt(f_s.quantity), _m_prec)
			for d_d, d_s in zip(doc.diamond_detail, _snc.diamond_detail):
				d_d.quantity = round(flt(d_s.quantity), _s_prec)
			for g_d, g_s in zip(doc.gemstone_detail, _snc.gemstone_detail):
				g_d.quantity = round(flt(g_s.quantity), _s_prec)
			for o_d, o_s in zip(doc.other_detail or [], _snc.other_detail or []):
				o_d.quantity = round(flt(o_s.quantity), _m_prec)

		# Pre-compute metal_and_finding_weight with Diamond Inclusive/Exclusive
		# logic so threshold checks in _process_metal_detail1 / _process_finding_detail1
		# are correct (quantities are already reset from SNC BOM above).
		_pre_making_charges_on = frappe.db.get_value(
			"Customer", self.customer, "compute_making_charges_on"
		)
		_pre_chain = sum(
			r.quantity for r in doc.finding_detail if r.finding_category == "Chains"
		)
		_pre_non_chain = sum(
			r.quantity for r in doc.finding_detail if r.finding_category != "Chains"
		)
		_pre_metal = sum(r.quantity for r in doc.metal_detail)
		_pre_dia_gms = round(sum(r.quantity for r in doc.diamond_detail) / 5, 3)
		_pre_gem_gms = round(sum(r.quantity for r in doc.gemstone_detail) / 5, 3)
		_pre_other = sum(r.quantity for r in (doc.other_detail or []))
		_pre_gross = (
			_pre_metal
			+ _pre_non_chain
			+ _pre_chain
			+ _pre_dia_gms
			+ _pre_gem_gms
			+ _pre_other
		)

		if _pre_making_charges_on == "Diamond Inclusive":
			doc.metal_and_finding_weight = round(
				_pre_gross - _pre_gem_gms - _pre_other - _pre_chain, 3
			)
		else:
			doc.metal_and_finding_weight = round(
				_pre_gross - _pre_dia_gms - _pre_gem_gms - _pre_other - _pre_chain, 3
			)

		_process_gemstone_detail(self, doc, ctx, cctx)
		_process_metal_detail1(self, doc, ctx, cctx)
		_process_finding_detail1(self, doc, ctx, cctx)
		_process_diamond_detail(self, doc, ctx, row, cctx)
		_update_bom_totals(self, doc, row, ctx, item_code, serial_no, cctx=cctx)

		doc.save(ignore_permissions=True)

	elif not row.bom and frappe.db.exists("Tracking Bom", row.custom_tracking_bom):
		# row.bom = row.custom_tracking_bom
		frappe.db.set_value(
			"Tracking Bom",
			row.custom_tracking_bom,
			{
				"bom_type": "Sales Order",
				"reference_doctype": "Sales Order",
				"reference_docname": self.name,
				"gold_rate_with_gst": self.gold_rate_with_gst,
			},
		)
		doc = frappe.get_doc("Tracking Bom", row.custom_tracking_bom)
		row.gold_bom_rate = doc.gold_bom_amount
		row.diamond_bom_rate = doc.diamond_bom_amount
		row.gemstone_bom_rate = doc.gemstone_bom_amount
		row.other_bom_rate = doc.other_bom_amount
		row.making_charge = doc.making_charge
		row.bom_rate = doc.total_bom_amount
		row.rate = doc.total_bom_amount


def _bulk_update_child_rows(self):
	"""
	Replace N individual set_value calls with a single
	UPDATE … CASE WHEN … END statement.
	"""
	if not self.items:
		return

	child_doctype = self.items[0].doctype
	fields = [
		"bom",
		"bom_no",
		"rate",
		"amount",
		"gold_bom_rate",
		"diamond_bom_rate",
		"gemstone_bom_rate",
		"other_bom_rate",
		"making_charge",
		"custom_company_rm_weight",
		"custom_customer_weight",
	]

	# Build CASE blocks
	cases = {f: [] for f in fields}
	names = []

	for row in self.items:
		n = frappe.db.escape(row.name)
		names.append(n)
		vals = {
			"bom": row.bom or "",
			"bom_no": getattr(row, "bom_no", row.bom) or "",
			"rate": row.rate or 0,
			"amount": row.amount or 0,
			"gold_bom_rate": row.gold_bom_rate or 0,
			"diamond_bom_rate": row.diamond_bom_rate or 0,
			"gemstone_bom_rate": row.gemstone_bom_rate or 0,
			"other_bom_rate": row.other_bom_rate or 0,
			"making_charge": row.making_charge or 0,
			"custom_company_rm_weight": getattr(row, "custom_company_rm_weight", 0)
			or 0,
			"custom_customer_weight": getattr(row, "custom_customer_weight", 0) or 0,
		}
		for f in fields:
			v = frappe.db.escape(str(vals[f]))
			cases[f].append(f"WHEN {n} THEN {v}")

	set_clause = ", ".join(
		f"`{f}` = CASE `name` {' '.join(cases[f])} END" for f in fields
	)
	name_list = ", ".join(names)

	frappe.db.sql(
		f"""
        UPDATE `tab{child_doctype}`
        SET {set_clause}
        WHERE `name` IN ({name_list})
    """
	)


def create_new_bom1(self):
	"""
	Process all BOM rows synchronously with caching.
	No RQ jobs. No chunking. No queue explosion.

	Performance improvements vs original:
			• _get_company_context  : queried once per serial_no/SO  (was once per row)
			• _get_making_charge    : queried once per unique combo   (was once per row)
			• Metal purity lookups  : queried once per (customer, type, touch)
			• CCP doc               : queried once per customer
			• Gemstone PL type      : queried once per customer
			• Child row writes      : 1 bulk UPDATE                  (was N set_value calls)
	"""
	_clear_caches()  # always start fresh — never use stale data from a prior call

	self.total = 0
	ctx = _get_bom_context(self)

	for i, row in enumerate(self.items):
		try:
			_process_single_row(self, row, ctx)
		except Exception:
			frappe.log_error(
				frappe.get_traceback(),
				f"BOM failed — {self.name} row {i} serial_no {row.serial_no}",
			)

	_bulk_update_child_rows(self)
	# frappe.db.set_value(self.doctype, self.name, "total", self.total)
	# frappe.db.commit()
	for row in self.items:
		row.net_amount = flt(row.amount) / flt(row.conversion_factor or 1)
		row.base_amount = flt(row.amount)


def create_sales_order_bom(self, row, diamond_grade_data):
	doc = frappe.copy_doc(frappe.get_doc("BOM", row.quotation_bom))
	# doc = get_mapped_doc(
	# 	"BOM",
	# 	row.quotation_bom,
	# 	{
	# 		"BOM": {
	# 			"doctype": "BOM",
	# 		}
	# 	},
	# 	ignore_permissions=True,
	# )
	try:
		doc.custom_creation_doctype = self.doctype
		doc.is_default = 0
		doc.is_active = 1
		doc.bom_type = "Sales Order"
		doc.gold_rate_with_gst = self.gold_rate_with_gst
		doc.customer = self.customer
		doc.selling_price_list = self.selling_price_list
		doc.reference_doctype = "Sales Order"
		doc.reference_docname = self.name
		doc.custom_creation_docname = None
		# doc.save(ignore_permissions=True)
		for diamond in doc.diamond_detail:
			if row.diamond_grade:
				diamond.diamond_grade = row.diamond_grade
				diamond.quality = self.custom_diamond_quality

			else:
				if not diamond_grade_data.get(row.diamond_quality):
					diamond_grade_data[row.diamond_quality] = frappe.db.get_value(
						"Customer Diamond Grade",
						{
							"parent": doc.customer,
							"diamond_quality": row.diamond_quality,
						},
						"diamond_grade_1",
					)

				diamond.diamond_grade = diamond_grade_data.get(row.diamond_quality)
			if row.diamond_quality:
				diamond.quality = row.diamond_quality

		# This Save will Call before_save and validate method in BOM and Rates Will be Calculated as diamond_quality is calculated too
		doc.save(ignore_permissions=True)
		doc.db_set("custom_creation_docname", self.name)
		row.bom = doc.name
		row.gold_bom_rate = doc.gold_bom_amount
		row.diamond_bom_rate = doc.diamond_bom_amount
		row.gemstone_bom_rate = doc.gemstone_bom_amount
		row.other_bom_rate = doc.other_bom_amount
		row.making_charge = doc.making_charge
		row.bom_rate = doc.total_bom_amount
		row.rate = doc.total_bom_amount
		# frappe.msgprint(f"{row.rate}HERE13")
		self.total = doc.total_bom_amount
		# frappe.throw(f"{self.total}")
	except Exception as e:
		frappe.logger("utils").exception(e)
		frappe.log_error(
			title=f"Error while creating Sales Order from {row.quotation_bom}",
			message=str(e),
		)
		frappe.throw(_("Row {0} {1}").format(row.idx, e))


def validate_snc(self):
	for row in self.items:
		if row.serial_no:
			if self.docstatus == 2:
				frappe.db.set_value("Serial No", row.serial_no, "status", "Active")
			else:
				frappe.db.set_value("Serial No", row.serial_no, "status", "Reserved")


def submit_bom(self):
	for row in self.items:
		if row.bom:
			bom_doc = frappe.get_doc("BOM", row.bom)
			if bom_doc.docstatus == 0:
				bom_doc.submit()
			# frappe.enqueue(enqueue_submit_bom, job_name="Submitting SO BOM", bom=row.bom)


# def enqueue_submit_bom(bom):
# 	bom_doc = frappe.get_doc("BOM", bom)
# 	if bom_doc.docstatus == 0:
# 		bom_doc.submit()


def cancel_bom(self):
	for row in self.items:
		if row.bom:
			bom = frappe.get_doc("BOM", row.bom)
			bom.is_active = 0
			row.bom = ""


def validate_serial_number(self):
	if getattr(self, "skip_serial_validation", False):
		return

	for row in self.items:
		if row.serial_no and (not row.quotation_item or not row.prevdoc_docname):
			# serial_nos = [s.strip() for s in row.serial_no.split('\n') if s.strip()]

			# for serial in serial_nos:
			existing = frappe.db.sql(
				"""
				SELECT soi.name, soi.parent
				FROM `tabSales Order Item` soi
				JOIN `tabSales Order` so ON soi.parent = so.name
				WHERE so.docstatus = 1
					AND soi.serial_no = %s

			""",
				(row.serial_no),
				as_dict=True,
			)
			# if existing:
			# 	so_name = existing[0].parent
			# 	frappe.throw(f"Serial No {row.serial_no} is already used in submitted Sales Order {so_name}.")


from io import BytesIO

import frappe
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


@frappe.whitelist()
def xl_preview_sales_order(docname):
	doc = frappe.get_doc("Sales Order", docname)
	rows_diamond = []

	# Excel columns (Diamond Rate pehle, Diamond Amount baadme shift kiya hai)
	columns = [
		"Index",
		"Item Code",
		"Serial No",
		"Item Name",
		"Diamond Quality",
		"PCS",
		"Diamond Weight",
		"Average",
		"Total Cts",
		"Grams",
		"Total Diamond Rate",
		"Diamond Amount",
		"Gross Weight",
		"Gemstone Weight",
		"Other Weight",
		"Gold Rate",
		"Net Weight",
		"Gold Amount",
		"Customer Purity",
		"Chain Weight",
		"Chain Amount",
		"Chain Purity",
		"Per Gram MC",
		"Chain MC",
		"Chain Wastage %",
		"Chain Wastage Amount",
		"Jewellery Per Gram MC",
		"Jewellery MC",
		"Gold Wastage %",
		"Jewellery Wastage",
		"Gemstone Pcs",
		"Gemstone Cts",
		"Gemstone Amount",
		"Cert Charge",
		"Hallmark Charge",
		"Total Amt",
	]

	# --- Populate rows_diamond ---
	for item in doc.items:
		#  New logic for BOM selection
		bom_name = None
		if item.custom_tracking_bom:
			bom_name = item.custom_tracking_bom
		elif (
			hasattr(item, "bom") and item.bom
		):  # Check if BOM field exists in Sales Order Item
			bom_name = item.bom

		if not bom_name:
			continue  # Skip if no BOM found

		bom_doc = frappe.get_doc("BOM", bom_name)

		total_qty = sum([float(d.quantity or 0) for d in bom_doc.diamond_detail])
		# frappe.throw(f"{total_qty}")
		grams = total_qty * 0.2

		gross_weight = float(bom_doc.gross_weight or 0)
		gross_weight = round(gross_weight, 2)

		gemstone_weight = float(bom_doc.total_gemstone_weight_in_gms or 0)
		other_weight = float(bom_doc.other_weight or 0)
		net_weight = float(bom_doc.metal_and_finding_weight or 0)

		gemstone_pcs_rows = (
			[float(g.pcs or 0) for g in bom_doc.gemstone_detail]
			if bom_doc.gemstone_detail
			else []
		)
		gemstone_cts_rows = (
			[float(g.quantity or 0) for g in bom_doc.gemstone_detail]
			if bom_doc.gemstone_detail
			else []
		)
		gemstone_amount_rows = (
			[
				float(g.gemstone_rate_for_specified_quantity or 0)
				for g in bom_doc.gemstone_detail
			]
			if bom_doc.gemstone_detail
			else []
		)

		chain_weight_val, chain_mc_val, chain_wastage_val = 0.0, 0.0, 0.0
		chain_weight, chain_amount, chain_mc, chain_wastage, chain_purity = (
			0,
			0,
			0,
			0,
			0,
		)
		per_gram_mc, chain_wastage_amount = 0, 0
		net_weight_from_findings = 0.0

		if bom_doc.finding_detail:
			for f in bom_doc.finding_detail:
				qty = float(f.quantity or 0)
				if f.finding_category and f.finding_category.lower() == "chains":
					chain_weight_val += qty
					chain_purity = float(f.customer_metal_purity or 0)
					per_gram_mc = float(f.making_rate or 0)
					# frappe.throw(f"{per_gram_mc}")
					chain_mc_val = float(f.making_amount or 0)
					chain_wastage_val = float(f.wastage_rate or 0)
				else:
					net_weight_from_findings += qty

		net_weight_display = net_weight + net_weight_from_findings

		#  Net Weight se chain weight minus karna ---
		if chain_weight > 0:
			net_weight_display = net_weight_display - chain_weight

		if bom_doc.metal_detail:
			customer_metal_purity = float(
				bom_doc.metal_detail[0].customer_metal_purity or 0
			)
			gold_wastage = float(bom_doc.metal_detail[0].wastage_rate or 0)
			jewellery_per_gram_mc = float(bom_doc.metal_detail[0].making_rate or 0)
		else:
			customer_metal_purity, gold_wastage, jewellery_per_gram_mc = 0.0, 0, 0

		quotation_gold_rate = float(doc.gold_rate or 0)
		calculated_gold_rate = (quotation_gold_rate * customer_metal_purity) / 100
		calculated_gold_rate = float(f"{calculated_gold_rate:.2f}")  # Always 2 decimals
		# --- FIXED CHAIN AMOUNT CALCULATION ---
		if chain_weight_val > 0:
			chain_weight = chain_weight_val
			quotation_gold_rate = float(doc.gold_rate or 0)
			# chain_amount = (quotation_gold_rate * chain_purity / 100) * chain_weight
			chain_amount = chain_weight * calculated_gold_rate
			chain_mc = chain_mc_val
			# per_gram_mc = float(f.making_rate or 0)
			chain_wastage = chain_wastage_val
			chain_wastage_amount = (
				(chain_amount * chain_wastage_val) if chain_wastage_val else 0
			)

		cert_charge = float(bom_doc.certification_amount or 0)
		hallmark_charge = float(bom_doc.hallmarking_amount or 0)

		for i, diamond in enumerate(bom_doc.diamond_detail):
			pcs = float(diamond.pcs or 0)
			qty = float(diamond.quantity or 0)  #
			qty = float(f"{qty:.2f}")
			# total_qty = sum([float(diamond.quantity or 0)])
			avg = (qty / pcs) if pcs else 0
			rate = float(diamond.total_diamond_rate or 0)
			# diamond_amount = rate * qty
			diamond_amount = diamond.diamond_rate_for_specified_quantity

			gold_amount_val = (
				calculated_gold_rate * bom_doc.metal_weight if i == 0 else 0
			)
			# gold_amount_val = calculated_gold_rate * net_weight_display if i == 0 else 0
			jewellery_wastage_val = (
				gold_amount_val * (gold_wastage / 100) if i == 0 else 0
			)

			gemstone_pcs_val = gemstone_pcs_rows[i] if i < len(gemstone_pcs_rows) else 0
			gemstone_cts_val = gemstone_cts_rows[i] if i < len(gemstone_cts_rows) else 0
			gemstone_amount_val = (
				gemstone_amount_rows[i] if i < len(gemstone_amount_rows) else 0
			)

			jewellery_mc_val = (
				net_weight_display * jewellery_per_gram_mc if i == 0 else 0
			)

			total_amt = (
				hallmark_charge
				+ cert_charge
				+ jewellery_mc_val
				+ gemstone_amount_val
				+ gold_amount_val
				+ jewellery_wastage_val
				+ diamond_amount
			)

			rows_diamond.append(
				[
					item.idx if i == 0 else "",
					item.item_code if i == 0 else "",
					item.serial_no if i == 0 else "",
					bom_doc.item_category if i == 0 else "",
					item.diamond_quality,
					pcs,
					f"{qty:.2f}",
					f"{avg:.3f}",  #
					round(total_qty, 3) if (i == 0 and total_qty != 0) else "",
					round(grams, 2) if (i == 0 and grams != 0) else "",
					round(rate, 2),  #
					round(diamond_amount, 2),  #
					round(gross_weight, 2) if (i == 0 and gross_weight != 0) else "",
					round(gemstone_weight, 2)
					if (i == 0 and gemstone_weight != 0)
					else "",
					round(other_weight, 2) if (i == 0 and other_weight != 0) else "",
					f"{calculated_gold_rate:.2f}" if i == 0 else "",
					bom_doc.metal_weight if i == 0 else "",
					f"{gold_amount_val:.2f}" if i == 0 else "",
					customer_metal_purity if i == 0 else "",
					round(chain_weight, 2) if i == 0 else "",
					round(chain_amount, 2) if i == 0 else "",
					chain_purity if i == 0 else "",
					round(per_gram_mc, 2) if i == 0 else "",
					round(chain_mc, 2) if i == 0 else "",
					round(chain_wastage, 2) if i == 0 else "",
					round(chain_wastage_amount, 2) if i == 0 else "",
					round(jewellery_per_gram_mc, 2) if i == 0 else "",
					round(jewellery_mc_val, 2) if i == 0 else "",
					round(gold_wastage, 2) if i == 0 else "",
					round(jewellery_wastage_val, 2) if i == 0 else "",
					gemstone_pcs_val if gemstone_pcs_val != 0 else "",
					round(gemstone_cts_val, 2) if gemstone_cts_val != 0 else "",
					round(gemstone_amount_val, 2) if gemstone_amount_val != 0 else "",
					round(cert_charge, 2) if i == 0 else "",
					round(hallmark_charge, 2) if i == 0 else "",
					round(total_amt, 2),
				]
			)

	# --- SUM ROW ---
	sum_row = [""] * len(columns)
	sum_row[5] = round(sum(float(r[5] or 0) for r in rows_diamond), 2)
	sum_row[6] = round(sum(float(r[6] or 0) for r in rows_diamond), 2)
	sum_row[8] = round(sum(float(r[8] or 0) for r in rows_diamond), 2)
	sum_row[10] = round(
		sum(float(r[10] or 0) for r in rows_diamond), 2
	)  # Total Diamond Rate
	sum_row[11] = round(
		sum(float(r[11] or 0) for r in rows_diamond), 2
	)  # Diamond Amount
	sum_row[12] = round(sum(float(r[12] or 0) for r in rows_diamond), 2)
	sum_row[13] = round(sum(float(r[13] or 0) for r in rows_diamond), 2)
	sum_row[14] = round(sum(float(r[14] or 0) for r in rows_diamond), 2)
	sum_row[16] = round(sum(float(r[16] or 0) for r in rows_diamond), 3)
	sum_row[17] = round(sum(float(r[17] or 0) for r in rows_diamond), 2)
	sum_row[19] = round(sum(float(r[19] or 0) for r in rows_diamond), 3)
	sum_row[20] = round(sum(float(r[20] or 0) for r in rows_diamond), 2)
	sum_row[23] = round(sum(float(r[23] or 0) for r in rows_diamond), 2)
	sum_row[25] = round(sum(float(r[25] or 0) for r in rows_diamond), 2)
	sum_row[27] = round(
		sum(float(r[27] or 0) for r in rows_diamond), 2
	)  #  Jewellery MC total
	sum_row[29] = round(sum(float(r[29] or 0) for r in rows_diamond), 2)
	sum_row[30] = round(sum(float(r[30] or 0) for r in rows_diamond), 2)
	sum_row[31] = round(sum(float(r[31] or 0) for r in rows_diamond), 2)
	sum_row[32] = round(sum(float(r[32] or 0) for r in rows_diamond), 2)
	sum_row[33] = round(sum(float(r[33] or 0) for r in rows_diamond), 2)
	sum_row[34] = round(sum(float(r[34] or 0) for r in rows_diamond), 2)
	sum_row[35] = round(sum(float(r[35] or 0) for r in rows_diamond), 2)

	rows_diamond.append(sum_row)

	# --- Create Workbook ---
	wb = Workbook()
	ws = wb.active
	ws.title = "Diamond Detail"

	# --- Add Company Name ---
	company_name = "M/S. GURUKRUPA EXPORT PVT LIMITED"
	ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
	cell = ws.cell(row=1, column=1, value=company_name)
	cell.font = Font(bold=True, size=15)
	cell.alignment = Alignment(horizontal="center", vertical="center")

	# --- Add Headers ---
	for col_num, column_title in enumerate(columns, 1):
		c = ws.cell(row=2, column=col_num, value=column_title)
		c.font = Font(bold=True)
		c.alignment = Alignment(horizontal="center", vertical="center")

	# --- Add Data Rows ---
	for row_num, row_data in enumerate(rows_diamond, 3):
		for col_num, cell_value in enumerate(row_data, 1):
			ws.cell(row=row_num, column=col_num, value=cell_value)

	# --- Auto column width ---
	for i, column in enumerate(columns, 1):
		ws.column_dimensions[get_column_letter(i)].width = 15

	# --- Save to BytesIO and Download ---
	output = BytesIO()
	wb.save(output)
	output.seek(0)
	frappe.local.response.filecontent = output.read()
	frappe.local.response.filename = f"Diamond_Detail_SO_{docname}.xlsx"
	frappe.local.response.type = "download"


@frappe.whitelist()
def get_customer_approval_data(customer_approval_data):
	doc = frappe.get_doc("Customer Approval", customer_approval_data)
	return doc


@frappe.whitelist()
def customer_approval_filter(doctype, txt, searchfield, start, page_len, filters):
	CustomerApproval = frappe.qb.DocType("Customer Approval")
	StockEntry = frappe.qb.DocType("Stock Entry")

	query = (
		frappe.qb.from_(CustomerApproval)
		.left_join(StockEntry)
		.on(CustomerApproval.name == StockEntry.custom_customer_approval_reference)
		.select(CustomerApproval.name)
		.where(
			(
				(StockEntry.custom_customer_approval_reference != CustomerApproval.name)
				| (StockEntry.custom_customer_approval_reference.isnull())
			)
			& (CustomerApproval.docstatus == 1)
			& (CustomerApproval[searchfield].like(f"%{txt}%"))
		)
	)

	if filters.get("date"):
		query = query.where(CustomerApproval.date == filters["date"])

	dialoge_filter = query.run(as_dict=True)

	return dialoge_filter


def validate_item_dharm(self):
	precision = frappe.db.get_value("Customer", self.customer, "custom_precision_variable")
	allowed = ("Outright", "Outwork", "Certification","Branch Sales","Repairing","Hybrid")
	if self.sales_type in allowed:
		if self.sales_type == "Hybrid" and self.company not in HYBRID_ENABLED_COMPANIES:
			frappe.throw(_("Hybrid Sales Type is not yet enabled for company {0}").format(self.company))

		customer_payment_term_doc = frappe.get_doc(
			"Customer Payment Terms",
			{"customer": self.customer}
		)

		e_invoice_items = []

		for row in self.items:

			gross_weighh = frappe.get_value("BOM", row.bom, "gross_weight")
			row.custom_gross_weight = gross_weighh


		# Prepare invoice items as before
		for row in customer_payment_term_doc.customer_payment_details:
			item_type = row.item_type
			e_invoice_item = frappe.get_doc("E Invoice Item", item_type)

			if self.sales_type == "Hybrid":
				# Hybrid carries no tax_rate of its own — job-work item
				# types (repair/labour, used for customer-supplied
				# material) get the Outwork rate, everything else
				# (owned material, making, hallmarking, certification)
				# gets the Outright rate.
				lookup_sales_type = (
					"Outwork"
					if (e_invoice_item.is_for_repair or e_invoice_item.is_for_labour)
					else "Outright"
				)
			else:
				lookup_sales_type = self.sales_type

			matched_sales_type_row = None
			for st_row in e_invoice_item.sales_type:
				if st_row.sales_type == lookup_sales_type:
					matched_sales_type_row = st_row
					break

			if self.sales_type and not matched_sales_type_row:
				continue

			e_invoice_items.append({
				"item_type": item_type,
				"is_for_metal": e_invoice_item.is_for_metal,
				"is_for_hallmarking":e_invoice_item.is_for_hallmarking,
				"is_for_certification":e_invoice_item.is_for_certification,
				"is_for_labour": e_invoice_item.is_for_labour,
				"is_for_diamond": e_invoice_item.is_for_diamond,
				"diamond_type": e_invoice_item.diamond_type,
				"is_for_making": e_invoice_item.is_for_making,
				"is_for_finding": e_invoice_item.is_for_finding,
				"is_for_finding_making": e_invoice_item.is_for_finding_making,
				"is_for_gemstone": e_invoice_item.is_for_gemstone,
				"metal_type": e_invoice_item.metal_type,
				"metal_purity": e_invoice_item.metal_purity,
				"uom": e_invoice_item.uom,
				"finding_category":e_invoice_item.finding_category,
				"tax_rate": matched_sales_type_row.tax_rate if matched_sales_type_row else 0,
				"is_for_repair":e_invoice_item.is_for_repair
			})

		self.set("custom_invoice_item", [])
		aggregated_metal_items = {}
		aggregated_metal_labour_items = {}
		aggregated_metal_making_items = {}
		aggregated_hallmarking_items = {}
		aggregated_certification_items = {}
		aggregated_diamond_items = {}
		aggregated_gemstone_items = {}
		aggregated_finding_items = {}
		aggregated_finding_making_items = {}
		aggregated_repairing_items = {}
		for item in self.items:
			bom_doc = None
			if item.bom:
				bom_doc = frappe.get_doc("BOM", item.bom)
			if bom_doc:
				if bom_doc.hallmarking_amount:
					# frappe.throw("hii")
					for e_item in e_invoice_items:
						if (
							e_item["is_for_hallmarking"]
						):  
							key = (e_item["item_type"], e_item["uom"])
							
							if key not in aggregated_hallmarking_items:
								
								aggregated_hallmarking_items[key] = {
									"item_code": e_item["item_type"],
									"item_name": e_item["item_type"],
									"uom": e_item["uom"],
									"qty": 0,
									"amount": 0,
									"tax_rate": e_item["tax_rate"],
									"tax_amount": 0,
									"amount_with_tax": 0,
									"delivery_date": self.delivery_date
								}

							aggregated_hallmarking_items[key]["amount"] += bom_doc.hallmarking_amount
							# frappe.msgprint(f"hii{bom_doc.hallmarking_amount}")		
							aggregated_hallmarking_items[key]["qty"]+=1
							if bom_doc.item_category=='Earrings':
								aggregated_hallmarking_items[key]["qty"]+=1
							tax_rate_decimal = aggregated_hallmarking_items[key]["tax_rate"] / 100
							aggregated_hallmarking_items[key]["tax_amount"] += bom_doc.hallmarking_amount * tax_rate_decimal

							aggregated_hallmarking_items[key]["amount_with_tax"] = (
									aggregated_hallmarking_items[key]["amount"] +
								aggregated_hallmarking_items[key]["tax_amount"]
							)
							
				if bom_doc.certification_amount:
					# frappe.throw("hii")
					for e_item in e_invoice_items:
						if (
							e_item["is_for_certification"]
						):
							key = (e_item["item_type"], e_item["uom"])
							if key not in aggregated_certification_items:
								aggregated_certification_items[key] = {
									"item_code": e_item["item_type"],
									"item_name": e_item["item_type"],
									"uom": e_item["uom"],
									"qty": 0,
									"amount": 0,
									"tax_rate": e_item["tax_rate"],
									"tax_amount": 0,
									"amount_with_tax": 0,
									"delivery_date": self.delivery_date
								}
							aggregated_certification_items[key]["amount"] += bom_doc.certification_amount
							# frappe.msgprint(f"hii{bom_doc.certification_amount}")
							aggregated_certification_items[key]["qty"]+=1
							tax_rate_decimal = aggregated_certification_items[key]["tax_rate"] / 100
							aggregated_certification_items[key]["tax_amount"] += bom_doc.certification_amount * tax_rate_decimal

							aggregated_certification_items[key]["amount_with_tax"] = (
									aggregated_certification_items[key]["amount"] +
								aggregated_certification_items[key]["tax_amount"]
							)
				for metal in bom_doc.metal_detail:
					
					if not metal.is_customer_item:
						# frappe.msgprint(f"hkhii{metal.metal_touch},{metal.stock_uom},{metal.metal_type})")
						for e_item in e_invoice_items:
							if (
								e_item["is_for_metal"] and
								metal.metal_type == e_item["metal_type"] and
								metal.metal_touch == e_item["metal_purity"] and
								metal.stock_uom == e_item["uom"]
							):
								key = (e_item["item_type"], e_item["uom"])
								
								if key not in aggregated_metal_items:
									aggregated_metal_items[key] = {
										"item_code": e_item["item_type"],
										"item_name": e_item["item_type"],
										"uom": e_item["uom"],
										"qty": 0,
										"amount": 0,
										"tax_rate": e_item["tax_rate"],
										"tax_amount": 0,
										"amount_with_tax": 0,
										"delivery_date": self.delivery_date
									}

								multiplied_qty = metal.quantity * item.qty
								
								# metal_rate = metal.se_rate if self.company == "KG GK Jewellers Private Limited" and self.customer == "GJCU0009" else metal.rate
								# making_amount=metal.making_amount
								metal_rate=metal.rate
								metal_amount = round(metal_rate * multiplied_qty , precision) 
								# Sum quantities and amounts
								aggregated_metal_items[key]["qty"] += multiplied_qty
								aggregated_metal_items[key]["amount"] += metal_amount
								# frappe.throw(f"hiii")

								# Calculate tax amount
								tax_rate_decimal = aggregated_metal_items[key]["tax_rate"] / 100
								aggregated_metal_items[key]["tax_amount"] += metal_amount * tax_rate_decimal

								aggregated_metal_items[key]["amount_with_tax"] = (
									aggregated_metal_items[key]["amount"] +
									aggregated_metal_items[key]["tax_amount"]
								)
								break
								
						for e_item in e_invoice_items:
							if (
								e_item["is_for_making"] and
								metal.metal_type == e_item["metal_type"] and
								metal.metal_touch == e_item["metal_purity"] and
								metal.stock_uom == e_item["uom"]
							):
								key = (e_item["item_type"], e_item["uom"])

								if key not in aggregated_metal_making_items:
									aggregated_metal_making_items[key] = {
										"item_code": e_item["item_type"],
										"item_name": e_item["item_type"],
										"uom": e_item["uom"],
										"qty": 0,
										"rate": metal.making_rate,  # initial rate, will be overwritten with average later
										"amount": 0,
										"tax_rate": e_item["tax_rate"],
										"tax_amount": 0,
										"amount_with_tax": 0,
										"delivery_date": self.delivery_date
									}

								# is_per_pc = (
								# 	flt(metal.making_rate) > 0
								# 	and abs(flt(metal.making_amount) - flt(metal.making_rate)) < 0.01
								# )
								# if is_per_pc:
								# 	frappe.msgprint(f"oiuhgvbjb{metal.making_amount},{metal.making_rate}")
								multiplied_qty =  metal.quantity * item.qty
								metal_making_amount = round(metal.making_amount * item.qty  + (metal.wastage_amount * item.qty) , precision)
								aggregated_metal_making_items[key]["qty"] += multiplied_qty
								# frappe.msgprint(f"poihuuhg{metal_making_amount}, {multiplied_qty}")
								aggregated_metal_making_items[key]["amount"] += metal_making_amount

								tax_rate_decimal = aggregated_metal_making_items[key]["tax_rate"] / 100
								aggregated_metal_making_items[key]["tax_amount"] += metal_making_amount * tax_rate_decimal

								aggregated_metal_making_items[key]["amount_with_tax"] = (
										aggregated_metal_making_items[key]["amount"] +
									aggregated_metal_making_items[key]["tax_amount"]
								)
								break
							
								# frappe.throw(f"{multiplied_qty},{aggregated_metal_items[key]["qty"]}")
					else:
						
						for e_item in e_invoice_items:
							# frappe.throw(f"{metal.stock_uom},{e_item["uom"]}")
							if (
								e_item["is_for_repair"] 
								# metal.metal_type == e_item["metal_type"] and
								# metal.metal_touch == e_item["metal_purity"] and
								# metal.stock_uom == e_item["uom"]
							): 
           						
								key = (e_item["item_type"])
								
								# frappe.throw("hii")
								if key not in aggregated_repairing_items:
									aggregated_repairing_items[key] = {
										"item_code": e_item["item_type"],
										"item_name": e_item["item_type"],
										# "uom": e_item["uom"],
										"qty": 0,
										"rate": metal.making_rate,  # initial rate, will be overwritten with average later
										"amount": 0,
										"tax_rate": e_item["tax_rate"],
										"tax_amount": 0,
										"amount_with_tax": 0,
										"delivery_date": self.delivery_date
									}

								multiplied_qty = metal.quantity * item.qty
								metal_making_amount = metal.making_rate * multiplied_qty  
								aggregated_repairing_items[key]["qty"] += multiplied_qty
								aggregated_repairing_items[key]["amount"] += metal_making_amount

								tax_rate_decimal = aggregated_repairing_items[key]["tax_rate"] / 100
								aggregated_repairing_items[key]["tax_amount"] += metal_making_amount * tax_rate_decimal

								aggregated_repairing_items[key]["amount_with_tax"] = (
										aggregated_repairing_items[key]["amount"] +
									aggregated_repairing_items[key]["tax_amount"]
								)
								break
						for e_item in e_invoice_items:
							
							if (
								e_item["is_for_labour"]
								# and metal.stock_uom == e_item["uom"]
								# and metal.metal_type == e_item["metal_type"]
								and metal.metal_touch == e_item["metal_purity"]
							):
								key = (e_item["item_type"], e_item["uom"])
								if key not in aggregated_metal_labour_items:
									aggregated_metal_labour_items[key] = {
										"item_code": e_item["item_type"],
										"item_name": e_item["item_type"],
										"uom": e_item["uom"],
										"qty": 0,
										"amount": 0,
										"tax_rate": e_item["tax_rate"],
										"tax_amount": 0,
										"amount_with_tax": 0,
										"delivery_date": self.delivery_date
									}

								multiplied_qty = metal.quantity * item.qty
								# metal_rate = metal.se_rate if self.company == "KG GK Jewellers Private Limited" and self.customer == "GJCU0009" else metal.making_rate
								metal_rate =metal.making_rate
								metal_amount = metal_rate * multiplied_qty

								aggregated_metal_labour_items[key]["qty"] += multiplied_qty
								aggregated_metal_labour_items[key]["amount"] += metal_amount
								tax_rate_decimal = aggregated_metal_labour_items[key]["tax_rate"] / 100
								aggregated_metal_labour_items[key]["tax_amount"] += metal_amount * tax_rate_decimal
								aggregated_metal_labour_items[key]["amount_with_tax"] = (
									aggregated_metal_labour_items[key]["amount"] +
									aggregated_metal_labour_items[key]["tax_amount"]
								)
								
						

				for diamond in bom_doc.diamond_detail:
					if not diamond.is_customer_item:
						# frappe.msgprint(f"gytffy{diamond.diamond_type},{diamond.stock_uom}")
						for e_item in e_invoice_items:
							if (
								e_item["is_for_diamond"]
								and e_item["diamond_type"] == diamond.diamond_type
								and e_item["uom"] == diamond.stock_uom
							):
								key = (e_item["item_type"], e_item["uom"])
								# frappe.msgprint("hii")
								if key not in aggregated_diamond_items:
									aggregated_diamond_items[key] = {
										"item_code": e_item["item_type"],
										"item_name": e_item["item_type"],
										"uom": e_item["uom"],
										"qty": 0,
										"rate": 0,
										"amount": 0,
										"tax_rate": e_item["tax_rate"],
										"tax_amount": 0,
										"amount_with_tax": 0,
										"delivery_date": self.delivery_date
									}

								multiplied_qty = diamond.quantity * item.qty
								diamond_rate = diamond.se_rate if self.company == "KG GK Jewellers Private Limited" and self.customer == "GJCU0009" else diamond.total_diamond_rate
								if self.sales_type == "Hybrid":
									# Hybrid: only the base stone value is Outright
									# material here — handling_rate is always
									# job-work value, folded into the labour/
									# Outwork bucket below regardless of
									# is_customer_item.
									diamond_amount = round(flt(diamond.quantity) * flt(diamond.total_diamond_rate) * item.qty, precision)
								else:
									diamond_amount = round(float(diamond.diamond_rate_for_specified_quantity) , precision)
								# frappe.msgprint(f"hii{diamond_amount}")
								aggregated_diamond_items[key]["qty"] += multiplied_qty
								aggregated_diamond_items[key]["amount"] += diamond_amount

								# Calculate average rate after accumulation
								if aggregated_diamond_items[key]["qty"] > 0:
									aggregated_diamond_items[key]["rate"] = aggregated_diamond_items[key]["amount"] / aggregated_diamond_items[key]["qty"]
								else:
									aggregated_diamond_items[key]["rate"] = 0

								tax_rate_decimal = aggregated_diamond_items[key]["tax_rate"] / 100
								aggregated_diamond_items[key]["tax_amount"] += diamond_amount * tax_rate_decimal

								aggregated_diamond_items[key]["amount_with_tax"] = (
									aggregated_diamond_items[key]["amount"] +
									aggregated_diamond_items[key]["tax_amount"]
								)

						if self.sales_type == "Hybrid":
							# Handling charge is always job-work value, whether
							# the stone is company-owned or customer-supplied —
							# route it to the same labour/Outwork bucket
							# used by the customer-supplied branch below.
							handling_amount = round(flt(diamond.quantity) * flt(diamond.handling_rate) * item.qty, precision)
							if handling_amount:
								for e_item in e_invoice_items:
									if e_item["is_for_labour"] and e_item["uom"] == diamond.stock_uom:
										key = (e_item["item_type"], e_item["uom"])
										if key not in aggregated_metal_labour_items:
											aggregated_metal_labour_items[key] = {
												"item_code": e_item["item_type"],
												"item_name": e_item["item_type"],
												"uom": e_item["uom"],
												"qty": 0,
												"rate": 0,
												"amount": 0,
												"tax_rate": e_item["tax_rate"],
												"tax_amount": 0,
												"amount_with_tax": 0,
												"delivery_date": self.delivery_date
											}
										multiplied_qty = diamond.quantity * item.qty
										aggregated_metal_labour_items[key]["qty"] += multiplied_qty
										aggregated_metal_labour_items[key]["amount"] += handling_amount
										tax_rate_decimal = aggregated_metal_labour_items[key]["tax_rate"] / 100
										aggregated_metal_labour_items[key]["tax_amount"] += handling_amount * tax_rate_decimal
										aggregated_metal_labour_items[key]["amount_with_tax"] = (
											aggregated_metal_labour_items[key]["amount"] +
											aggregated_metal_labour_items[key]["tax_amount"]
										)
										break
					else:
						
						for e_item in e_invoice_items:
							if (
								e_item["is_for_repair"]
								
							):
								key = (e_item["item_type"])
								
								# frappe.throw("hii")
								if key not in aggregated_repairing_items:
									aggregated_repairing_items[key] = {
										"item_code": e_item["item_type"],
										"item_name": e_item["item_type"],
										"uom": e_item["uom"],
										"qty": 0,
										"rate": 0,
										"amount": 0,
										"tax_rate": e_item["tax_rate"],
										"tax_amount": 0,
										"amount_with_tax": 0,
										"delivery_date": self.delivery_date
									}

								multiplied_qty = (diamond.quantity * item.qty)/5
								diamond_rate = diamond.se_rate if self.company == "KG GK Jewellers Private Limited" and self.customer == "GJCU0009" else diamond.total_diamond_rate
								diamond_amount = round(float(diamond.diamond_rate_for_specified_quantity) , precision)

								aggregated_repairing_items[key]["qty"] += multiplied_qty
								aggregated_repairing_items[key]["amount"] += diamond_amount

								# Calculate average rate after accumulation
								if aggregated_repairing_items[key]["qty"] > 0:
									aggregated_repairing_items[key]["rate"] = aggregated_repairing_items[key]["amount"] / aggregated_repairing_items[key]["qty"]
								else:
									aggregated_repairing_items[key]["rate"] = 0

								tax_rate_decimal = aggregated_repairing_items[key]["tax_rate"] / 100
								aggregated_repairing_items[key]["tax_amount"] += diamond_amount * tax_rate_decimal

								aggregated_repairing_items[key]["amount_with_tax"] = (
									aggregated_repairing_items[key]["amount"] +
									aggregated_repairing_items[key]["tax_amount"]
								)
								break
						for e_item in e_invoice_items:
							if (
								e_item["is_for_labour"]
								# and e_item["diamond_type"] == diamond.diamond_type
								# and e_item["uom"] == diamond.stock_uom
							):
								key = (e_item["item_type"], e_item["uom"])

								if key not in aggregated_metal_labour_items:
									aggregated_metal_labour_items[key] = {
										"item_code": e_item["item_type"],
										"item_name": e_item["item_type"],
										"uom": e_item["uom"],
										"qty": 0,
										"rate": 0,
										"amount": 0,
										"tax_rate": e_item["tax_rate"],
										"tax_amount": 0,
										"amount_with_tax": 0,
										"delivery_date": self.delivery_date
									}

								multiplied_qty = diamond.quantity * item.qty
								diamond_rate = diamond.se_rate if self.company == "KG GK Jewellers Private Limited" and self.customer == "GJCU0009" else diamond.total_diamond_rate
								diamond_amount = float(diamond.diamond_rate_for_specified_quantity)

								aggregated_metal_labour_items[key]["qty"] += multiplied_qty/5
								aggregated_metal_labour_items[key]["amount"] += diamond_amount
								# Calculate average rate after accumulation
								if aggregated_metal_labour_items[key]["qty"] > 0:
									aggregated_metal_labour_items[key]["rate"] = aggregated_metal_labour_items[key]["amount"] / aggregated_metal_labour_items[key]["qty"]
								else:
									aggregated_metal_labour_items[key]["rate"] = 0

								tax_rate_decimal = aggregated_metal_labour_items[key]["tax_rate"] / 100
								aggregated_metal_labour_items[key]["tax_amount"] += diamond_amount * tax_rate_decimal

								aggregated_metal_labour_items[key]["amount_with_tax"] = (
									aggregated_metal_labour_items[key]["amount"] +
									aggregated_metal_labour_items[key]["tax_amount"]
								)		

				for gemstone in bom_doc.gemstone_detail:
					for e_item in e_invoice_items:
						if not gemstone.is_customer_item:
							if (
								e_item["is_for_gemstone"]
								and e_item["uom"] == gemstone.stock_uom
							):
								key = (e_item["item_type"], e_item["uom"])

								if key not in aggregated_gemstone_items:
									aggregated_gemstone_items[key] = {
										"item_code": e_item["item_type"],
										"item_name": e_item["item_type"],
										"uom": e_item["uom"],
										"qty": 0,
										"rate": gemstone.total_gemstone_rate,  # initial rate; average will be calculated later
										"amount": 0,
										"tax_rate": e_item["tax_rate"],
										"tax_amount": 0,
										"amount_with_tax": 0,
										"delivery_date": self.delivery_date
									}

								multiplied_qty = gemstone.quantity * item.qty
								gemstone_rate = gemstone.se_rate if self.company == "KG GK Jewellers Private Limited" and self.customer == "GJCU0009" else gemstone.total_gemstone_rate
								gemstone_amount = round(float(gemstone.gemstone_rate_for_specified_quantity) , precision)

								aggregated_gemstone_items[key]["qty"] += multiplied_qty
								aggregated_gemstone_items[key]["amount"] += gemstone_amount

								# Calculate average rate after accumulation
								if aggregated_gemstone_items[key]["qty"] > 0:
									aggregated_gemstone_items[key]["rate"] = aggregated_gemstone_items[key]["amount"] / aggregated_gemstone_items[key]["qty"]
								else:
									aggregated_gemstone_items[key]["rate"] = 0

								tax_rate_decimal = aggregated_gemstone_items[key]["tax_rate"] / 100
								aggregated_gemstone_items[key]["tax_amount"] += gemstone_amount * tax_rate_decimal

								aggregated_gemstone_items[key]["amount_with_tax"] = (
									aggregated_gemstone_items[key]["amount"] +
									aggregated_gemstone_items[key]["tax_amount"]
								)
						else:
							if (
							e_item["is_for_labour"]
							and e_item["uom"] == gemstone.stock_uom
						):
								key = (e_item["item_type"], e_item["uom"])

								if key not in aggregated_metal_labour_items:
									aggregated_metal_labour_items[key] = {
										"item_code": e_item["item_type"],
										"item_name": e_item["item_type"],
										"uom": e_item["uom"],
										"qty": 0,
										"rate": gemstone.total_gemstone_rate,  # initial rate; average will be calculated later
										"amount": 0,
										"tax_rate": e_item["tax_rate"],
										"tax_amount": 0,
										"amount_with_tax": 0,
										"delivery_date": self.delivery_date
									}

								multiplied_qty = gemstone.quantity * item.qty
								gemstone_rate = gemstone.se_rate if self.company == "KG GK Jewellers Private Limited" and self.customer == "GJCU0009" else gemstone.total_gemstone_rate
								gemstone_amount = round(float(gemstone.gemstone_rate_for_specified_quantity) , precision)

								aggregated_metal_labour_items[key]["qty"] += multiplied_qty/5
								aggregated_metal_labour_items[key]["amount"] += gemstone_amount
								# Calculate average rate after accumulation
								if aggregated_metal_labour_items[key]["qty"] > 0:
									aggregated_metal_labour_items[key]["rate"] = aggregated_metal_labour_items[key]["amount"] / aggregated_metal_labour_items[key]["qty"]
								else:
									aggregated_metal_labour_items[key]["rate"] = 0

								tax_rate_decimal = aggregated_metal_labour_items[key]["tax_rate"] / 100
								aggregated_metal_labour_items[key]["tax_amount"] += gemstone_amount * tax_rate_decimal

								aggregated_metal_labour_items[key]["amount_with_tax"] = (
									aggregated_metal_labour_items[key]["amount"] +
									aggregated_metal_labour_items[key]["tax_amount"]
								)

				for finding in bom_doc.finding_detail:
					if not finding.is_customer_item:
						finding_handled = False
						for e_item in e_invoice_items:
							if (e_item["is_for_finding"] and e_item["metal_type"] == finding.metal_type and e_item["metal_purity"] == finding.metal_touch and e_item["uom"] == finding.stock_uom and e_item["finding_category"] == finding.finding_category):
								finding_handled = True
								key = (e_item["item_type"], e_item["uom"])
								if key not in aggregated_finding_items:
									aggregated_finding_items[key] = {
										"item_code": e_item["item_type"],
										"item_name": e_item["item_type"],
										"uom": e_item["uom"],
										"qty": 0,
										"rate": 0,
										"amount": 0,
										"tax_rate": e_item["tax_rate"],
										"tax_amount": 0,
										"amount_with_tax": 0,
										"delivery_date": self.delivery_date
									}
								multiplied_qty = finding.quantity * item.qty
								making_amount = round(finding.making_amount , precision)
								finding_rate = 0 
								if self.company == "KG GK Jewellers Private Limited" and self.customer == "GJCU0009":
									finding_rate = finding.se_rate 
								elif self.company == "KG GK Jewellers Private Limited" and self.customer == "GJCU0009":
									finding_rate = finding.se_rate
								elif self.company == "KG GK Jewellers Private Limited" and self.customer == "GJCU0009":
									finding_rate = finding.se_rate 
								finding_making_amount = (finding.rate * multiplied_qty) 
								# frappe.throw(f"{multiplied_qty}")
								aggregated_finding_items[key]["qty"] += multiplied_qty
								aggregated_finding_items[key]["amount"] += finding_making_amount

								aggregated_finding_items[key]["rate"] = finding_rate
								
								tax_rate_decimal = aggregated_finding_items[key]["tax_rate"] / 100
								aggregated_finding_items[key]["tax_amount"] += finding_making_amount * tax_rate_decimal

								aggregated_finding_items[key]["amount_with_tax"] = (
									aggregated_finding_items[key]["amount"] +
									aggregated_finding_items[key]["tax_amount"]
								)
								break

						if not finding_handled:
							for e_item in e_invoice_items:
								if (e_item["is_for_metal"] and finding.metal_type == e_item["metal_type"] and finding.metal_touch == e_item["metal_purity"] and finding.stock_uom == e_item["uom"] and e_item["finding_category"] is None):
									key = (e_item["item_type"], e_item["uom"])
									if key not in aggregated_metal_items:
										aggregated_metal_items[key] = {
											"item_code": e_item["item_type"],
											"item_name": e_item["item_type"],
											"uom": e_item["uom"],
											"qty": 0,
											"amount": 0,
											"tax_rate": e_item["tax_rate"],
											"tax_amount": 0,
											"amount_with_tax": 0,
											"delivery_date": self.delivery_date,
											"rate": 0
										}
									
									finding_rate = finding.se_rate if self.company == "KG GK Jewellers Private Limited" and self.customer == "GJCU0009" else finding.rate
									multiplied_qty = finding.quantity * item.qty
									making_amount = round(finding.making_amount , precision)
									finding_making_amount = (finding.rate * multiplied_qty)
									
									aggregated_metal_items[key]["qty"] += multiplied_qty
									aggregated_metal_items[key]["amount"] += finding.amount
									aggregated_metal_items[key]["rate"] = finding_rate
									
									tax_rate_decimal = aggregated_metal_items[key]["tax_rate"] / 100
									aggregated_metal_items[key]["tax_amount"] += finding_making_amount * tax_rate_decimal
									aggregated_metal_items[key]["amount_with_tax"] = (
										aggregated_metal_items[key]["amount"] + 
										aggregated_metal_items[key]["tax_amount"]
									)
									break

						
						finding_making_handled = False
						for e_item in e_invoice_items:
							if (e_item["is_for_finding_making"] and e_item["metal_type"] == finding.metal_type and e_item["metal_purity"] == finding.metal_touch and e_item["uom"] == finding.stock_uom and e_item["finding_category"] == finding.finding_category):
								finding_making_handled = True
								key = (e_item["item_type"], e_item["uom"])
								if key not in aggregated_finding_making_items:
									aggregated_finding_making_items[key] = {
										"item_code": e_item["item_type"],
										"item_name": e_item["item_type"],
										"uom": e_item["uom"],
										"qty": 0,
										"rate": finding.making_rate,
										"amount": 0,
										"tax_rate": e_item["tax_rate"],
										"tax_amount": 0,
										"amount_with_tax": 0,
										"delivery_date": self.delivery_date
									}
								
								multiplied_qty = finding.quantity * item.qty
								# frappe.throw(f"{finding.quantity},{item.qty}")
								making_amount = round(finding.making_amount , precision)
								finding_making_amount = (finding.making_amount * item.qty) + (finding.wastage_amount * item.qty)
								
								aggregated_finding_making_items[key]["qty"] += multiplied_qty
								aggregated_finding_making_items[key]["amount"] += finding_making_amount
								# frappe.throw(f"{aggregated_finding_making_items[key]["amount"]}")
								# if aggregated_finding_making_items[key]["qty"] > 0:
								# 	aggregated_finding_making_items[key]["rate"] = aggregated_finding_making_items[key]["amount"] / aggregated_finding_making_items[key]["qty"]
								# else:
								# 	aggregated_finding_making_items[key]["rate"] = 0
								
								tax_rate_decimal = aggregated_finding_making_items[key]["tax_rate"] / 100
								aggregated_finding_making_items[key]["tax_amount"] += finding_making_amount * tax_rate_decimal
								aggregated_finding_making_items[key]["amount_with_tax"] = (
									aggregated_finding_making_items[key]["amount"] +
									aggregated_finding_making_items[key]["tax_amount"]
								)
								break
						
						if not finding_making_handled:
							is_finding_per_pc = (
								flt(finding.making_rate) > 0
								and abs(flt(finding.making_amount) - flt(finding.making_rate)) < 0.01
							)
							# if is_finding_per_pc:
							# 	continue
							for e_item in e_invoice_items:
								if (e_item["is_for_making"] and e_item["metal_type"] == finding.metal_type and e_item["metal_purity"] == finding.metal_touch and e_item["uom"] == finding.stock_uom):
									key = (e_item["item_type"], e_item["uom"])
									if key not in aggregated_metal_making_items:
										aggregated_metal_making_items[key] = {
											"item_code": e_item["item_type"],
											"item_name": e_item["item_type"],
											"uom": e_item["uom"],
											"qty": 0,
											"rate": finding.making_rate,
											"amount": 0,
											"tax_rate": e_item["tax_rate"],
											"tax_amount": 0,
											"amount_with_tax": 0,
											"delivery_date": self.delivery_date
										}
									
									multiplied_qty = finding.quantity * item.qty
									finding_making_amount = round(finding.making_amount * item.qty + finding.wastage_amount * item.qty, precision)
									aggregated_metal_making_items[key]["qty"] += multiplied_qty
									aggregated_metal_making_items[key]["amount"] += finding_making_amount

									tax_rate_decimal = aggregated_metal_making_items[key]["tax_rate"] / 100
									aggregated_metal_making_items[key]["tax_amount"] += finding_making_amount * tax_rate_decimal
									aggregated_metal_making_items[key]["amount_with_tax"] = (
										aggregated_metal_making_items[key]["amount"] +
										aggregated_metal_making_items[key]["tax_amount"]
									)
									break
					else:
						for e_item in e_invoice_items:
							if (e_item["is_for_repair"] ):
								key = (e_item["item_type"])
								if key not in aggregated_repairing_items:
									aggregated_repairing_items[key] = {
										"item_code": e_item["item_type"],
										"item_name": e_item["item_type"],
										# "uom": e_item["uom"],
										"qty": 0,
										"rate": finding.making_rate,
										"amount": 0,
										"tax_rate": e_item["tax_rate"],
										"tax_amount": 0,
										"amount_with_tax": 0,
										"delivery_date": self.delivery_date
									}
								
								multiplied_qty = finding.quantity * item.qty
								making_amount = finding.making_amount
								finding_making_amount = (finding.making_rate * multiplied_qty)
								aggregated_repairing_items[key]["qty"] += multiplied_qty
								aggregated_repairing_items[key]["amount"] += finding_making_amount

								if aggregated_repairing_items[key]["qty"] > 0:
									aggregated_repairing_items[key]["rate"] = aggregated_repairing_items[key]["amount"] / aggregated_repairing_items[key]["qty"]
								else:
									aggregated_repairing_items[key]["rate"] = 0
								
								tax_rate_decimal = aggregated_repairing_items[key]["tax_rate"] / 100
								aggregated_repairing_items[key]["tax_amount"] += finding_making_amount * tax_rate_decimal
								aggregated_repairing_items[key]["amount_with_tax"] = (
									aggregated_repairing_items[key]["amount"] +
									aggregated_repairing_items[key]["tax_amount"]
								)
								break
						for e_item in e_invoice_items:
							if (e_item["is_for_labour"] ):
								key = (e_item["item_type"], e_item["uom"])
								if key not in aggregated_metal_labour_items:
									aggregated_metal_labour_items[key] = {
										"item_code": e_item["item_type"],
										"item_name": e_item["item_type"],
										"uom": e_item["uom"],
										"qty": 0,
										"rate": finding.making_rate,
										"amount": 0,
										"tax_rate": e_item["tax_rate"],
										"tax_amount": 0,
										"amount_with_tax": 0,
										"delivery_date": self.delivery_date
									}
								
								multiplied_qty = finding.quantity * item.qty
								making_amount = finding.making_amount
								finding_making_amount = (finding.making_rate * multiplied_qty)
								aggregated_metal_labour_items[key]["qty"] += multiplied_qty
								aggregated_metal_labour_items[key]["amount"] += finding_making_amount

								if aggregated_metal_labour_items[key]["qty"] > 0:
									aggregated_metal_labour_items[key]["rate"] = aggregated_metal_labour_items[key]["amount"] / aggregated_metal_labour_items[key]["qty"]
								else:
									aggregated_metal_labour_items[key]["rate"] = 0
								
								tax_rate_decimal = aggregated_metal_labour_items[key]["tax_rate"] / 100
								aggregated_metal_labour_items[key]["tax_amount"] += finding_making_amount * tax_rate_decimal
								aggregated_metal_labour_items[key]["amount_with_tax"] = (
									aggregated_metal_labour_items[key]["amount"] +
									aggregated_metal_labour_items[key]["tax_amount"]
								)
								break

		# Hybrid: making charge is always job-work value (Outwork rate),
		# regardless of is_customer_item — metal/finding making charges
		# above got aggregated into the Outright-taxed "making" buckets
		# by default, so for Hybrid fold them into the labour/
		# Outwork bucket instead of listing them separately.
		if self.sales_type == "Hybrid":
			for making_dict in (aggregated_metal_making_items, aggregated_finding_making_items):
				for key, val in making_dict.items():
					labour_item = next(
						(e for e in e_invoice_items if e["is_for_labour"] and e["uom"] == val["uom"]),
						None,
					)
					if not labour_item:
						continue
					target_key = (labour_item["item_type"], labour_item["uom"])
					if target_key not in aggregated_metal_labour_items:
						aggregated_metal_labour_items[target_key] = {
							"item_code": labour_item["item_type"],
							"item_name": labour_item["item_type"],
							"uom": labour_item["uom"],
							"qty": 0,
							"rate": 0,
							"amount": 0,
							"tax_rate": labour_item["tax_rate"],
							"tax_amount": 0,
							"amount_with_tax": 0,
							"delivery_date": val["delivery_date"],
						}
					target = aggregated_metal_labour_items[target_key]
					target["qty"]    += val["qty"]
					target["amount"] += val["amount"]
					tax_rate_decimal  = target["tax_rate"] / 100
					target["tax_amount"]     += val["amount"] * tax_rate_decimal
					target["amount_with_tax"] = target["amount"] + target["tax_amount"]
				making_dict.clear()

		# After aggregation, calculate average rate = total amount / total qty per key
		for key, val in aggregated_diamond_items.items():
			val["amount"] = round(val["amount"], precision)
			val["rate"] = val["amount"] / val["qty"] if val["qty"] else 0
			val["rate"] = round(val["rate"], precision)
			self.append("custom_invoice_item", val)
   
		for key, val in aggregated_metal_items.items():
			if val["qty"] > 0:
				
				average_rate = val["amount"] / val["qty"]
			else:
				average_rate = 0
			val["rate"] = average_rate
			val["rate"] = round(val["rate"], precision)
			val["amount"] = round(val["amount"], precision)
			self.append("custom_invoice_item", val)
   
		for key, val in aggregated_finding_items.items():
			val["amount"] = round(val["amount"], precision)
			val["rate"] = val["amount"] / val["qty"] if val["qty"] else 0
			val["rate"] = round(val["rate"], precision)
			
			self.append("custom_invoice_item", val)
   
		for key, val in aggregated_gemstone_items.items():
			val["amount"] = round(val["amount"], precision)
			val["rate"] = val["amount"] / val["qty"] if val["qty"] else 0
			val["rate"] = round(val["rate"], precision)
			
			self.append("custom_invoice_item", val)
   
		for key, val in aggregated_metal_making_items.items():
			val["amount"] = round(val["amount"], precision)
			val["rate"] = val["amount"] / val["qty"] if val["qty"] else 0
			val["rate"] = round(val["rate"], precision)
			self.append("custom_invoice_item", val)

		for key, val in aggregated_finding_making_items.items():
			val["amount"] = round(val["amount"], precision)
			val["rate"] = val["amount"] / val["qty"] if val["qty"] else 0
			val["rate"] = round(val["rate"], precision)
			
			self.append("custom_invoice_item", val)
   
		for key, val in aggregated_hallmarking_items.items():
			val["amount"] = round(val["amount"], precision)
			val["rate"] = val["amount"] / val["qty"] if val["qty"] else 0
			val["rate"] = round(val["rate"], precision)
			self.append("custom_invoice_item", val)

		for key, val in aggregated_metal_labour_items.items():
			val["amount"] = round(val["amount"], precision)
			val["rate"] = val["amount"] / val["qty"] if val["qty"] else 0
			val["qty"] = round(val["qty"],2)
			val["rate"] = round(val["rate"], precision)
			self.append("custom_invoice_item", val)
   
		for key, val in aggregated_repairing_items.items():
			val["rate"] = val["amount"] / val["qty"] if val["qty"] else 0
			val["rate"] = round(val["rate"], precision)
			val["amount"] = round(val["amount"], precision)
			self.append("custom_invoice_item", val)
		
		for key, val in aggregated_certification_items.items():
			val["rate"] = val["amount"] / val["qty"] if val["qty"] else 0
			self.append("custom_invoice_item", val)


def validate_quotation_item(self):
	if not self.custom_invoice_item:
		for row in self.items:
			if row.prevdoc_docname:
				quotation_id = row.prevdoc_docname
				invoice_items = frappe.get_all(
					"Quotation E Invoice Item",
					filters={"parent": quotation_id},
					fields=["item_code", "item_name", "uom", "qty", "rate", "amount"],
				)
				if invoice_items:
					for invoice_item in invoice_items:
						self.append(
							"custom_invoice_item",
							{
								"item_code": invoice_item.item_code,
								"item_name": invoice_item.item_name,
								"uom": invoice_item.uom,
								"qty": invoice_item.qty,
								"rate": invoice_item.rate,
								"amount": invoice_item.amount,
							},
						)


def validate_sales_type(self):
	for r in self.items:
		if not r.prevdoc_docname and not r.custom_customer_approval:
			pass
	if not self.sales_type:
		frappe.throw("Sales Type is mandatory.")


def fetch_sales_type_from_quotation(doc, method=None):
	"""When a Sales Order is mapped from a Quotation (Create > Sales Order),
	carry over the Quotation's Sales Type (Quotation.custom_sales_type ->
	Sales Order.sales_type). The mapper's generic same-fieldname copy misses
	this since the fieldnames differ; Order Type already copies fine on its
	own since both doctypes use the same fieldname.

	Runs as the mapper's postprocess step (target.run_method("set_missing_values")
	inside quotation.py's _make_sales_order), which fires server-side while the
	mapped doc is being built -- before it's sent to the browser -- so the value
	is already populated on the fresh, unsaved Sales Order form.

	Skips if sales_type is already set (never overrides a manual choice) or if
	there's no prevdoc_docname (Sales Order not created from a Quotation), so it
	never interferes with the Customer-based get_sales_type() flow.
	"""
	if doc.sales_type:
		return
	quotation = next((r.prevdoc_docname for r in doc.items if r.prevdoc_docname), None)
	if not quotation:
		return
	sales_type = frappe.db.get_value("Quotation", quotation, "custom_sales_type")
	if sales_type:
		doc.sales_type = sales_type


import json


@frappe.whitelist()
def make_sales_order_batch(sales_orders, target_doc=None):
	if isinstance(sales_orders, str):
		sales_orders = json.loads(sales_orders)

	if target_doc:
		if isinstance(target_doc, str):
			target_doc = json.loads(target_doc)

		target_doc = frappe.get_doc(target_doc)
	else:
		target_doc = frappe.new_doc("Sales Order")

	target_doc.items = []

	for so_name in sales_orders:
		so = frappe.db.get_value("Sales Order", so_name, "*", as_dict=True)
		if not so:
			continue
		target_doc.custom_diamond_quality = so.custom_diamond_quality
		target_doc.order_type = so.order_type
		target_doc.sales_type = so.sales_type
		target_doc.custom_parent_sales_order = so.name
		items = frappe.get_all(
			"Sales Order Item", filters={"parent": so_name}, fields="*"
		)

		for it in items:
			snc_list = frappe.db.get_list(
				"Serial Number Creator",
				filters={"sales_order_id": so_name},
				fields=["name"],
			)

			stock_entries = []
			for snc in snc_list:
				stock_entry = frappe.db.get_value(
					"Stock Entry", {"custom_serial_number_creator": snc.name}, "name"
				)
				if stock_entry:
					stock_entries.append(stock_entry)

			available_serials = []
			for stock_entry in stock_entries:
				serial_no = frappe.db.sql(
					f"""
					SELECT sed.serial_no, sed.item_code
					FROM `tabStock Entry Detail` sed
					WHERE sed.parent = '{stock_entry}'
					AND sed.item_code = '{it.item_code}'
					ORDER BY sed.idx DESC
					LIMIT 1
				""",
					as_dict=1,
				)

				if serial_no and serial_no[0]["item_code"] == it.item_code:
					available_serials.append(serial_no[0]["serial_no"])

			if not available_serials:
				continue

			serial_count = 0
			for s_no in available_serials:
				if serial_count < it.qty:
					target_doc.append(
						"items",
						{
							"item_code": it.item_code,
							"item_name": it.item_name,
							"serial_no": s_no,
							"bom": frappe.db.get_value(
								"Serial No", s_no, "custom_bom_no"
							),
							"diamond_quality": so.custom_diamond_quality,
							"description": it.description,
							"qty": 1,
							"rate": it.rate,
							"warehouse": it.warehouse,
							"against_sales_order": so_name,
							"uom": it.uom,
						},
					)
					serial_count += 1
				else:
					break

	first_so = frappe.db.get_value("Sales Order", sales_orders[0], "*", as_dict=True)

	return target_doc
