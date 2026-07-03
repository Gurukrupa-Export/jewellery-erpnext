# from hamcrest import none

import frappe
from frappe.utils import flt
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import io
from frappe.utils.file_manager import save_file


# def customer_tolerance(customer, weight):
# 	"""Get customer tolerance for a given weight"""
# 	tolerance = frappe.db.exists("Customer Attributes", customer)
# 	if tolerance:
# 		customer_tolerance = frappe.get_all(
# 			"Metal Tolerance Table", 
# 			{"parent": customer, "parenttype": "Customer Attributes"}, 
# 			["range_type", "from_weight", "to_weight", "tolerance_range", "minus_percent", "plus_percent"]
# 		)
# 		for tolerance_row in customer_tolerance:
# 			if tolerance_row.get('range_type') and tolerance_row.get('range_type') == 'Prcentage':
# 				return tolerance_row.tolerance_range
# 			if weight >= tolerance_row.from_weight and weight <= tolerance_row.to_weight:
# 				return tolerance_row.tolerance_range
# 	return 0

# def customer_making_details(customer, metal_rate_gst, metal_rate, bom_doc):
# 	"""
# 	Get making charge and wastage amount for a customer based on metal rate and BOM details.

# 	Returns:
# 		tuple: (making_amt, labour_charge, per_pc_labour_charge, wastage_amt)
# 	"""
# 	metal = bom_doc.metal_detail[0] if bom_doc.metal_detail else None

# 	# ── Early exits ──────────────────────────────────────────────────
# 	if not metal:
# 		return 0, 0, 0, 0

# 	# ── Fetch matching Making Charge Price record ────────────────────
# 	mc = frappe.get_all(
# 		"Making Charge Price",
# 		filters={
# 			"customer": customer,
# 			"metal_type": bom_doc.metal_type,
# 			"setting_type": bom_doc.setting_type,
# 			"from_gold_rate": ["<=", metal_rate_gst],
# 			"to_gold_rate": [">=", metal_rate_gst],
# 			"metal_touch": bom_doc.metal_touch,
# 		},
# 		fields=["name"],
# 		limit=1,
# 	)

# 	if not mc:
# 		return 0, 0, 0, 0

# 	# ── Fetch subcategory pricing info ───────────────────────────────
# 	sub_info = frappe.db.get_value(
# 		"Making Charge Price Item Subcategory",
# 		{
# 			"parent": mc[0]["name"],
# 			"subcategory": bom_doc.item_subcategory,
# 		},
# 		["rate_per_gm", "rate_per_pc", "wastage", "rate_per_gm_threshold"],
# 		as_dict=True,
# 	)

# 	if not sub_info:
# 		return 0, 0, 0, 0

# 	if not (metal.get("quantity") and metal.get("metal_type") and metal.get("metal_touch")):
# 		return 0, 0, 0, 0

# 	# ── Determine making rate based on weight threshold ──────────────
# 	threshold = flt(sub_info.rate_per_gm_threshold) or 2
# 	weight = flt(bom_doc.metal_and_finding_weight)
# 	is_below_threshold = weight < threshold

# 	labour_charge = flt(sub_info.rate_per_gm) or 0
# 	per_pc_labour_charge = flt(sub_info.rate_per_pc) or 0

# 	making_rate = flt(sub_info.rate_per_pc) if is_below_threshold else flt(sub_info.rate_per_gm)
# 	wastage_rate = 0 if is_below_threshold else flt(sub_info.wastage) / 100

# 	making_amt = making_rate if is_below_threshold else making_rate * metal.quantity

# 	# ── Calculate wastage using customer metal purity ────────────────
# 	customer_metal_purity = frappe.db.get_value(
# 		"Metal Criteria",
# 		{
# 			"parent": customer,
# 			"metal_type": metal.metal_type,
# 			"metal_touch": metal.metal_touch,
# 		},
# 		"metal_purity",
# 	)

# 	if not customer_metal_purity:
# 		return making_amt, labour_charge, per_pc_labour_charge, 0

# 	calculated_gold_rate = (flt(customer_metal_purity) * metal_rate_gst) / (100 + metal_rate)
# 	line_gold_amt = round(calculated_gold_rate * metal.quantity, 2)
# 	wastage_amt = line_gold_amt * wastage_rate

# 	return making_amt, labour_charge, per_pc_labour_charge, wastage_amt


@frappe.whitelist()
def download_cost_sheet(sales_order):
	"""Generate and download a COST Sheet Excel file with headers only."""
	customer_name = frappe.db.get_value('Sales Order', sales_order ,'customer_name')
	so_doc = frappe.get_doc("Sales Order", sales_order)
	if customer_name == "Caratlane Trading Private Limited":
		wb = Workbook()
		ws = wb.active
		ws.title = "COST Sheet"

		# ── Styling ──────────────────────────────────────────────────────
		header_font = Font(bold=True, size=10)
		center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
		thin_border = Border(
			left=Side(style="thin"),
			right=Side(style="thin"),
			top=Side(style="thin"),
			bottom=Side(style="thin"),
		)
		header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
		diamond_quality_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # Light yellow for diamond quality columns

		# ── Fetch Sales Order and Customer Data ──────────────────────────
		customer = so_doc.get("customer", "")
		metal_rate = so_doc.get('gold_rate', 0)
		metal_rate_gst = so_doc.get('gold_rate_with_gst', 0)

		# ── Fetch Customer Diamond Qualities (dynamic columns) ───────────
		customer_diamond_qualitys = frappe.get_all(
			"Customer Diamond Grade",
			{"parent": customer},
			["diamond_quality"],
			order_by="idx asc",  # Maintain consistent column order
		)

		# Extract quality names as a list for easy indexing
		diamond_quality_names = [row["diamond_quality"] for row in customer_diamond_qualitys]
		num_dq_cols = len(diamond_quality_names)  # Number of dynamic diamond quality columns

		# ── Calculate dynamic column offsets ─────────────────────────────
		# Fixed columns up to 52 remain unchanged.
		# Diamond Quality dynamic columns start at 53.
		# Columns after diamond qualities (Metal, Labor & Wastage) are shifted by num_dq_cols.

		DQ_START      = 53                          # First dynamic diamond quality column
		DQ_END        = 52 + num_dq_cols            # Last dynamic diamond quality column (inclusive)
		METAL_START   = DQ_END + 1                  # Was 53, now shifted
		METAL_END     = DQ_END + 2                  # Was 54
		LABOR_START   = DQ_END + 3                  # Was 55
		LABOR_END     = DQ_END + 7                  # Was 59

		# Individual column indices (used when writing data rows)
		COL_METAL_RATE  = DQ_END + 1   # "Rate"          (was 53)
		COL_METAL_VAL   = DQ_END + 2   # "Value"         (was 54)
		COL_TOTAL_DIA   = DQ_END + 3   # "Total Dia"     (was 55)
		COL_CENT_PER_GM = DQ_END + 4   # "Cent per gm"   (was 56)
		COL_LABOR       = DQ_END + 5   # "Labor"         (was 57)
		COL_PER_PC_LAB  = DQ_END + 6   # "Per Pc Labor"  (was 58)
		COL_WASTAGE     = DQ_END + 7   # "Wastage"       (was 59)
		COL_TOTAL       = DQ_END + 8   # "Total"         (was 60)
		COL_TOTAL_PRICE = DQ_END + 9   # "Total Price"   (was 61)
		COL_TECHNIQUE   = DQ_END + 10  # "Technique"     (was 62)

		# ── Row 1: Merged group headers ─────────────────────────────────
		merged_headers = [
			(10, 18, "Stone Breakup"),
			(19, 21, "Stone Quality and Price"),
			(22, 24, "Dia. & Gemstone"),
			(25, 28, "Component  Information For Metal"),
			(29, 34, "Finding /Chain  Information"),
			(35, 37, "Net Gold Weight Without Finding"),
			(38, 40, "Net Gold Weight With Finding"),
			(METAL_START, METAL_END, "Metal"),
			(LABOR_START, LABOR_END, "Labor & Wastage Total"),
		]

		# Add dynamic diamond quality group header (only if there are qualities)
		if num_dq_cols > 0:
			merged_headers.append((DQ_START, DQ_END, "Diamond Quality Price"))

		for col_start, col_end, label in merged_headers:
			cell = ws.cell(row=1, column=col_start, value=label)
			cell.font = header_font
			cell.alignment = center_align
			cell.border = thin_border

			# Use a distinct fill for the diamond quality group header
			if label == "Diamond Quality Price":
				cell.fill = diamond_quality_fill
			else:
				cell.fill = header_fill

			if col_start != col_end:
				ws.merge_cells(
					start_row=1,
					start_column=col_start,
					end_row=1,
					end_column=col_end,
				)

		# ── Row 2: Fixed column headers (1–52) ───────────────────────────
		row2_headers = {
			1: "Sr. No.",
			2: "Caratlane \nSKU Code",
			3: "ERP Variant Name",
			4: "Vendor Style Code",
			5: "Images",
			6: "Gold Kt",
			7: "Gold Colour",
			8: "Product Type",
			9: "Ring/Bangle/Bracelet/\nNecklace Size",
			10: "Stone Type",
			11: "Dia Sieve/Col Stone",
			12: "Shape",
			13: "MM Size",
			14: "Quantity",
			15: "Individual Stone Wt",
			16: "Total Stone Wt",
			17: "Setting Type",
			18: "Type",
			19: "Stone Quality",
			20: "Colour",
			21: "Cut",
			22: "Rate PCT",
			23: "Value",
			24: "Gross weight",
			25: "Colour",
			26: "Karat ",
			27: "Quantity",
			28: " Gold weight",
			29: "Name ( specify the finding type)",
			30: "Quantity",
			31: "Weight",
			32: "Colour",
			33: "Karat",
			34: "Finding Type",
			35: "Min",
			36: "Avg",
			37: "Max",
			38: "Min",
			39: "Avg",
			40: "Max",
			41: "Finishing Information. (Rhodium/ Plating)",
			42: "Shipping Days",
			43: "Size",
			44: "Theme",
			45: "Occasion",
			46: "14kt",
			47: "18kt",
			48: "22kt",
			49: "Platinum",
			50: "White Variance",
			51: "Yellow Variance",
			52: "Pink Variance",
		}

		# ── Row 2: Append Dynamic Diamond Quality Headers (53 … 52+N) ────
		for i, quality_name in enumerate(diamond_quality_names):
			col_idx = DQ_START + i
			row2_headers[col_idx] = quality_name

		# ── Row 2: Append Shifted Static Headers after diamond qualities ──
		row2_headers[COL_METAL_RATE]  = "Rate"
		row2_headers[COL_METAL_VAL]   = "Value"
		row2_headers[COL_TOTAL_DIA]   = "Total Dia"
		row2_headers[COL_CENT_PER_GM] = "Cent per gm"
		row2_headers[COL_LABOR]       = "Labor"
		row2_headers[COL_PER_PC_LAB]  = "Per Pc Labor"
		row2_headers[COL_WASTAGE]     = "Wastage"
		row2_headers[COL_TOTAL]       = "Total"
		row2_headers[COL_TOTAL_PRICE] = "Total Price"
		row2_headers[COL_TECHNIQUE]   = "Technique"

		# ── Write Row 2 headers to worksheet ────────────────────────────
		for col_idx, header_text in row2_headers.items():
			cell = ws.cell(row=2, column=col_idx, value=header_text)
			cell.font = header_font
			cell.alignment = center_align
			cell.border = thin_border

			# Distinct fill for dynamic diamond quality columns
			if DQ_START <= col_idx <= DQ_END:
				cell.fill = diamond_quality_fill
			else:
				cell.fill = header_fill

		# ── Styling for data rows ────────────────────────────────────────
		data_font = Font(size=10)
		bold_font = Font(bold=True, size=10)
		total_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

		# ── Data Rows ────────────────────────────────────────────────────
		current_row = 3  # data starts after header rows
		sr_no = 0

		for soi in so_doc.items:
			if not soi.bom:
				continue

			sr_no += 1
			bom = frappe.get_doc("BOM", soi.bom)

			if not bom:
				continue

			# ── Item-level data ──────────────────────────────────────────
			metal = bom.metal_detail[0] if bom.metal_detail else None
			finding = bom.finding_detail[0] if bom.finding_detail else None

			# Product Size from Item Variant Attributes
			product_size = ""
			try:
				if bom.get("product_size") and bom.get("sizer_type"):
					product_size = f"{bom.get('product_size')} {bom.get('sizer_type')}".strip()
				if not product_size:
					item_doc = frappe.get_doc("Item", soi.item_code)
					size_val = ""
					sizer_type = ""
					for attr in item_doc.attributes:
						if attr.attribute == "Product Size":
							size_val = attr.attribute_value or ""
						elif attr.attribute == "Sizer Type":
							sizer_type = attr.attribute_value or ""
					if size_val:
						product_size = f"{size_val} {sizer_type}".strip()
			except Exception:
				product_size = ""


			# ── Build stone rows ─────────────────────────────────────────
			stone_rows = []

			# Diamond rows
			for d in bom.diamond_detail:
				pcs = flt(d.pcs) or 0
				qty = flt(d.quantity) or 0
				individual_wt = flt(qty / pcs, 4) if pcs else 0
				setting = bom.get('sub_setting_type1', '').replace("Setting", "").replace("setting", "").strip()

				stone_rows.append({
					"stone_type": "Diamond",
					"sieve_col_stone": d.sieve_size_range or "",
					"shape": d.stone_shape or "",
					"mm_size": "",
					"quantity": pcs,
					"quality": d.quality or "",
					"total_stone_wt": qty,
					"individual_stone_wt": individual_wt,
					"setting_type": setting,
					"type": "Wax",
					"total_diamond_rate": d.total_diamond_rate,
				})

			# Gemstone rows
			for g in bom.gemstone_detail:
				pcs = flt(g.pcs) or 0
				qty = flt(g.quantity) or 0
				individual_wt = flt(qty / pcs, 4) if pcs else 0
				setting = bom.get('sub_setting_type1', '').replace("Setting", "").replace("setting", "").strip()

				stone_rows.append({
					"stone_type": g.gemstone_type or "",
					"sieve_col_stone": g.gemstone_size or "",
					"shape": g.stone_shape or "",
					"mm_size": g.gemstone_size or "",
					"quantity": pcs,
					"total_stone_wt": qty,
					"individual_stone_wt": individual_wt,
					"setting_type": setting,
					"type": "",
				})

			# If BOM has no stones, still create one row for the item
			if not stone_rows:
				stone_rows.append({
					"stone_type": "", "sieve_col_stone": "", "shape": "",
					"mm_size": "", "quantity": 0, "total_stone_wt": 0,
					"individual_stone_wt": 0, "setting_type": "", "type": "",
				})

			# ── Write stone rows ─────────────────────────────────────────
			total_qty = 0
			total_stone_wt = 0

			for idx, stone in enumerate(stone_rows):
				row_data = {}
				is_first_row = (idx == 0)

				# Stone columns (filled on every row)
				row_data[10] = stone["stone_type"]
				row_data[11] = stone["sieve_col_stone"]
				row_data[13] = stone["mm_size"]
				row_data[14] = stone["quantity"]
				row_data[15] = stone["individual_stone_wt"]
				row_data[16] = stone["total_stone_wt"]
				row_data[18] = stone["type"]

				stone_quality = stone.get("quality", "").split("-")
				row_data[19] = stone_quality[0] if len(stone_quality) > 0 else ""
				row_data[20] = stone_quality[1] if len(stone_quality) > 1 else ""
				row_data[21] = "Brilliant"

				row_data[22] = stone.get("total_diamond_rate", '')
				row_data[23] = row_data[22] * row_data[16] if row_data[22] and row_data[16] else ''

				# Accumulate totals
				total_qty += flt(stone["quantity"])
				total_stone_wt += flt(stone["total_stone_wt"])

				if is_first_row:
					# Item info
					row_data[1] = sr_no
					row_data[3] = soi.item_code
					row_data[6] = metal.metal_touch if metal else ""
					row_data[7] = metal.metal_colour if metal else ""
					row_data[8] = bom.item_category or ""
					row_data[9] = product_size

					row_data[12] = stone["shape"]
					row_data[17] = stone["setting_type"]

					row_data[24] = flt(bom.gross_weight)

					row_data[25] = metal.metal_colour if metal else ""
					row_data[26] = metal.metal_touch if metal else ""
					row_data[27] = 1
					row_data[28] = flt(metal.quantity) if metal else ""

					row_data[29] = finding.finding_type if finding else ""
					row_data[30] = finding.qty if finding else ""
					row_data[31] = finding.quantity if finding else ""
					row_data[32] = finding.metal_colour if finding else ""
					row_data[33] = finding.metal_touch if finding else ""
					row_data[34] = finding.metal_type if finding else ""

					tolerance_range = customer_tolerance(customer, row_data[28])

					tolerance_amount_without_finding = (row_data[28] or 0) * (tolerance_range / 100)
					row_data[36] = row_data[28] or 0
					row_data[35] = ((row_data[28] or 0) - tolerance_amount_without_finding) or ""
					row_data[37] = ((row_data[28] or 0) + tolerance_amount_without_finding) or ""

					row_data[39] = (row_data[36] or 0) + (row_data[31] or 0)
					tolerance_amount_with_finding = (row_data[39] or 0) * (tolerance_range / 100)
					row_data[38] = ((row_data[39] or 0) - tolerance_amount_with_finding) or ""
					row_data[40] = ((row_data[39] or 0) + tolerance_amount_with_finding) or ""

					row_data[43] = ""
					row_data[44] = ""
					row_data[45] = ""
					row_data[46] = "Yes" if row_data[33] and row_data[33].lower() == "14kt" else "No"
					row_data[47] = "Yes" if row_data[33] and row_data[33].lower() == "18kt" else "No"
					row_data[48] = "Yes" if row_data[33] and row_data[33].lower() == "22kt" else "No"
					row_data[49] = "Yes" if row_data[33] and row_data[33].lower() == "platinum" else "No"
					row_data[50] = "Yes" if row_data[7] and row_data[7].lower() == "white" else "No"
					row_data[51] = "Yes" if row_data[7] and row_data[7].lower() == "yellow" else "No"
					row_data[52] = "Yes" if row_data[7] and row_data[7].lower() == "pink" else "No"


					# ── Shifted Metal & Labor columns ─────────────────────
					row_data[COL_METAL_RATE]  = metal_rate
					row_data[COL_TOTAL_DIA]   = flt(bom.diamond_weight)
					row_data[COL_CENT_PER_GM] = (
						round(row_data[COL_TOTAL_DIA] / row_data[39], 3)
						if row_data[39] else ""
					)

					if metal_rate_gst:
						customer_price_data = customer_making_details(customer, metal_rate_gst, metal_rate, bom)
						row_data[COL_LABOR]      = customer_price_data[1]
						row_data[COL_PER_PC_LAB] = customer_price_data[2]
						row_data[COL_WASTAGE]    = customer_price_data[3]
						row_data[COL_TOTAL]      = customer_price_data[0]
				
				# ── Dynamic Diamond Quality Columns (Yes/No) ──────────────────────
				stone_quality_raw = stone.get("quality", "")
				for i, quality_name in enumerate(diamond_quality_names):
					col_idx = DQ_START + i
					row_data[col_idx] = "Yes" if stone_quality_raw == quality_name else "No"

				# Write the row to worksheet
				for col_idx, value in row_data.items():
					cell = ws.cell(row=current_row, column=col_idx, value=value)
					cell.font = data_font
					cell.alignment = center_align

					# Apply distinct fill to diamond quality data cells
					# if DQ_START <= col_idx <= DQ_END:
					# 	cell.fill = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")

				current_row += 1

			# ── Total Count row ──────────────────────────────────────────
			total_row_data = {
				1: sr_no,
				10: "Total Count",
				14: total_qty,
				16: total_stone_wt,
			}

			for col_idx, value in total_row_data.items():
				cell = ws.cell(row=current_row, column=col_idx, value=value)
				cell.font = bold_font
				cell.alignment = center_align
				cell.border = thin_border
				cell.fill = total_fill

			current_row += 1
			current_row += 2  # 2 empty rows between items for visual separation

		# ── Auto-adjust column widths ────────────────────────────────────
		total_cols = COL_TECHNIQUE  # last column index
		for col_idx in range(1, total_cols + 1):
			col_letter = get_column_letter(col_idx)
			header_val = row2_headers.get(col_idx, "")
			max_line_len = max((len(line) for line in str(header_val).split("\n")), default=8)
			ws.column_dimensions[col_letter].width = max(max_line_len + 4, 10)

		# ── Set row heights ──────────────────────────────────────────────
		ws.row_dimensions[1].height = 30
		ws.row_dimensions[2].height = 40

		# ── Write to response ────────────────────────────────────────────
		file_data = io.BytesIO()
		wb.save(file_data)
		file_data.seek(0)

		filename = f"COST_Sheet_{sales_order}.xlsx"

		frappe.response["filename"] = filename
		frappe.response["filecontent"] = file_data.getvalue()
		frappe.response["type"] = "download"
  
  
	if customer_name == "Novel Jewels Limited":
		import io
		from openpyxl import Workbook
		from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

		workbook = Workbook()
		sheet = workbook.active
		sheet.title = "Proto Sheet"

		headers = [
			"sr","Vendor Name","PO No","Vendor Design No","Vendor Ref No","Gold Kt",
			"Product Size","Gold Colour","Complexity","Complexity (code)",
			"Findings","Category","Sub Category", "Group","Sub Group","SKU",
			"HSN CODE ( 8 digit)","DESCRIPTION","Qty","Gross Wt","Net Wt","Wt UOM","Rate UOM",
			"Rate","Value","Labour/ Dia Handling / Other per ct or gm charges","Labour Amt","Loss / CS Handling / Other %",
			"Loss Amt","Stone Shape","Setting Type","Stone Cut/ Cab","Stone Quality",
			"Stone Color","Stone Size (mm)","Stone Sieve Size","Addional Charges",
		]

		sheet.append(headers)

		# -------------------- STYLES --------------------
		blue_fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")    # Light Blue
		yellow_fill = PatternFill(start_color="FFF200", end_color="FFF200", fill_type="solid")  # Yellow
		orange_fill = PatternFill(start_color="F4B183", end_color="F4B183", fill_type="solid")  # Orange

		header_font = Font(bold=True, color="000000")
		center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

		thin_border = Border(
			left=Side(style='thin'),
			right=Side(style='thin'),
			top=Side(style='thin'),
			bottom=Side(style='thin')
		)

		# Header Style Default Blue
		for cell in sheet[1]:
			cell.font = header_font
			cell.fill = blue_fill
			cell.alignment = center_alignment
			cell.border = thin_border

		# Header Yellow Columns (DESCRIPTION, Qty, Gross Wt, Net Wt)
		# DESCRIPTION = 19, Qty = 20, Gross Wt = 21, Net Wt = 22
		yellow_columns = [19, 20, 21, 22]
		for col in yellow_columns:
			sheet.cell(row=1, column=col).fill = yellow_fill

		# Header Orange Columns (Value, Labour, Loss)
		# Value=26, Labour Charges=27, Labour Amt=28, Loss %=29, Loss Amt=30
		orange_columns = [26, 27, 28, 29, 30]
		for col in orange_columns:
			sheet.cell(row=1, column=col).fill = orange_fill

		sheet.row_dimensions[1].height = 40

		# Column Widths
		sheet.column_dimensions['A'].width = 8
		sheet.column_dimensions['B'].width = 18
		sheet.column_dimensions['C'].width = 25
		sheet.column_dimensions['D'].width = 30
		sheet.column_dimensions['E'].width = 18
		sheet.column_dimensions['F'].width = 35
		sheet.column_dimensions['G'].width = 15
		sheet.column_dimensions['H'].width = 25
		sheet.column_dimensions['I'].width = 25
		sheet.column_dimensions['J'].width = 25
		sheet.column_dimensions['K'].width = 25
		sheet.column_dimensions['L'].width = 25
		sheet.column_dimensions['M'].width = 25
		sheet.column_dimensions['N'].width = 25
		sheet.column_dimensions['O'].width = 25
		sheet.column_dimensions['P'].width = 25
		sheet.column_dimensions['Q'].width = 25
		sheet.column_dimensions['R'].width = 25
		sheet.column_dimensions['S'].width = 25
		sheet.column_dimensions['T'].width = 25
		sheet.column_dimensions['U'].width = 25
		sheet.column_dimensions['V'].width = 25
		sheet.column_dimensions['W'].width = 25
		sheet.column_dimensions['X'].width = 25
		sheet.column_dimensions['Y'].width = 25
		sheet.column_dimensions['Z'].width = 25
		sheet.column_dimensions['AA'].width = 25
		sheet.column_dimensions['AB'].width = 25
		sheet.column_dimensions['AC'].width = 25
		sheet.column_dimensions['AD'].width = 25
		sheet.column_dimensions['AE'].width = 25
		sheet.column_dimensions['AF'].width = 25
		sheet.column_dimensions['AG'].width = 25
		sheet.column_dimensions['AH'].width = 25
		sheet.column_dimensions['AI'].width = 25
		sheet.column_dimensions['AJ'].width = 25
		sheet.column_dimensions['AK'].width = 25
		sheet.column_dimensions['AL'].width = 25

		# -------------------- DATA INSERT --------------------
		for row in so_doc.items:
			total_value = 0
			lbr_amount = 0
			wastage_amt = 0

			metal_touch = frappe.db.get_value('BOM', row.get('bom'), 'metal_touch')
			product_size = frappe.db.get_value('BOM', row.get('bom'), 'product_size')
			metal_colour = frappe.db.get_value('BOM', row.get('bom'), 'metal_colour')
			setting_type = frappe.db.get_value('BOM', row.get('bom'), 'setting_type')
			category_item = frappe.db.get_value('BOM', row.get('bom'), 'item_category')
			subcategory = frappe.db.get_value('BOM', row.get('bom'), 'item_subcategory')
			diamond_qlty =  row.get('diamond_quality')
			if diamond_qlty:
				fg, si = diamond_qlty.split("-")
			else:
				fg = ""
				si = ""


			bom_finding = frappe.get_all("BOM Finding Detail",
				filters={'parent': row.get('bom')},
				fields=['finding_type', 'finding_category']
			)

			finding = []
			if bom_finding:
				for item in bom_finding:
					fnd = frappe.get_value("Customer Finding Detail", {
						'gk_finding_sub_category': item.get('finding_type'),
						'gk_finding_category': item.get('finding_category'),
						'parent': so_doc.customer,
					}, 'description_finding')

					if fnd:
						finding.append(fnd)

			category = frappe.get_all("Customer Category Detail", filters={
				'gk_sub_category': frappe.db.get_value('BOM', row.get('bom'), 'item_subcategory'),
				'gk_category': frappe.db.get_value('BOM', row.get('bom'), 'item_category'),
				'parent': so_doc.customer,
			}, fields=['customer_subcategory', 'customer_category'], limit_page_length=1)

			digit_code = frappe.get_all("Item Theme Code Detail", filters={
				'customer': so_doc.customer,
				'parent': row.get('item_code'),
			}, fields=['theme_code', 'digit_15code'], limit_page_length=1)

			metal_detail = frappe.get_all("BOM Metal Detail", filters={
				'parent': row.get('bom')
			}, fields=['*'])
			prodct_size = frappe.db.get_value('Novel Size Master',
                            {
                                'customer':so_doc.customer,
                                'item_category':category_item,
                                'product_size_in':product_size
                                },"product_size")
			making_charge =  frappe.db.get_all('Making Charge Price',filters = {
				'customer':so_doc.customer,
				'setting_type':setting_type,
				'metal_touch':metal_touch,
			},fields=['name'])
			making_charges = None
			if making_charge:
				making_charges = frappe.db.get_all("Making Charge Price Item Subcategory",filters={
							"parent": making_charge[0].name,
							"subcategory":subcategory
						},fields=['mfg_complexity_code','wastage'])
			making_code = None
			if making_charges:
				if making_charges[0].get('mfg_complexity_code') == "Simple":
					making_code = 'A'
				elif making_charges[0].get('mfg_complexity_code') == "Medium":
					making_code = 'B'
				elif making_charges[0].get('mfg_complexity_code') == "Complex":
					making_code = 'C'
			total_value = total_value + metal_detail[0].get('amount') if metal_detail else 0 
			lbr_amount = lbr_amount + metal_detail[0].get('making_amount') if metal_detail else 0
			wastage_amt = wastage_amt + metal_detail[0].get('wastage_amount') if metal_detail else 0
			row_data = [
				row.get('idx'),
				so_doc.get('company'),
				"VC-0436",
				row.get('serial_no') or '',
				row.get('item_code') or '',
				metal_touch or '',
				prodct_size or '',
				metal_colour or '',
				making_charges[0].get('mfg_complexity_code') if making_charges else "",
				making_code,
				", ".join(finding) if finding else '',
				category[0].get('customer_category') if category else "",
				category[0].get('customer_subcategory') if category else "",
				"Studded Group",
				"Studded - DIS",
				digit_code[0].get('digit_15code') if digit_code else "",
				digit_code[0].get('theme_code') if digit_code else "",
				frappe.db.get_value('BOM', row.get('bom'), 'item_category'),
				"",
				frappe.db.get_value('BOM', row.get('bom'), 'gross_weight'),
				metal_detail[0].get('quantity') if metal_detail else "",
				"GM",
				"WT",
				metal_detail[0].get('rate') if metal_detail else "",
				metal_detail[0].get('amount') if metal_detail else "",
				metal_detail[0].get('making_rate') if metal_detail else "",
				metal_detail[0].get('making_amount') if metal_detail else "",
				making_charges[0].get('wastage') if making_charges else "",
				metal_detail[0].get('wastage_amount') if metal_detail else "",
				"",
				"",
				"",
				"",
				"",
				"",
				"",
				"",
				metal_detail[0].get('amount') if metal_detail else 0 + metal_detail[0].get('making_amount') if metal_detail else 0,
			]

			sheet.append(row_data)

			# Apply Borders and Alignment
			last_row = sheet.max_row
			for col in range(1, len(headers) + 1):
				cell = sheet.cell(row=last_row, column=col)
				cell.alignment = center_alignment
				cell.border = thin_border

				# Apply Orange fill to Value/Labour/Loss columns
				if col in orange_columns:
					cell.fill = orange_fill


			# -------------------- DIAMOND DETAILS --------------------
			diamond_detail = frappe.get_all("BOM Diamond Detail", filters={
				'parent': row.get('bom')
			}, fields=['*'])
			total_se_rate = 0
   
			for item in diamond_detail:
				total_se_rate = total_se_rate + item.get('se_rate')
				total_value = total_value + item.get('diamond_rate_for_specified_quantity') or 0 
				stone_code = frappe.db.sql("""
					SELECT rm.customer_code
					FROM `tabCustomer RM Code Detail` rm
					LEFT JOIN `tabCustomer RM Code` rmc ON rm.parent = rmc.name
					WHERE 
						IFNULL(rm.stone_shape, '') = IFNULL(%s, '')
						AND rmc.customer = %s
						AND IFNULL(rm.diamond_type, '') = IFNULL(%s, '')
						AND IFNULL(rm.diamond_quality, '') = IFNULL(%s, '')
						AND CAST(IFNULL(rm.size_in_mm, 0) AS DECIMAL(10,3)) = CAST(%s AS DECIMAL(10,3))
							
					""",
					(
						item.get("stone_shape"),
						so_doc.customer,
						item.get("diamond_type"),
						row.get('diamond_quality'),
						item.get("size_in_mm"),
					),
					as_dict=1)
				row_data = [
					"", "", "", "", "", "", "", "", "", "", "", "", "", "",
     				"",
					stone_code[0].get('customer_code') if stone_code else "",
         			"",
					"Diamond",
					item.get('pcs') or '',
					"",
					item.get('quantity') or '',
					"CT",
					"WT",
					item.get('total_diamond_rate') or '',
					item.get('diamond_rate_for_specified_quantity') or 0,
					item.get('handling_rate') or '',
					"",
					"",
					"",
					item.get('stone_shape') or '',
					setting_type or '',
					"",
					si,
					fg,
					str(item.get('size_in_mm')) + " MM" or '',
					item.get('diamond_sieve_size') or '',
					"",
					item.get('diamond_rate_for_specified_quantity') or 0,
				]

				sheet.append(row_data)

				last_row = sheet.max_row
				for col in range(1, len(headers) + 1):
					cell = sheet.cell(row=last_row, column=col)
					cell.alignment = center_alignment
					cell.border = thin_border

					# Orange fill for Value/Labour/Loss
					if col in orange_columns:
						cell.fill = orange_fill


			# -------------------- GEMSTONE DETAILS --------------------
			gemstone_detail = frappe.get_all("BOM Gemstone Detail", filters={
				'parent': row.get('bom')
			}, fields=['*'])
			total_se_rate = 0
			for item in gemstone_detail:
				total_se_rate = total_se_rate + item.get('se_rate')
				total_value = total_value + item.get('se_rate') or 0,
				row_data = [
					"", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "",
					item.get('gemstone_type') or '',
					item.get('pcs') or '',
					"",
					item.get('quantity') or '',
					"CT",
					"WT",
					item.get('total_gemstone_rate') or '',
					item.get('se_rate') or '',
					"",
					"",
					"",
					"",
					item.get('stone_shape') or '',
					setting_type or '',
					item.get('cut_or_cab') or '',
					"",
					"",
					item.get('gemstone_size') or '',
					"",
					"",
					item.get('se_rate')  or 0,

				]

				sheet.append(row_data)

				last_row = sheet.max_row
				for col in range(1, len(headers) + 1):
					cell = sheet.cell(row=last_row, column=col)
					cell.alignment = center_alignment
					cell.border = thin_border

					# Orange fill for Value/Labour/Loss
					if col in orange_columns:
						cell.fill = orange_fill
      
			row_data = [
					"", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "",
					'',
					'',
					"",
					'',
					"",
					"",
					'',
					total_value,
					"",
					lbr_amount,
					"",
					wastage_amt,
					'',
					'',
					'',
					"",
					"",
					'',
					"",
					"",
					"",

				]

			sheet.append(row_data)
			total_row = sheet.max_row
			for col in range(1, len(headers) + 1):
				cell = sheet.cell(row=total_row, column=col)
				cell.fill = yellow_fill
				cell.border = thin_border
				cell.alignment = center_alignment


		# -------------------- TOTAL ROW (FULL YELLOW) --------------------
		# Optional total row like screenshot
		# This will just highlight last row if needed, you can remove if not required.

		# Example: If you want to highlight every blank row between items then you can do later.

		# Check if row inserted
		if sheet.max_row <= 1:
			frappe.throw("Cost Sheet Can Not Download")

		# Save workbook into memory
		output = io.BytesIO()
		workbook.save(output)
		output.seek(0)

		filename = f"COST_Sheet_{sales_order}.xlsx"

		frappe.response["filename"] = filename
		frappe.response["filecontent"] = output.getvalue()
		frappe.response["type"] = "download"
  
  
  
  
		
	# if customer_name.strip() == "Reliance Retail Limited":
	# 	import io
	# 	from openpyxl import Workbook
	# 	from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

	# 	workbook = Workbook()
	# 	sheet = workbook.active
	# 	sheet.title = "Proto Sheet"

	# 	headers = [
	# 		"Sr no.","Design Image","DESIGNS","Vendor Design Code","Vendor Name","VENDOR CODE(As per Axepta)",
	# 		"Metal (Gold/Silver/Platinum)","Purity","Metal Color(White/Geru/Rose Gold/Black)",
	# 		"Product Category(Bangle/SET/MS/FingerRing)","Sub-product(bali, hanging, studs,Gents or ladies ring etc)",
	# 		"Article code(BAN/BLT/SET/MSR)","RJ ref size code","Finding type(70 / 71 / 72 / 75)",
	# 		"UOM (1/B1/B2/B4/B6)","COMPLEXITY CODE","Manufacturing code",
	# 		"Production route(handmade /Casting/Machinemade)","14kt gross wt","14Kt nett wt","Color Stone pcs",
	# 		"Color Stone Wt","Diamond Pcs","Diamond Wt","Stone Name",
	# 		"Shape(Round / Baguette / Princess / Tapers / MQ / Rose / Pear)","Item number","Code",
	# 		"Group Size","Child Sieve Size","Stone Pcs","Avg Wt","Stone Wt",
	# 		"UOM(Cts/Gms)","Plating","SettingType(As per Master)","Plating Type(As per Master)",
	# 		"Stone Rate/Cts","Making Charge per/gram","Wastage %","Hallmarking","TAG","GK-remark"
	# 	]

	# 	sheet.append(headers)

	# 	# Styles
	# 	blue_fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")
	# 	yellow_fill = PatternFill(start_color="FFF200", end_color="FFF200", fill_type="solid")
	# 	orange_fill = PatternFill(start_color="F4B183", end_color="F4B183", fill_type="solid")

	# 	header_font = Font(bold=True, color="000000")
	# 	center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

	# 	thin_border = Border(
	# 		left=Side(style='thin'),
	# 		right=Side(style='thin'),
	# 		top=Side(style='thin'),
	# 		bottom=Side(style='thin')
	# 	)

	# 	for cell in sheet[1]:
	# 		cell.font = header_font
	# 		cell.fill = blue_fill
	# 		cell.alignment = center_alignment
	# 		cell.border = thin_border

	# 	yellow_columns = [19, 20, 21, 22]
	# 	for col in yellow_columns:
	# 		sheet.cell(row=1, column=col).fill = yellow_fill

	# 	orange_columns = [26, 27, 28, 29, 30]
	# 	for col in orange_columns:
	# 		sheet.cell(row=1, column=col).fill = orange_fill

	# 	sheet.row_dimensions[1].height = 40

	# 	for row in so_doc.items:

	# 		metal_type = frappe.db.get_value('BOM', row.get('bom'), 'metal_type')
	# 		metal_touch = frappe.db.get_value('BOM', row.get('bom'), 'metal_touch')
	# 		metal_color = frappe.db.get_value('BOM', row.get('bom'), 'metal_colour')
	# 		item_category = frappe.db.get_value('BOM', row.get('bom'), 'item_category')
	# 		item_sub_category = frappe.db.get_value('BOM', row.get('bom'), 'item_subcategory')
	# 		setting_type = frappe.db.get_value('BOM', row.get('bom'), 'setting_type')
	# 		making_charge = frappe.db.get_value('BOM', row.get('bom'), 'making_charge')

	# 		category = frappe.db.get_value(
	# 			"Customer Category Detail",
	# 			{
	# 				'gk_category': item_category,
	# 				'parent': so_doc.customer
	# 			},
	# 			'customer_category'
	# 		)

	# 		sub_category = frappe.db.get_all(
	# 			"Customer Category Detail",
	# 			{
	# 				'gk_category': item_category,
	# 				"gk_sub_category": item_sub_category,
	# 				'parent': so_doc.customer
	# 			},
	# 			['customer_subcategory', 'code_category']
	# 		)

	# 		diamond_pcs = frappe.db.get_all(
	# 			"BOM Diamond Detail",
	# 			filters={"parent": row.get("bom")},
	# 			fields=["*"]
	# 		)
	# 		finding_code = []
	# 		finding_pcs = frappe.db.get_all(
	# 			"BOM Finding Detail",
	# 			filters={"parent": row.get("bom")},
	# 			fields=["*"]
	# 		)
	# 		if finding_pcs:
	# 			for i in finding_pcs:
	# 				sub_category = frappe.db.get_all(
	# 					"Customer Finding Detail",
	# 					{
	# 						"gk_finding_sub_category": i.get('finding_type'),
	# 						'parent': so_doc.customer
	# 					},
	# 					[ 'code_finding']
	# 				)
	# 				if sub_category:
	# 					finding_code.append(sub_category[0].get('code_finding'))
						

	# 		first_diamond = diamond_pcs[0] if diamond_pcs else None

	# 		# MAIN ROW (FIRST DIAMOND DETAIL IN SAME ROW)
	# 		row_data = [
	# 			row.idx,
	# 			"",
	# 			"GK",
	# 			row.item_code,
	# 			"GURUKRUPA EXPORT PRIVATE LIMITED",
	# 			"60450001",
	# 			metal_type or "",
	# 			metal_touch or "",
	# 			metal_color or "",
	# 			category if category else "",
	# 			sub_category[0].get('customer_subcategory') if sub_category else "",
	# 			sub_category[0].get('code_category') if sub_category else "",
	# 			"",
	# 			", ".join(finding_code) if finding_code else '',
	# 			"",
	# 			"",
	# 			"",
	# 			"",
	# 			frappe.db.get_value('BOM', row.get('bom'), 'gross_weight') or "",
	# 			frappe.db.get_value('BOM', row.get('bom'), 'metal_and_finding_weight') or "",
	# 			"",  # Color Stone pcs
	# 			"",  # Color Stone Wt
	# 			first_diamond.get("pcs") if first_diamond else "",
	# 			first_diamond.get("quantity") if first_diamond else "",
	# 			"Diamond" if first_diamond else "",
	# 			first_diamond.get("stone_shape") if first_diamond else "",
	# 			"",
	# 			"",
	# 			first_diamond.get("diamond_sieve_size") if first_diamond else "",
	# 			first_diamond.get("sieve_size_range") if first_diamond else "",
	# 			first_diamond.get("pcs") if first_diamond else "",
	# 			first_diamond.get("weight_per_pcs") if first_diamond else "",
	# 			first_diamond.get("quantity") if first_diamond else "",
	# 			"CTS" if first_diamond else "",
	# 			"",
	# 			setting_type if setting_type else '',
	# 			"",
	# 			"",
	# 			making_charge if making_charge else "",
	# 			"",
	# 			"",
	# 			row.get('serial_no' or ''),
	# 			"",
	# 		]

	# 		sheet.append(row_data)

	# 		# REMAINING DIAMOND DETAILS IN NEXT ROWS (INDEX 1 ONWARDS)
	# 		for d in diamond_pcs[1:]:
	# 			diamond_row = [
	# 				"", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "",
	# 				"Diamond",
	# 				d.get("stone_shape") or "",
	# 				"",
	# 				"",
	# 				d.get('diamond_sieve_size') or "",
	# 				d.get('sieve_size_range') or "",
	# 				d.get('pcs') or "",
	# 				d.get('weight_per_pcs') or "",
	# 				d.get('quantity') or "",
	# 				"CTS",
	# 				"",
	# 				setting_type or "",
	# 				"",
	# 				"",
	# 				"",
	# 				"",
	# 				"",
	# 				"",
	# 				"",
	# 				"",
	# 			]
	# 			sheet.append(diamond_row)

	# 	# Add blank row at end
	# 	blank_row = [""] * len(headers)
	# 	sheet.append(blank_row)

	# 	last_row = sheet.max_row
	# 	for col in range(1, len(headers) + 1):
	# 		cell = sheet.cell(row=last_row, column=col)
	# 		cell.alignment = center_alignment
	# 		cell.border = thin_border
	# 		if col in orange_columns:
	# 			cell.fill = orange_fill

	# 	# Save workbook
	# 	output = io.BytesIO()
	# 	workbook.save(output)
	# 	output.seek(0)

	# 	filename = f"COST_Sheet_{sales_order}.xlsx"

	# 	frappe.response["filename"] = filename
	# 	frappe.response["filecontent"] = output.getvalue()
	# 	frappe.response["type"] = "download"
  
  
  
  
  
  
	if customer_name.strip() == "Titan Company Limited":
		import io
		from openpyxl import Workbook
		from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

		workbook = Workbook()
		sheet = workbook.active
		sheet.title = "Proto Sheet"

		headers = [
			"SI No.","Prod Code/Dia Code","Ven Ref","Gold","Plating","Finsihing",
			"KT","Gross Wt","Net Wt",
			"Gold Rate","Gold Value",
			"Size Only","Stone Qty","Stone Cts",
			"Wt Per Stone","Stone type","StoneQuality",
			"Stone Shape","Cut/Cab","Setting Type","Stone Rate / cts",
			"Stone value","Handling Chg ","Stone Broken","Labour Val @400/gm",
			"Setting Rate","Setting val","Rhodium",
			"Finishing","Gold loss @3%","Total Labour","Total Value","LINE ID",
		]

		sheet.append(headers)

		# Styles
		blue_fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")
		yellow_fill = PatternFill(start_color="FFF200", end_color="FFF200", fill_type="solid")
		orange_fill = PatternFill(start_color="F4B183", end_color="F4B183", fill_type="solid")

		header_font = Font(bold=True, color="000000")
		center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

		thin_border = Border(
			left=Side(style='thin'),
			right=Side(style='thin'),
			top=Side(style='thin'),
			bottom=Side(style='thin')
		)

		for cell in sheet[1]:
			cell.font = header_font
			cell.fill = blue_fill
			cell.alignment = center_alignment
			cell.border = thin_border

		yellow_columns = [19, 20, 21, 22]
		for col in yellow_columns:
			sheet.cell(row=1, column=col).fill = yellow_fill

		orange_columns = [26, 27, 28, 29, 30]
		for col in orange_columns:
			sheet.cell(row=1, column=col).fill = orange_fill

		sheet.row_dimensions[1].height = 40

		for row in so_doc.items:

			# metal_type = frappe.db.get_value('BOM', row.get('bom'), 'metal_type')
			metal_touch = frappe.db.get_value('BOM', row.get('bom'), 'metal_touch')
			metal_color = frappe.db.get_value('BOM', row.get('bom'), 'metal_colour')
			gross_weight = frappe.db.get_value('BOM', row.get('bom'), 'gross_weight')
			net_weight = frappe.db.get_value('BOM', row.get('bom'), 'metal_and_finding_weight')
			item_sub_category = frappe.db.get_value('BOM', row.get('bom'), 'item_subcategory')
			setting_type = frappe.db.get_value('BOM', row.get('bom'), 'setting_type')
			making_charge = frappe.db.get_value('BOM', row.get('bom'), 'making_charge')

			digit_14code = frappe.db.get_value(
				"Item Theme Code Detail",
				{
					'customer': so_doc.customer,
					'parent': row.item_code
				},
				'digit_14code'
			)
   
			making_charge = frappe.db.get_value("Making Charge Price",{
				"customer": so_doc.customer,
				"metal_touch":metal_touch,
				"setting_type":setting_type,
				"mfg_complexity_code":row.custom_mfg_complexity_code,
			},['name'])

			making_charge_rate = None
			making_charge_wastage = None
			if making_charge:
				making_charge_rate = frappe.db.get_value("Making Charge Price Item Subcategory",{
					"parent":making_charge,
					"subcategory":item_sub_category
				},['rate_per_gm'])
				making_charge_wastage = frappe.db.get_value("Making Charge Price Item Subcategory",{
					"parent":making_charge,
					"subcategory":item_sub_category
				},["wastage"])

			diamond_pcs = frappe.db.get_all(
				"BOM Diamond Detail",
				filters={"parent": row.get("bom")},
				fields=["*"]
			)
			gemstone = frappe.db.get_all(
				"BOM Gemstone Detail",
				filters={"parent": row.get("bom")},
				fields=["*"]
			)
			metal = frappe.db.get_all(
				"BOM Metal Detail",
				filters={"parent": row.get("bom")},
				fields=["*"]
			)
			row_data = [
				"",
				"",
				"",
				"",
				"",
				"",
				"",
				"",
				"",
				"",
				"",
				"",
				"",
				"",
				"",
				"",
				"",
				"",
				"",
				"",
				"",
				"",
				"",
				"",
				"",
				"",
				"",
				"",
				"",
				"",
				"",
				"",
				"",
				"",
				"",
				"",
			]

			sheet.append(row_data)
			last_row = sheet.max_row
			for col in range(1, len(headers) + 1):
				cell = sheet.cell(row=last_row, column=col)
				cell.alignment = center_alignment
				cell.border = thin_border
				if col in orange_columns:
					cell.fill = orange_fill
			total_stone_qty = 0
			total_stone_cts = 0
			total_handling = 0
			total_amount = 0
			total_gold_rate = 0
			total_net_weight = 0


			for pcs in diamond_pcs:
				stone_code = frappe.db.sql("""
					SELECT rm.customer_code
					FROM `tabCustomer RM Code Detail` rm
					LEFT JOIN `tabCustomer RM Code` rmc ON rm.parent = rmc.name
					WHERE 
						IFNULL(rm.stone_shape, '') = IFNULL(%s, '')
						AND rmc.customer = %s
						AND IFNULL(rm.diamond_type, '') = IFNULL(%s, '')
						AND IFNULL(rm.diamond_quality, '') = IFNULL(%s, '')
						AND CAST(IFNULL(rm.size_in_mm, 0) AS DECIMAL(10,3)) = CAST(%s AS DECIMAL(10,3))
							
					""",
					(
						pcs.get("stone_shape"),
						so_doc.customer,
						pcs.get("diamond_type"),
						row.get('diamond_quality'),
						pcs.get("size_in_mm"),
					),
					as_dict=1)
				diamond_rate = frappe.db.get_all("Diamond Price List",filters={
					"diamond_quality":row.get('diamond_quality'),
					"size_in_mm":pcs.get('size_in_mm' or ''),
					"stone_code":stone_code[0].get('customer_code') if stone_code else "",
					"stone_shape":pcs.get('stone_shape' or ''),
					"customer":so_doc.customer
				},fields=['outright_handling_charges_rate','rate'])
    
				handling_rate = diamond_rate[0].get('outright_handling_charges_rate') * pcs.get('quantity' or 0) if diamond_rate[0].get('outright_handling_charges_rate') else 0
				diamond_rt = diamond_rate[0].get('rate') * pcs.get('quantity' or 0) if diamond_rate[0].get('rate') else 0

				total_stone_qty += pcs.get('pcs') or 0
				total_stone_cts += pcs.get('quantity') or 0
				total_amount += diamond_rt or 0
				total_handling += handling_rate
    
			for gems in gemstone:
				gemstone_rate = frappe.db.get_all("Gemstone Price List",filters={
					"gemstone_quality":gems.get('gemstone_quality'),
					"gemstone_size":gems.get('gemstone_size' or ''),
					"stone_shape":gems.get('stone_shape' or ''),
					"customer":so_doc.customer,
					"cut_or_cab": gems.get('cut_or_cab')
				},fields=['outright_handling_charges_rate','rate'])
    
				total_stone_qty += gems.get('pcs' or '') or 0		
				total_stone_cts += gems.get('quantity') or 0
				total_amount += gems.get('pcs') * gemstone_rate[0].get('rate') if gemstone_rate else 0
				total_handling += gems.get('pcs') * gemstone_rate[0].get('outright_handling_charges_rate') if gemstone_rate else 0

			total = (
				((net_weight * making_charge_wastage) /100 if making_charge_rate else 0)
				+ (net_weight * making_charge_rate if making_charge_rate else 0)
				+ total_handling
			)
			total_net_weight +=	net_weight if net_weight else 0
			total_gold_rate += net_weight * so_doc.gold_rate if net_weight else 0
			row_data = [
				"",
				digit_14code if digit_14code else "",
				row.get('serial_no' or ''),
				metal_color,
				"",
				"",
				metal_touch,
				gross_weight if gross_weight else "",
				net_weight if net_weight else "",
				so_doc.gold_rate or '',
				net_weight * so_doc.gold_rate if net_weight else "",
				"",
				total_stone_qty,
				total_stone_cts,
				"",
				"",
				"",
				"",
				"",
				"",
				"",
				total_amount,
				total_handling,
				"",
				(flt(net_weight) * flt(making_charge_rate) if making_charge_rate else ''),
				"",
				"",
				"",
				"",
				(net_weight * making_charge_wastage) /100 if making_charge_rate else '',
				total,
			]
			sheet.append(row_data)
			last_row = sheet.max_row
			for col in range(1, len(headers) + 1):
				cell = sheet.cell(row=last_row, column=col)
				cell.alignment = center_alignment
				cell.border = thin_border
				if col in orange_columns:
					cell.fill = orange_fill
			for pcs in diamond_pcs:
				stone_code = frappe.db.sql("""
					SELECT rm.customer_code
					FROM `tabCustomer RM Code Detail` rm
					LEFT JOIN `tabCustomer RM Code` rmc ON rm.parent = rmc.name
					WHERE 
						IFNULL(rm.stone_shape, '') = IFNULL(%s, '')
						AND rmc.customer = %s
						AND IFNULL(rm.diamond_type, '') = IFNULL(%s, '')
						AND IFNULL(rm.diamond_quality, '') = IFNULL(%s, '')
						AND CAST(IFNULL(rm.size_in_mm, 0) AS DECIMAL(10,3)) = CAST(%s AS DECIMAL(10,3))
							
					""",
					(
						pcs.get("stone_shape"),
						so_doc.customer,
						pcs.get("diamond_type"),
						row.get('diamond_quality'),
						pcs.get("size_in_mm"),
					),
					as_dict=1)
				diamond_rate = frappe.db.get_all("Diamond Price List",filters={
					"diamond_quality":row.get('diamond_quality'),
					"size_in_mm":pcs.get('size_in_mm' or ''),
					"stone_code":stone_code[0].get('customer_code') if stone_code else "",
					"stone_shape":pcs.get('stone_shape' or ''),
					"customer":so_doc.customer
				},fields=['outright_handling_charges_rate','rate'])
				row_data =[
					"",
					stone_code[0].get('customer_code') if stone_code else "",
					"",
					"",
					"",
					"",
					"",
					"",
					"",
					"",
					"",
					pcs.get('size_in_mm' or ''),
					pcs.get('pcs' or ''),
					pcs.get('quantity' or ''),
					pcs.get('weight_per_pcs' or ''),
					"Diamond",
					row.get('diamond_quality' or ''),
					pcs.get('stone_shape' or ''),
					"Cut",
					setting_type,
					diamond_rate[0].get('rate') if diamond_rate else'',
					# pcs.get('total_diamond_rate' or ''),
					# pcs.get('diamond_rate_for_specified_quantity' or ''),
     				diamond_rate[0].get('rate') * pcs.get('quantity' or 0) if diamond_rate[0].get('rate') else '',
					diamond_rate[0].get('outright_handling_charges_rate') * pcs.get('quantity' or 0) if diamond_rate[0].get('outright_handling_charges_rate') else '',
				]
				sheet.append(row_data)
				last_row = sheet.max_row
				for col in range(1, len(headers) + 1):
					cell = sheet.cell(row=last_row, column=col)
					cell.alignment = center_alignment
					cell.border = thin_border
					if col in orange_columns:
						cell.fill = orange_fill
			for gems in gemstone:
				stone_code = frappe.db.sql("""
					SELECT rm.customer_code
					FROM `tabCustomer RM Code Detail` rm
					LEFT JOIN `tabCustomer RM Code` rmc ON rm.parent = rmc.name
					WHERE 
						IFNULL(rm.stone_shape, '') = IFNULL(%s, '')
						AND rmc.customer = %s
						AND IFNULL(rm.gemstone_type, '') = IFNULL(%s, '')
						AND IFNULL(rm.gemstone_size, '') = IFNULL(%s, '')
							
					""",
					(
						gems.get("stone_shape"),
						so_doc.customer,
						gems.get("gemstone_type"),
						gems.get('gemstone_size'),
					),
					as_dict=1)
				gemstone_rate = frappe.db.get_all("Gemstone Price List",filters={
					"gemstone_quality":gems.get('gemstone_quality'),
					"gemstone_size":gems.get('gemstone_size' or ''),
					"stone_shape":gems.get('stone_shape' or ''),
					"customer":so_doc.customer,
					"cut_or_cab": gems.get('cut_or_cab')
				},fields=['outright_handling_charges_rate','rate'])
				row_data =[
					"",
					stone_code[0].get('customer_code') if stone_code else "",
					"",
					"",
					"",
					"",
					"",
					"",
					"",
					"",
					"",
					gems.get('gemstone_size' or ''),
					gems.get('pcs' or ''),
					gems.get('quantity' or ''),
					"",
					"Stone",
					gems.get('gemstone_type' or ''),
					gems.get('stone_shape' or ''),
					"Cut" if gems.get('cut_or_cab') == "Faceted" else 'Cab',
					setting_type,
					gemstone_rate[0].get('rate') if gemstone_rate else "",
					gems.get('pcs' or 0) * gemstone_rate[0].get('rate') if gemstone_rate else "",
					gems.get('pcs' or 0) * gemstone_rate[0].get('outright_handling_charges_rate') if gemstone_rate else "",
				]
				sheet.append(row_data)
				last_row = sheet.max_row
				for col in range(1, len(headers) + 1):
					cell = sheet.cell(row=last_row, column=col)
					cell.alignment = center_alignment
					cell.border = thin_border
					if col in orange_columns:
						cell.fill = orange_fill


		# Add blank row at end
			row_data = [
						"Grand Total",
						"",
						"",
						"",
						"",
						"",
						"",
						"",
						total_net_weight,
						"",
						total_gold_rate,
						"",
						total_stone_qty,
						total_stone_cts,
						"",
						"",
						"",
						"",
						"",
						"",
						"",
						total_amount,
						total_handling,
						"",
						net_weight * metal[0].get('making_rate') if metal else '',
						"",
						"",
						"",
						"",
						net_weight * metal[0].get('wastage_rate') if metal else '',
						total,
						"",
						"",
						"",
						"",
						"",
					]

			sheet.append(row_data)

		last_row = sheet.max_row
		for col in range(1, len(headers) + 1):
			cell = sheet.cell(row=last_row, column=col)
			cell.alignment = center_alignment
			cell.border = thin_border
			if col in orange_columns:
				cell.fill = orange_fill

		# Save workbook
		output = io.BytesIO()
		workbook.save(output)
		output.seek(0)

		filename = f"COST_Sheet_{sales_order}.xlsx"

		frappe.response["filename"] = filename
		frappe.response["filecontent"] = output.getvalue()
		frappe.response["type"] = "download"